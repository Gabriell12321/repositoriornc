#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inspecionar estrutura da planilha Cliente.xlsx
"""

from openpyxl import load_workbook
import sys

try:
    # Ler a planilha
    file_path = r'DADOS PUXAR RNC\Cliente.xlsx'
    print(f"📂 Abrindo planilha: {file_path}\n")
    
    # Carregar workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    print("=" * 80)
    print("📊 ESTRUTURA DA PLANILHA CLIENTE")
    print("=" * 80)
    
    # Ler cabeçalho (primeira linha)
    header = []
    for cell in ws[1]:
        header.append(cell.value)
    
    print(f"\n📋 Total de colunas: {len(header)}")
    print(f"📋 Total de linhas (aprox): {ws.max_row}")
    
    print("\n📝 Nomes das colunas:")
    for i, col in enumerate(header, 1):
        print(f"  {i}. {col}")
    
    print("\n📄 Primeiras 10 linhas de dados:")
    print("-" * 80)
    
    # Mostrar primeiras linhas
    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=11, values_only=True):
        row_count += 1
        print(f"\nLinha {row_count}:")
        for i, (col_name, value) in enumerate(zip(header, row), 1):
            if value is not None and str(value).strip():
                print(f"  {col_name}: {value}")
    
    # Contar total real de linhas com dados
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row if cell is not None):
            total_rows += 1
    
    print("\n" + "=" * 80)
    print(f"✅ Total de registros com dados: {total_rows}")
    print("✅ Inspeção concluída!")
    
    wb.close()
    
except FileNotFoundError:
    print(f"❌ Arquivo não encontrado: {file_path}")
    sys.exit(1)
except Exception as e:
    print(f"❌ Erro ao ler planilha: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
