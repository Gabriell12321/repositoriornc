#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SCRIPT DE IMPORTAÇÃO - CADASTRO DE CLIENTES
Importa dados da planilha Cliente.xlsx para o banco IPPEL
"""

from openpyxl import load_workbook
import sqlite3
from datetime import datetime
import sys

def clean_client_name(name):
    """Limpa e normaliza nome do cliente"""
    if not name:
        return None
    
    name = str(name).strip()
    
    # Remover espaços duplos
    while '  ' in name:
        name = name.replace('  ', ' ')
    
    return name if name else None

def import_clients(dry_run=True):
    """
    Importa clientes da planilha para o banco
    
    Args:
        dry_run: Se True, apenas simula a importação sem gravar no banco
    """
    
    file_path = r'DADOS PUXAR RNC\Cliente.xlsx'
    db_path = 'ippel_system.db'
    
    print("=" * 80)
    if dry_run:
        print("🔍 MODO DRY-RUN - Simulação de importação (sem gravar no banco)")
    else:
        print("💾 MODO REAL - Importando clientes para o banco de dados")
    print("=" * 80)
    
    try:
        # Conectar ao banco
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Verificar se tabela clientes existe, senão criar
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE,
                ativo INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        if not dry_run:
            conn.commit()
            print("✅ Tabela 'clientes' verificada/criada\n")
        
        # Carregar planilha
        print(f"📂 Abrindo planilha: {file_path}")
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Estatísticas
        total_planilha = 0
        novos_clientes = 0
        existentes = 0
        invalidos = 0
        clientes_lista = []
        
        # Ler dados da planilha (pular cabeçalho)
        print("📊 Processando dados...\n")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_planilha += 1
            
            cliente_nome = clean_client_name(row[0])
            
            if not cliente_nome:
                invalidos += 1
                continue
            
            # Verificar se cliente já existe no banco
            cursor.execute("SELECT id, nome FROM clientes WHERE nome = ?", (cliente_nome,))
            existing = cursor.fetchone()
            
            if existing:
                existentes += 1
                if dry_run:
                    print(f"  ⏭️  JÁ EXISTE: {cliente_nome}")
            else:
                novos_clientes += 1
                clientes_lista.append(cliente_nome)
                
                if dry_run:
                    print(f"  ✨ NOVO: {cliente_nome}")
                else:
                    # Inserir no banco
                    cursor.execute("""
                        INSERT INTO clientes (nome, ativo, created_at, updated_at)
                        VALUES (?, 1, ?, ?)
                    """, (cliente_nome, datetime.now().isoformat(), datetime.now().isoformat()))
        
        wb.close()
        
        # Commit se não for dry-run
        if not dry_run:
            conn.commit()
        
        # Estatísticas finais
        print("\n" + "=" * 80)
        print("📊 RESULTADO DA IMPORTAÇÃO")
        print("=" * 80)
        print(f"📋 Total na planilha: {total_planilha}")
        print(f"✨ Novos clientes: {novos_clientes}")
        print(f"⏭️  Já existentes: {existentes}")
        print(f"⚠️  Inválidos/vazios: {invalidos}")
        
        if not dry_run:
            # Verificar total no banco
            cursor.execute("SELECT COUNT(*) FROM clientes WHERE ativo = 1")
            total_db = cursor.fetchone()[0]
            print(f"\n📈 Total de clientes ativos no banco: {total_db}")
            
            print("\n✅ IMPORTAÇÃO CONCLUÍDA COM SUCESSO!")
        else:
            print("\n💡 Execute com '--apply' para aplicar as mudanças no banco")
        
        print("=" * 80)
        
        conn.close()
        
        return {
            'total_planilha': total_planilha,
            'novos': novos_clientes,
            'existentes': existentes,
            'invalidos': invalidos
        }
        
    except FileNotFoundError:
        print(f"❌ Arquivo não encontrado: {file_path}")
        sys.exit(1)
    except sqlite3.Error as e:
        print(f"❌ Erro no banco de dados: {e}")
        if not dry_run:
            conn.rollback()
        sys.exit(1)
    except Exception as e:
        print(f"❌ Erro durante importação: {e}")
        import traceback
        traceback.print_exc()
        if not dry_run:
            conn.rollback()
        sys.exit(1)

if __name__ == '__main__':
    # Verificar argumentos
    dry_run = True
    
    if len(sys.argv) > 1 and sys.argv[1] in ['--apply', '-a', '--real']:
        dry_run = False
        
        print("\n⚠️  ATENÇÃO: Você está prestes a importar dados para o banco!")
        print("⚠️  Esta operação irá adicionar novos clientes.")
        response = input("\n❓ Deseja continuar? (s/n): ")
        
        if response.lower() not in ['s', 'sim', 'y', 'yes']:
            print("❌ Operação cancelada pelo usuário.")
            sys.exit(0)
        
        print("\n🚀 Iniciando importação...\n")
    
    # Executar importação
    import_clients(dry_run=dry_run)
