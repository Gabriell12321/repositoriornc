 #!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sqlite3
import time
# Nota de manutenção: Em 2025-10-03 houve falha de inicialização
# (NameError: HAS_COMPRESS) porque a flag era usada antes de garantir definição.
# Bloco de import de flask_compress agora define HAS_COMPRESS de forma defensiva
# antes de qualquer uso subsequente.
import hashlib
from flask import Flask, request, jsonify, send_from_directory, session, redirect, url_for, render_template, abort, make_response, g
import socket
from services.db import DB_PATH, get_db_connection, return_db_connection, warm_pool
from services.cache import cache_query, get_cached_query, clear_expired_cache, clear_rnc_cache
from services.permissions import has_permission, has_department_permission, get_user_department
from services.groups import (
    get_all_groups,
    get_group_by_id,
    create_group,
    update_group,
    delete_group,
    get_users_by_group,
)
from services.rnc import (
    share_rnc_with_user,
    get_rnc_shared_users,
    can_user_access_rnc,
)

# Importações opcionais com fallback - otimizadas para startup mais rápido
# Garantir que flags booleanas são sempre definidas antes de qualquer uso
HAS_COMPRESS = False
# Importação lazy - só importa quando for usar compressão
Compress = None

HAS_TALISMAN = False
# Importação lazy - só importa quando for configurar segurança
Talisman = None

from werkzeug.middleware.proxy_fix import ProxyFix

# Flask-Limiter desabilitado para evitar problemas de configuração
HAS_LIMITER = False
def get_remote_address():
    return request.remote_addr if request else '127.0.0.1'

import re
import secrets
import sys
import os

# Adicionar o diretório utils ao path para importar formatação
utils_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'utils')
if utils_path not in sys.path:
    sys.path.append(utils_path)

# Importar funções de formatação
try:
    from utils.formatting import format_currency, format_number, format_percentage, format_data_for_dashboard, format_table_data
    HAS_FORMATTING = True
except ImportError:
    HAS_FORMATTING = False
    print(" Módulo de formatação não encontrado - valores não serão formatados")
    # Funções de fallback para formatação
    def format_currency(value):
        try:
            num = float(value) if value is not None else 0
            return f"R$ {num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "R$ 0,00"
    
    def format_number(value):
        try:
            num = float(value) if value is not None else 0
            return f"{num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "0,00"
    
    def format_percentage(value):
        try:
            num = float(value) if value is not None else 0
            return f"{num:.1f}%"
        except:
            return "0,0%"
    
    def format_data_for_dashboard(data):
        return data  # Retorna dados sem formatação no fallback
    
    def format_table_data(data):
        return data  # Retorna dados sem formatação no fallback
import queue
import time
import gc
import json
import threading
import tracemalloc
import datetime
import logging
import os
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timedelta
from routes.api import api as api_bp
from routes.auth import auth as auth_bp
from routes.rnc import rnc as rnc_bp
from routes.print_reports import print_reports as print_reports_bp
from routes.field_locks import field_locks_bp

try:
    from flask_socketio import SocketIO, emit, join_room, leave_room  # type: ignore
    HAS_SOCKETIO = True
except ImportError:  # Fallback dummy para executar sem dependência
    HAS_SOCKETIO = False
    class _DummySocketIO:
        def __init__(self, app, *args, **kwargs):
            self.app = app
        def on(self, *args, **kwargs):
            def _decorator(f):
                return f
            return _decorator
        def emit(self, *args, **kwargs):
            pass
        def run(self, app, host='0.0.0.0', port=5000, **kwargs):
            print(" flask_socketio não instalado - usando Flask padrão (sem WebSocket)")
            app.run(host=host, port=port, debug=kwargs.get('debug', False))
    def emit(*args, **kwargs):
        pass
    def join_room(*args, **kwargs):
        pass
    def leave_room(*args, **kwargs):
        pass
    SocketIO = _DummySocketIO  # type: ignore

from functools import wraps
from concurrent.futures import ThreadPoolExecutor

# psutil é opcional; se der conflito de versão, seguimos sem ele
try:
    import psutil  # type: ignore
    HAS_PSUTIL = True
except Exception:
    psutil = None  # type: ignore
    HAS_PSUTIL = False

try:
    tracemalloc.start()
except Exception:
    pass

app = Flask(__name__)

# Forçar recarga de templates
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.auto_reload = True

# Rate limiting global (se lib disponível)
try:
    import importlib
    _rl = importlib.import_module('services.rate_limit')
    _init_limiter = getattr(_rl, 'init_limiter', None)
    _defaults = os.environ.get('RATE_LIMIT_DEFAULTS', '200 per minute').split(';')
    _defaults = [d.strip() for d in _defaults if d.strip()]
    _limiter = _init_limiter(app, default_limits=_defaults) if _init_limiter else None
    HAS_LIMITER = _limiter is not None
    
    # Isentar endpoint de notificações do rate limit (polling frequente)
    if _limiter is not None:
        try:
            _limiter.exempt('/api/notifications/unread')
        except Exception as e:
            print(f" Não foi possível isentar /api/notifications/unread do rate limit: {e}")
except Exception:
    HAS_LIMITER = False

# Config de compressão (gzip/brotli se disponível) - lazy loading


def setup_compression():
    global HAS_COMPRESS, Compress
    if not HAS_COMPRESS:
        try:
            from flask_compress import Compress  # type: ignore
            HAS_COMPRESS = True
        except ImportError:
            print(" Flask-Compress não instalado - compressão desabilitada")
            return
    
if HAS_COMPRESS and Compress is not None:
    try:
        app.config.update(
            COMPRESS_MIMETYPES=[
                'text/html','text/css','text/xml','application/json',
                'application/javascript','text/javascript','image/svg+xml'
            ],
            COMPRESS_LEVEL=6,
            COMPRESS_MIN_SIZE=1024
        )
        try:  # Brotli opcional
            import brotli  # type: ignore  # noqa: F401
            app.config['COMPRESS_ALGORITHM'] = 'br'
            app.config['COMPRESS_BR_LEVEL'] = 5
        except Exception:
            app.config['COMPRESS_ALGORITHM'] = 'gzip'
        
        Compress(app)
    except Exception as _compress_err:
        print(f" Falha ao inicializar compressão: {_compress_err}")
        HAS_COMPRESS = False

# Compressão será configurada em background para não atrasar startup
# Thread será iniciado após inicialização completa da aplicação

app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Secret key estável: lê de variável de ambiente ou de arquivo local persistente
_secret_from_env = os.environ.get('IPPEL_SECRET_KEY')
if _secret_from_env:
    app.secret_key = _secret_from_env
else:
    _secret_file = os.path.join(os.path.dirname(__file__), 'ippel_secret.key')
    try:
        if os.path.exists(_secret_file):
            with open(_secret_file, 'r', encoding='utf-8') as f:
                app.secret_key = f.read().strip()
        else:
            _new_secret = secrets.token_hex(32)
            with open(_secret_file, 'w', encoding='utf-8') as f:
                f.write(_new_secret)
            app.secret_key = _new_secret
    except Exception:
        app.secret_key = secrets.token_hex(32)

# Configurações essenciais
app.config['JSON_SORT_KEYS'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['REMEMBER_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_NAME'] = 'ippel_session'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB por requisição

# Preferir assets minificados em produção
app.config['USE_MIN_ASSETS'] = os.environ.get('USE_MIN_ASSETS', '1') not in ('0', 'false', 'False')

def asset_url(filename: str) -> str:
    """
    Retorna URL para static, preferindo arquivo .min.* quando configurado e existente.
    Também adiciona um query param de versão baseado em mtime para cache busting.
    """
    try:
        use_min = bool(app.config.get('USE_MIN_ASSETS'))
        chosen = filename
        full_dir = app.static_folder or os.path.join(os.path.dirname(__file__), 'static')
        if use_min and '.min.' not in filename:
            parts = filename.rsplit('.', 1)
            if len(parts) == 2:
                candidate = f"{parts[0]}.min.{parts[1]}"
                full_path = os.path.join(full_dir, candidate)
                if os.path.exists(full_path):
                    chosen = candidate
        # Cache bust por mtime
        try:
            full_path_final = os.path.join(full_dir, chosen)
            v = str(int(os.path.getmtime(full_path_final))) if os.path.exists(full_path_final) else '1'
        except Exception:
            v = '1'
        return url_for('static', filename=chosen) + f"?v={v}"
    except Exception:
        # Fallback simples
        return url_for('static', filename=filename)

# Disponibiliza helper no Jinja
app.jinja_env.globals['asset_url'] = asset_url

# Security headers e CSP básicos (aplicados se Flask-Talisman estiver disponível)
try:
    if HAS_TALISMAN and Talisman is not None:
        csp = {
            'default-src': ["'self'"],
            'base-uri': ["'self'"],
            'frame-ancestors': ["'self'"],
            'object-src': ["'none'"],
            'form-action': ["'self'"],
            'img-src': ["'self'", 'data:', 'blob:', 'https://api.dicebear.com'],
            'style-src': ["'self'", "'unsafe-inline'"],  # manter inline até extrair estilos
            'font-src': ["'self'", 'data:'],
            'script-src': [
                "'self'",
                # Temporário: ainda há scripts inline e handlers; manter até migração
                "'unsafe-inline'",
                "'unsafe-eval'",  # Necessário para Socket.IO
                'https://cdn.jsdelivr.net',
                'https://cdnjs.cloudflare.com',
            ],
            'connect-src': ["'self'", 'ws:', 'wss:', 'https://cdnjs.cloudflare.com', 'https://cdn.jsdelivr.net'],
            'manifest-src': ["'self'"],
        }
        # Importante: não usar nonces enquanto ainda existem event handlers inline,
        # pois a presença de nonce invalida 'unsafe-inline' para handlers (CSP spec)
        Talisman(
            app,
            force_https=False,
            content_security_policy=csp,
            # content_security_policy_nonce_in=['script-src'],  # desabilitado temporariamente
            session_cookie_secure=False,
            frame_options='SAMEORIGIN',
            referrer_policy='no-referrer',
            feature_policy=None,
        )
        # Adiciona um header "report-only" mais estrito para mapear violações sem quebrar páginas
        @app.after_request
        def add_csp_report_only(resp):
            try:
                ro_directives = {
                    'default-src': "'self'",
                    'base-uri': "'self'",
                    'frame-ancestors': "'self'",
                    'object-src': "'none'",
                    'form-action': "'self'",
                    'img-src': "'self' data: blob: https://api.dicebear.com",
                    # temporário: permitir inline no report-only para reduzir ruído enquanto migramos
                                    'style-src': "'self' 'unsafe-inline' https://fonts.googleapis.com",
                'font-src': "'self' data: https://fonts.gstatic.com",
                    'script-src': "'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com",
                    'connect-src': "'self'",
                    'manifest-src': "'self'",
                    'report-uri': "/csp-report",
                }
                policy = '; '.join([f"{k} {v}" for k, v in ro_directives.items()])
                resp.headers['Content-Security-Policy-Report-Only'] = policy
            except Exception:
                pass
            return resp
        
        # Endpoint para receber relatórios CSP (report-uri)
        @app.post('/csp-report')
        def csp_report():
            try:
                payload = None
                ct = request.headers.get('Content-Type', '')
                if 'application/json' in ct or 'csp-report' in ct or 'reports+json' in ct:
                    payload = request.get_json(silent=True)
                if payload is None:
                    payload = {'raw': request.data.decode('utf-8', errors='ignore')}
            except Exception:
                payload = {'raw': ''}
            try:
                import importlib
                _sl = importlib.import_module('services.security_log')
                sec_log = getattr(_sl, 'sec_log', None)
                if sec_log:
                    sec_log('csp', 'violation_report', status='report-only', details=payload)
            except Exception:
                pass
            return ('', 204)
except Exception:
    pass

# Fallback: se Talisman não estiver disponível, aplica CSP mínima e report-only manualmente
if not HAS_TALISMAN:
    @app.after_request
    def add_basic_csp(resp):
        try:
            policy = " ".join([
                "default-src 'self';",
                "base-uri 'self';",
                "frame-ancestors 'self';",
                "object-src 'none';",
                "form-action 'self';",
                "img-src 'self' data: blob: https://api.dicebear.com;",
                "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com;",
                "font-src 'self' data: https://fonts.gstatic.com;",
                "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com;",
                "connect-src 'self' https://cdnjs.cloudflare.com https://cdn.jsdelivr.net;",
                "manifest-src 'self';",
            ])
            resp.headers['Content-Security-Policy'] = policy

            # Report-Only estrito
            report_only = " ".join([
                "default-src 'self';",
                "base-uri 'self';",
                "frame-ancestors 'self';",
                "object-src 'none';",
                "form-action 'self';",
                "img-src 'self' data: blob: https://api.dicebear.com;",
                "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com;",
                "font-src 'self' data: https://fonts.gstatic.com;",
                "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com;",
                "connect-src 'self' https://cdnjs.cloudflare.com https://cdn.jsdelivr.net;",
                "manifest-src 'self';",
                "report-uri /csp-report;",
            ])
            resp.headers['Content-Security-Policy-Report-Only'] = report_only
        except Exception:
            pass
        return resp

    @app.post('/csp-report')
    def csp_report_fallback():
        try:
            payload = None
            ct = request.headers.get('Content-Type', '')
            if 'application/json' in ct or 'csp-report' in ct or 'reports+json' in ct:
                payload = request.get_json(silent=True)
            if payload is None:
                payload = {'raw': request.data.decode('utf-8', errors='ignore')}
        except Exception:
            payload = {'raw': ''}
        try:
            import importlib
            _sl = importlib.import_module('services.security_log')
            sec_log = getattr(_sl, 'sec_log', None)
            if sec_log:
                sec_log('csp', 'violation_report', status='report-only', details=payload)
        except Exception:
            pass
        return ('', 204)

# Security logger inicial
try:
    import importlib
    _sl = importlib.import_module('services.security_log')
    _setup_sec = getattr(_sl, 'setup_security_logger', None)
    if _setup_sec:
        _setup_sec(app)
except Exception:
    pass

# Rota para servir logo(s) existentes na raiz do projeto (compatibilidade com templates)
@app.route('/LOGOIPPEL.JPEG')
def _serve_logo_legacy_upper():
    try:
        root_dir = os.path.dirname(os.path.abspath(__file__))
        return send_from_directory(root_dir, 'IPPELLOGO.jpg')
    except Exception:
        abort(404)

@app.route('/IPPELLOGO.jpg')
def _serve_logo_legacy():
    try:
        root_dir = os.path.dirname(os.path.abspath(__file__))
        return send_from_directory(root_dir, 'IPPELLOGO.jpg')
    except Exception:
        abort(404)

# SocketIO com eventlet para melhor performance
try:
    # Verificar se SSL está ativo
    import os as _ssl_check
    ssl_cert_path = _ssl_check.path.join(_ssl_check.path.dirname(__file__), 'ssl_certs', 'cert.pem')
    ssl_key_path = _ssl_check.path.join(_ssl_check.path.dirname(__file__), 'ssl_certs', 'key.pem')
    use_https = _ssl_check.path.exists(ssl_cert_path) and _ssl_check.path.exists(ssl_key_path)
    
    socketio = SocketIO(app, 
                       async_mode='eventlet',
                       cors_allowed_origins="*",
                       logger=True,
                       engineio_logger=True,
                       ping_timeout=60,
                       ping_interval=25,
                       # Configurações específicas para HTTPS
                       path='/socket.io/',
                       transports=['polling', 'websocket'],
                       # Configurações de transporte seguro
                       always_connect=False,
                       # Timeout específico para HTTPS
                       http_compression=True,
                       compression_threshold=1024,
                       # Headers CORS específicos para HTTPS
                       cors_credentials=True if use_https else False,
                       # Configuração de polling específica para HTTPS
                       json=None,  # Usar JSON padrão
                       # Configurações específicas para SSL
                       allow_unsafe_werkzeug=True)
    print(f" SocketIO inicializado com eventlet (HTTPS: {use_https})")
except Exception as e:
    print(f" SocketIO falhou: {e}")
    socketio = None

# Rota de verificação de saúde do Socket.IO
@app.route('/socketio-health')
def socketio_health():
    """Verifica se o Socket.IO está funcionando"""
    if socketio is None:
        return jsonify({
            'status': 'error',
            'message': 'Socket.IO não inicializado'
        }), 500
    
    # Verificar se está rodando com HTTPS
    import os as _check_ssl
    ssl_cert_path = _check_ssl.path.join(_check_ssl.path.dirname(__file__), 'ssl_certs', 'cert.pem')
    ssl_key_path = _check_ssl.path.join(_check_ssl.path.dirname(__file__), 'ssl_certs', 'key.pem')
    use_https = _check_ssl.path.exists(ssl_cert_path) and _check_ssl.path.exists(ssl_key_path)
    
    return jsonify({
        'status': 'ok',
        'async_mode': socketio.async_mode,
        'https_enabled': use_https,
        'protocol': 'wss' if use_https else 'ws',
        'transports': ['polling', 'websocket'],
        'message': f'Socket.IO funcionando corretamente ({("HTTPS" if use_https else "HTTP")})'
    })

@app.route('/socketio-test')
def socketio_test():
    """Página de teste do Socket.IO"""
    return render_template('socketio_test.html')

# Logging básico
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Pool e cache agora são providos por services/*

# Registrar Blueprints
app.register_blueprint(api_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(rnc_bp)
app.register_blueprint(print_reports_bp)
app.register_blueprint(field_locks_bp)

# Registrar Blueprint de Notificações
try:
    from services.notifications_api import notifications_bp
    app.register_blueprint(notifications_bp)
    logger.info(" Blueprint de notificações registrado")
except Exception as e:
    logger.error(f" Erro ao registrar blueprint de notificações: {e}")

# Inicializar Socket.IO para notificações
if socketio is not None:
    try:
        from services.notification_socketio import init_notification_socketio
        init_notification_socketio(socketio)
        logger.info(" Socket.IO de notificações inicializado")
    except Exception as e:
        logger.error(f" Erro ao inicializar Socket.IO de notificações: {e}")

# JWT: parse Authorization Bearer and attach g.user_id if valid
@app.before_request
def _jwt_before_request(): 
    try:
        authz = request.headers.get('Authorization') or ''
        if authz.lower().startswith('bearer '):
            token = authz.split(' ', 1)[1].strip()
            try:
                import importlib
                _jwt = importlib.import_module('services.jwt_auth')
                payload = _jwt.decode_token(token, verify_type=None)
                g.jwt = payload
                g.user_id = int(payload.get('sub')) if payload.get('sub') is not None else None
                
                # CORRIGIDO: Verificar se o usuário está ativo antes de permitir acesso
                if g.user_id:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute('SELECT is_active FROM users WHERE id = ?', (g.user_id,))
                    user_row = cursor.fetchone()
                    conn.close()
                    
                    # Se usuário não existe ou está inativo, limpar dados da sessão
                    if not user_row or not user_row[0]:
                        g.user_id = None
                        g.jwt = None
                        if 'user_id' in session:
                            session.clear()
                        return
                
                # For compatibility with existing permission decorators, mirror into session if absent
                if g.user_id and 'user_id' not in session:
                    session['user_id'] = g.user_id
                    session['user_name'] = payload.get('name')
                    session['user_email'] = payload.get('email')
                    session['user_department'] = payload.get('department')
                    session['user_role'] = payload.get('role')
            except Exception:
                pass
    except Exception:
        pass

# Middleware para verificar se usuário autenticado via sessão está ativo
@app.before_request
def _check_active_user():
    # Verificar apenas se há user_id na sessão (não JWT)
    if 'user_id' in session and not hasattr(g, 'user_id'):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute('SELECT is_active FROM users WHERE id = ?', (session['user_id'],))
            user_row = cursor.fetchone()
            conn.close()
            
            # Se usuário não existe ou está inativo, limpar sessão
            if not user_row or not user_row[0]:
                session.clear()
                return jsonify({'success': False, 'message': 'Usuário desativado'}), 401
        except Exception:
            pass

# =============== Optional Julia Analytics proxy ===============
try:
    from services.analytics_client import get_summary as _get_analytics_summary  # type: ignore
except Exception:
    _get_analytics_summary = None  # type: ignore


@app.get('/api/analytics/summary')
def api_analytics_summary_proxy():
    try:
        if not _get_analytics_summary:
            return jsonify({'success': False, 'message': 'Analytics client indisponível'}), 404
        data = _get_analytics_summary()  # may be None if not configured
        if not data:
            return jsonify({'success': False, 'message': 'Serviço Analytics não configurado'}), 404
        return jsonify({'success': True, 'summary': data})
    except Exception as e:
        try:
            logger.error(f"Erro ao obter resumo analytics: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao consultar analytics'}), 500

# =============== Optional Go Reports proxy ===============
try:
    from services.reports_client import get_rnc_pdf as _get_rnc_pdf  # type: ignore
except Exception:
    _get_rnc_pdf = None  # type: ignore


@app.get('/api/reports/rnc/<int:rnc_id>.pdf')
def api_reports_rnc_pdf(rnc_id: int):
    try:
        if not _get_rnc_pdf:
            return jsonify({'success': False, 'message': 'Reports client indisponível'}), 404
        content = _get_rnc_pdf(rnc_id)
        if not content:
            return jsonify({'success': False, 'message': 'Serviço de relatórios não configurado'}), 404
        resp = make_response(content)
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Cache-Control'] = 'no-store'
        return resp
    except Exception as e:
        try:
            logger.error(f"Erro ao obter PDF RNC {rnc_id}: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao gerar relatório'}), 500

# =============== Optional Kotlin Utils proxy ===============
try:
    from services.kotlin_client import get_qr_png as _get_qr_png  # type: ignore
except Exception:
    _get_qr_png = None  # type: ignore


@app.get('/api/utils/qr.png')
def api_utils_qr_png():
    try:
        if not _get_qr_png:
            return jsonify({'success': False, 'message': 'Kotlin utils client indisponível'}), 404
        text = request.args.get('text')
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        try:
            size = int(request.args.get('size', '256'))
        except Exception:
            size = 256
        content = _get_qr_png(text, size=size)
        if not content:
            return jsonify({'success': False, 'message': 'Serviço Kotlin Utils não configurado'}), 404
        resp = make_response(content)
        resp.headers['Content-Type'] = 'image/png'
        resp.headers['Cache-Control'] = 'no-store'
        return resp
    except Exception as e:
        try:
            logger.error(f"Erro ao gerar QR: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao gerar QR'}), 500

# =============== Optional Swift Tools proxy ===============
try:
    from services.swift_client import sha256 as _swift_sha256  # type: ignore
except Exception:
    _swift_sha256 = None  # type: ignore


@app.post('/api/utils/hash/sha256')
def api_utils_hash_sha256():
    try:
        if not _swift_sha256:
            return jsonify({'success': False, 'message': 'Swift tools client indisponível'}), 404
        text = None
        if request.is_json:
            data = request.get_json(silent=True) or {}
            text = data.get('text')
        if not text:
            text = request.form.get('text') or request.data.decode('utf-8', errors='ignore')
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        digest = _swift_sha256(text)
        if not digest:
            return jsonify({'success': False, 'message': 'Serviço Swift Tools não configurado'}), 404
        return jsonify({'success': True, 'sha256': digest})
    except Exception as e:
        try:
            logger.error(f"Erro ao calcular sha256: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao calcular hash'}), 500

# =============== Optional Scala Tools proxy ===============
try:
    from services.scala_client import b64_encode as _scala_b64_encode, b64_decode as _scala_b64_decode  # type: ignore
except Exception:
    _scala_b64_encode = None  # type: ignore
    _scala_b64_decode = None  # type: ignore


@app.post('/api/utils/b64/encode')
def api_utils_b64_encode():
    try:
        if not _scala_b64_encode:
            return jsonify({'success': False, 'message': 'Scala tools client indisponível'}), 404
        text = request.data.decode('utf-8', errors='ignore')
        if not text and request.is_json:
            text = (request.get_json(silent=True) or {}).get('text') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro texto é obrigatório'}), 400
        out = _scala_b64_encode(text)
        if not out:
            return jsonify({'success': False, 'message': 'Serviço Scala Tools não configurado'}), 404
        return jsonify({'success': True, 'data': out})
    except Exception as e:
        try:
            logger.error(f"Erro b64 encode: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha no encode'}), 500


@app.post('/api/utils/b64/decode')
def api_utils_b64_decode():
    try:
        if not _scala_b64_decode:
            return jsonify({'success': False, 'message': 'Scala tools client indisponível'}), 404
        text = request.data.decode('utf-8', errors='ignore')
        if not text and request.is_json:
            text = (request.get_json(silent=True) or {}).get('data') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro data é obrigatório'}), 400
        out = _scala_b64_decode(text)
        if out is None:
            return jsonify({'success': False, 'message': 'Serviço Scala Tools não configurado ou base64 inválido'}), 400
        return jsonify({'success': True, 'data': out})
    except Exception as e:
        try:
            logger.error(f"Erro b64 decode: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha no decode'}), 500

# =============== RNC Number API ===============
@app.get('/api/rnc/next-number')
def api_rnc_next_number():
    """Retorna o próximo número RNC que será gerado"""
    try:
        from main_system import RNCSystem, DB_PATH
        rnc_system = RNCSystem(DB_PATH)
        next_number = rnc_system.generate_rnc_number()
        return jsonify({'success': True, 'next_number': next_number})
    except Exception as e:
        try:
            logger.error(f"Erro ao obter próximo número RNC: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao gerar próximo número'}), 500

# =============== Optional Nim Tools proxy ===============
try:
    from services.nim_client import get_uuid as _nim_uuid, get_token as _nim_token  # type: ignore
except Exception:
    _nim_uuid = None  # type: ignore
    _nim_token = None  # type: ignore


@app.get('/api/utils/uuid')
def api_utils_uuid():
    try:
        if not _nim_uuid:
            return jsonify({'success': False, 'message': 'Nim tools client indisponível'}), 404
        u = _nim_uuid()
        if not u:
            return jsonify({'success': False, 'message': 'Serviço Nim Tools não configurado'}), 404
        return jsonify({'success': True, 'uuid': u})
    except Exception as e:
        try:
            logger.error(f"Erro get uuid: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao gerar uuid'}), 500


@app.get('/api/utils/token')
def api_utils_token():
    try:
        if not _nim_token:
            return jsonify({'success': False, 'message': 'Nim tools client indisponível'}), 404
        try:
            size = int(request.args.get('size', '32'))
        except Exception:
            size = 32
        t = _nim_token(size=size)
        if not t:
            return jsonify({'success': False, 'message': 'Serviço Nim Tools não configurado'}), 404
        return jsonify({'success': True, 'token': t})
    except Exception as e:
        try:
            logger.error(f"Erro get token: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao gerar token'}), 500

# =============== Optional V Tools proxy ===============
try:
    from services.v_client import slugify as _v_slugify  # type: ignore
except Exception:
    _v_slugify = None  # type: ignore


@app.get('/api/utils/slug')
def api_utils_slug():
    try:
        if not _v_slugify:
            return jsonify({'success': False, 'message': 'V tools client indisponível'}), 404
        text = request.args.get('text') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        s = _v_slugify(text)
        if not s:
            return jsonify({'success': False, 'message': 'Serviço V Tools não configurado'}), 404
        return jsonify({'success': True, 'slug': s})
    except Exception as e:
        try:
            logger.error(f"Erro slug: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao gerar slug'}), 500

# =============== Optional Haskell Tools proxy ===============
try:
    from services.haskell_client import levenshtein as _hs_lev  # type: ignore
except Exception:
    _hs_lev = None  # type: ignore


@app.post('/api/utils/levenshtein')
def api_utils_levenshtein():
    try:
        if not _hs_lev:
            return jsonify({'success': False, 'message': 'Haskell tools client indisponível'}), 404
        a = b = ''
        if request.is_json:
            data = request.get_json(silent=True) or {}
            a = data.get('a') or ''
            b = data.get('b') or ''
        if not a or not b:
            # Try body "a;b"
            raw = request.data.decode('utf-8', errors='ignore')
            parts = raw.split(';')
            if len(parts) >= 2:
                a, b = parts[0], parts[1]
        if not a or not b:
            return jsonify({'success': False, 'message': 'Parâmetros a e b são obrigatórios'}), 400
        dist = _hs_lev(a, b)
        if dist is None:
            return jsonify({'success': False, 'message': 'Serviço Haskell Tools não configurado'}), 404
        return jsonify({'success': True, 'distance': dist})
    except Exception as e:
        try:
            logger.error(f"Erro levenshtein: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao calcular distância'}), 500

# =============== Optional Zig Tools proxy ===============
try:
    from services.zig_client import xxh3 as _zig_xxh3  # type: ignore
except Exception:
    _zig_xxh3 = None  # type: ignore


@app.post('/api/utils/xxh3')
def api_utils_xxh3():
    try:
        if not _zig_xxh3:
            return jsonify({'success': False, 'message': 'Zig tools client indisponível'}), 404
        text = request.data.decode('utf-8', errors='ignore')
        if not text and request.is_json:
            text = (request.get_json(silent=True) or {}).get('text') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        value = _zig_xxh3(text)
        if not value:
            return jsonify({'success': False, 'message': 'Serviço Zig Tools não configurado'}), 404
        return jsonify({'success': True, 'xxh3': value})
    except Exception as e:
        try:
            logger.error(f"Erro xxh3: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao calcular xxh3'}), 500

# =============== Optional Crystal Tools proxy ===============
try:
    from services.crystal_client import sha256 as _cr_sha256  # type: ignore
except Exception:
    _cr_sha256 = None  # type: ignore


@app.post('/api/utils/sha256')
def api_utils_sha256():
    try:
        if not _cr_sha256:
            return jsonify({'success': False, 'message': 'Crystal tools client indisponível'}), 404
        text = request.data.decode('utf-8', errors='ignore')
        if not text and request.is_json:
            text = (request.get_json(silent=True) or {}).get('text') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        digest = _cr_sha256(text)
        if not digest:
            return jsonify({'success': False, 'message': 'Serviço Crystal Tools não configurado'}), 404
        return jsonify({'success': True, 'sha256': digest})
    except Exception as e:
        try:
            logger.error(f"Erro sha256 (crystal): {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha ao calcular sha256'}), 500

# =============== Optional Deno Tools proxy ===============
try:
    from services.deno_client import url_encode as _deno_encode, url_decode as _deno_decode  # type: ignore
except Exception:
    _deno_encode = None  # type: ignore
    _deno_decode = None  # type: ignore


@app.post('/api/utils/url/encode')
def api_utils_url_encode():
    try:
        if not _deno_encode:
            return jsonify({'success': False, 'message': 'Deno tools client indisponível'}), 404
        text = request.data.decode('utf-8', errors='ignore')
        if not text and request.is_json:
            text = (request.get_json(silent=True) or {}).get('text') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        out = _deno_encode(text)
        if out is None:
            return jsonify({'success': False, 'message': 'Serviço Deno Tools não configurado'}), 404
        return jsonify({'success': True, 'data': out})
    except Exception as e:
        try:
            logger.error(f"Erro URL encode: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha no URL encode'}), 500


@app.post('/api/utils/url/decode')
def api_utils_url_decode():
    try:
        if not _deno_decode:
            return jsonify({'success': False, 'message': 'Deno tools client indisponível'}), 404
        text = request.data.decode('utf-8', errors='ignore')
        if not text and request.is_json:
            text = (request.get_json(silent=True) or {}).get('text') or ''
        if not text:
            return jsonify({'success': False, 'message': 'Parâmetro text é obrigatório'}), 400
        out = _deno_decode(text)
        if out is None:
            return jsonify({'success': False, 'message': 'Serviço Deno Tools não configurado ou inválido'}), 404
        return jsonify({'success': True, 'data': out})
    except Exception as e:
        try:
            logger.error(f"Erro URL decode: {e}")
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Falha no URL decode'}), 500

# Rota de debug de sessão
@app.route('/api/debug/session')
def api_debug_session():
    data = {k: session.get(k) for k in ['user_id','user_name','user_email','user_department','user_role']}
    data['has_session'] = 'user_id' in session
    return jsonify({'success': True, 'session': data})

# Usuários online (para eventos WebSocket)
online_users = {}

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(2)  # Timeout de 2 segundos para não atrasar o startup
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        return local_ip
    except Exception:
        return "127.0.0.1"

def start_backup_scheduler(interval_seconds: int = 43200) -> None:
    """Thread simples para backups periódicos do SQLite usando API de backup."""
    def _worker():
        while True:
            try:
                # Snapshot para arquivo com timestamp
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                # Diretório de backup: variável de ambiente IPPEL_BACKUP_DIR (se definida) ou caminho padrão solicitado
                dest_dir = os.environ.get('IPPEL_BACKUP_DIR', r"G:\My Drive\BACKUP BANCO DE DADOS IPPEL")
                try:
                    os.makedirs(dest_dir, exist_ok=True)
                except Exception:
                    # Fallback: se não conseguir criar o diretório desejado, usar a pasta do projeto
                    dest_dir = os.path.dirname(os.path.abspath(__file__))
                dest = os.path.join(dest_dir, f"ippel_system_backup_{ts}.db")
                src = sqlite3.connect(DB_PATH, timeout=30.0)
                dst = sqlite3.connect(dest, timeout=30.0)
                with dst:
                    src.backup(dst)
                src.close(); dst.close()
                logger.info(f"Backup criado: {dest}")
            except Exception as e:
                try:
                    logger.error(f"Erro no backup: {e}")
                except Exception:
                    pass
            time.sleep(interval_seconds)
    threading.Thread(target=_worker, name='BackupScheduler', daemon=True).start()

def performance_monitor():
    """Monitor leve de performance (placeholder)."""
    while True:
        try:
            clear_expired_cache()
        except Exception:
            pass
        time.sleep(60)

def init_database():
    """Inicializa o banco de dados: tabelas, colunas e usuário admin padrão."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # PRAGMAs de performance
    cursor.execute('PRAGMA journal_mode=WAL')
    cursor.execute('PRAGMA synchronous=NORMAL')
    cursor.execute('PRAGMA cache_size=10000')
    cursor.execute('PRAGMA temp_store=MEMORY')
    cursor.execute('PRAGMA mmap_size=268435456')
    cursor.execute('PRAGMA optimize')

    # Tabela de grupos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            manager_user_id INTEGER,
            sub_manager_user_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (manager_user_id) REFERENCES users (id),
            FOREIGN KEY (sub_manager_user_id) REFERENCES users (id)
        )
    ''')
    # Adicionar colunas de gerente/sub-gerente se não existirem
    for col_sql in [
        'ALTER TABLE groups ADD COLUMN manager_user_id INTEGER',
        'ALTER TABLE groups ADD COLUMN sub_manager_user_id INTEGER']:
        try:
            cursor.execute(col_sql)
        except sqlite3.OperationalError:
            pass
    # Tabela de permissões por grupo
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS group_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            permission_name TEXT NOT NULL,
            permission_value BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (group_id) REFERENCES groups (id),
            UNIQUE(group_id, permission_name)
        )
    ''')

    # Tabela de usuários
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            department TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            permissions TEXT DEFAULT '[]',
            group_id INTEGER,
            avatar_key TEXT,
            avatar_prefs TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_active BOOLEAN DEFAULT 1,
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
    ''')
    # Garantir colunas novas (idempotente)
    for col_sql in [
        'ALTER TABLE users ADD COLUMN group_id INTEGER',
        'ALTER TABLE users ADD COLUMN avatar_key TEXT',
        'ALTER TABLE users ADD COLUMN avatar_prefs TEXT']:
        try:
            cursor.execute(col_sql)
        except sqlite3.OperationalError:
            pass

    # Tabela de RNCs
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS rncs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rnc_number TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            equipment TEXT,
            client TEXT,
            priority TEXT DEFAULT 'Média',
            status TEXT DEFAULT 'Pendente',
            user_id INTEGER,
            assigned_user_id INTEGER,
            is_deleted BOOLEAN DEFAULT 0,
            deleted_at TIMESTAMP,
            finalized_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (assigned_user_id) REFERENCES users (id)
        )
    ''')
    # Colunas adicionais em rncs
    for col_sql in [
        'ALTER TABLE rncs ADD COLUMN assigned_user_id INTEGER',
        'ALTER TABLE rncs ADD COLUMN is_deleted BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN deleted_at TIMESTAMP',
        'ALTER TABLE rncs ADD COLUMN finalized_at TIMESTAMP',
        'ALTER TABLE rncs ADD COLUMN price REAL DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN disposition_usar BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN disposition_retrabalhar BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN disposition_rejeitar BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN disposition_sucata BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN disposition_devolver_estoque BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN disposition_devolver_fornecedor BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN inspection_aprovado BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN inspection_reprovado BOOLEAN DEFAULT 0',
        'ALTER TABLE rncs ADD COLUMN inspection_ver_rnc TEXT',
        'ALTER TABLE rncs ADD COLUMN signature_inspection_date TEXT',
        'ALTER TABLE rncs ADD COLUMN signature_engineering_date TEXT',
        'ALTER TABLE rncs ADD COLUMN signature_inspection2_date TEXT',
        'ALTER TABLE rncs ADD COLUMN signature_inspection_name TEXT',
        'ALTER TABLE rncs ADD COLUMN signature_engineering_name TEXT',
        'ALTER TABLE rncs ADD COLUMN signature_inspection2_name TEXT',
        'ALTER TABLE rncs ADD COLUMN assigned_group_id INTEGER',
        'ALTER TABLE rncs ADD COLUMN causador_user_id INTEGER',
        'ALTER TABLE rncs ADD COLUMN ass_responsavel TEXT']:
        try:
            cursor.execute(col_sql)
        except sqlite3.OperationalError:
            pass

    # Compartilhamento de RNCs
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS rnc_shares (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rnc_id INTEGER NOT NULL,
            shared_by_user_id INTEGER NOT NULL,
            shared_with_user_id INTEGER NOT NULL,
            permission_level TEXT DEFAULT 'view',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (rnc_id) REFERENCES rncs (id) ON DELETE CASCADE,
            FOREIGN KEY (shared_by_user_id) REFERENCES users (id),
            FOREIGN KEY (shared_with_user_id) REFERENCES users (id),
            UNIQUE(rnc_id, shared_with_user_id)
        )
    ''')

    # Tabela de itens de valores/hora para orçamento de RNC
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS rnc_valores_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rnc_id INTEGER NOT NULL,
            codigo TEXT NOT NULL,
            descricao TEXT NOT NULL,
            setor TEXT,
            valor_hora REAL NOT NULL,
            horas REAL NOT NULL,
            subtotal REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (rnc_id) REFERENCES rncs (id) ON DELETE CASCADE
        )
    ''')

    # Tabelas auxiliares
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS chat_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rnc_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            message TEXT NOT NULL,
            message_type TEXT DEFAULT 'text',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (rnc_id) REFERENCES rncs (id),
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            rnc_id INTEGER,
            type TEXT NOT NULL,
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            is_read BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (rnc_id) REFERENCES rncs (id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS private_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_id INTEGER NOT NULL,
            recipient_id INTEGER NOT NULL,
            message TEXT NOT NULL,
            message_type TEXT DEFAULT 'text',
            is_read BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sender_id) REFERENCES users (id),
            FOREIGN KEY (recipient_id) REFERENCES users (id)
        )
    ''')
    
    # Tabela para rastrear mensagens visualizadas (evitar notificações duplicadas)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS message_views (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            rnc_id INTEGER NOT NULL,
            last_viewed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (rnc_id) REFERENCES rncs (id),
            UNIQUE(user_id, rnc_id)
        )
    ''')
    
    # Tabela de valores/hora
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS valores_hora (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            setor TEXT NOT NULL,
            descricao TEXT NOT NULL,
            valor_hora REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Usuário admin padrão
    cursor.execute('SELECT 1 FROM users WHERE role = "admin" LIMIT 1')
    if not cursor.fetchone():
        admin_password = generate_password_hash('admin123')
        cursor.execute('''
            INSERT INTO users (name, email, password_hash, department, role, permissions)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', ('Administrador', 'admin@ippel.com.br', admin_password, 'TI', 'admin', '["all"]'))

    conn.commit()
    conn.close()
    # Mover ensure_rnc_extra_columns() para background para não atrasar startup
    def delayed_column_check():
        time.sleep(5)  # Aguarda 5 segundos após início do servidor
    try:
        ensure_rnc_extra_columns()
        ensure_chat_viewed_at_column()
        ensure_notifications_table_migration()
    except Exception:
        pass
    threading.Thread(target=delayed_column_check, daemon=True).start()

def ensure_rnc_extra_columns():
    """Garante que a tabela rncs possua colunas extras usadas no modo Responder/Visualização.

    Campos:
      - instruction_retrabalho TEXT
      - cause_rnc TEXT
      - action_rnc TEXT
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('PRAGMA table_info(rncs)')
        cols = [row[1] for row in cursor.fetchall()]
        additions = []
        if 'instruction_retrabalho' not in cols:
            additions.append("ALTER TABLE rncs ADD COLUMN instruction_retrabalho TEXT")
        if 'cause_rnc' not in cols:
            additions.append("ALTER TABLE rncs ADD COLUMN cause_rnc TEXT")
        if 'action_rnc' not in cols:
            additions.append("ALTER TABLE rncs ADD COLUMN action_rnc TEXT")
        for sql in additions:
            try:
                cursor.execute(sql)
            except Exception:
                pass
        conn.commit()
        conn.close()
    except Exception as e:
        try:
            logger.error(f"Erro ao garantir colunas extras dos RNCs: {e}")
        except Exception:
            pass

def ensure_chat_viewed_at_column(existing_conn=None):
    """Garante que a tabela chat_messages possua a coluna viewed_at para funcionalidade WhatsApp-style"""
    try:
        if existing_conn:
            conn = existing_conn
            should_close = False
        else:
            conn = sqlite3.connect(DB_PATH)
            should_close = True
            
        cursor = conn.cursor()
        
        # Verificar se a coluna viewed_at existe
        cursor.execute('PRAGMA table_info(chat_messages)')
        cols = [row[1] for row in cursor.fetchall()]
        
        if 'viewed_at' not in cols:
            logger.info(" Adicionando coluna viewed_at à tabela chat_messages")
            cursor.execute("ALTER TABLE chat_messages ADD COLUMN viewed_at TIMESTAMP DEFAULT NULL")
            conn.commit()
            logger.info(" Coluna viewed_at adicionada com sucesso")
        
        if should_close:
            conn.close()
    except Exception as e:
        logger.error(f" Erro ao garantir coluna viewed_at em chat_messages: {e}")

def ensure_notifications_table_migration():
    """Migração da tabela notifications para compatibilidade com enhanced_notifications"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar estrutura atual da tabela notifications
        cursor.execute('PRAGMA table_info(notifications)')
        cols = {row[1]: row for row in cursor.fetchall()}
        
        # Verificar se precisa migrar de user_id para to_user_id
        if 'user_id' in cols and 'to_user_id' not in cols:
            logger.info(" Migrando tabela notifications: user_id -> to_user_id")
            
            # Criar nova tabela com estrutura correta
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS notifications_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    type TEXT NOT NULL,
                    priority TEXT DEFAULT 'normal',
                    title TEXT NOT NULL,
                    message TEXT NOT NULL,
                    data TEXT,
                    from_user_id INTEGER,
                    to_user_id INTEGER NOT NULL,
                    rnc_id INTEGER,
                    is_read BOOLEAN DEFAULT 0,
                    is_dismissed BOOLEAN DEFAULT 0,
                    read_at TIMESTAMP,
                    dismissed_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    expires_at TIMESTAMP,
                    FOREIGN KEY (from_user_id) REFERENCES users(id),
                    FOREIGN KEY (to_user_id) REFERENCES users(id),
                    FOREIGN KEY (rnc_id) REFERENCES rncs(id)
                )
            ''')
            
            # Copiar dados da tabela antiga para a nova (user_id -> to_user_id)
            cursor.execute('''
                INSERT INTO notifications_new 
                (id, type, title, message, to_user_id, rnc_id, is_read, created_at)
                SELECT id, type, title, message, user_id, rnc_id, is_read, created_at
                FROM notifications
            ''')
            
            # Remover tabela antiga e renomear a nova
            cursor.execute('DROP TABLE notifications')
            cursor.execute('ALTER TABLE notifications_new RENAME TO notifications')
            
            # Criar índices
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_notifications_user_read ON notifications(to_user_id, is_read)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_notifications_created ON notifications(created_at)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_notifications_type ON notifications(type)')
            
            conn.commit()
            logger.info(" Tabela notifications migrada com sucesso!")
            
        # Adicionar colunas faltantes se a tabela já usa to_user_id
        elif 'to_user_id' in cols:
            added_cols = []
            
            if 'priority' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN priority TEXT DEFAULT 'normal'")
                added_cols.append('priority')
            
            if 'data' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN data TEXT")
                added_cols.append('data')
            
            if 'from_user_id' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN from_user_id INTEGER")
                added_cols.append('from_user_id')
            
            if 'is_dismissed' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN is_dismissed BOOLEAN DEFAULT 0")
                added_cols.append('is_dismissed')
            
            if 'read_at' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN read_at TIMESTAMP")
                added_cols.append('read_at')
            
            if 'dismissed_at' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN dismissed_at TIMESTAMP")
                added_cols.append('dismissed_at')
            
            if 'expires_at' not in cols:
                cursor.execute("ALTER TABLE notifications ADD COLUMN expires_at TIMESTAMP")
                added_cols.append('expires_at')
            
            if added_cols:
                conn.commit()
                logger.info(f" Colunas adicionadas à tabela notifications: {', '.join(added_cols)}")
        
        conn.close()
        
    except Exception as e:
        logger.error(f" Erro ao migrar tabela notifications: {e}")
        import traceback
        logger.error(traceback.format_exc())

def get_user_by_email(email):
    """Buscar usuário por email"""
    cache_key = f"user_email_{email}"
    cached_result = get_cached_query(cache_key)
    if cached_result:
        return cached_result
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE email = ? AND is_active = 1', (email,))
        user = cursor.fetchone()
        cache_query(cache_key, user, ttl=600)  # Cache por 10 minutos
        return user
    finally:
        if conn:
            return_db_connection(conn)

def get_user_by_id(user_id):
    """Buscar usuário por ID"""
    cache_key = f"user_id_{user_id}"
    cached_result = get_cached_query(cache_key)
    if cached_result:
        return cached_result
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ? AND is_active = 1', (user_id,))
        user = cursor.fetchone()
        cache_query(cache_key, user, ttl=600)  # Cache por 10 minutos
        return user
    finally:
        if conn:
            return_db_connection(conn)

# Removed redundant function definitions since they're already imported from services.permissions
# The following functions are now used from imports:
# - has_permission (from services.permissions)
# - get_user_department (from services.permissions) 
# - has_department_permission (from services.permissions)

def get_all_users():
    """Buscar todos os usuários ativos"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.id, u.name, u.email, u.department, u.role, u.permissions, u.created_at, g.name as group_name
        FROM users u
        LEFT JOIN groups g ON u.group_id = g.id
        WHERE u.is_active = 1
        ORDER BY u.name
    ''')
    users = cursor.fetchall()
    conn.close()
    return users

def create_user(name, email, password, department, role, permissions):
    """Criar novo usuário"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    password_hash = generate_password_hash(password)
    import json
    permissions_json = json.dumps(permissions)
    
    try:
        # Verificar se o grupo existe, se não, criar automaticamente
        group_id = None
        if department:
            cursor.execute('SELECT id FROM groups WHERE name = ?', (department,))
            existing_group = cursor.fetchone()
            
            if existing_group:
                group_id = existing_group[0]
            else:
                # Criar o grupo automaticamente
                cursor.execute('INSERT INTO groups (name, description) VALUES (?, ?)', 
                             (department, f'Grupo criado automaticamente para {department}'))
                group_id = cursor.lastrowid
        
        cursor.execute('''
            INSERT INTO users (name, email, password_hash, department, role, permissions, group_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (name, email, password_hash, department, role, permissions_json, group_id))
        
        conn.commit()

        # Índices para performance em consultas do dashboard
        try:
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_created_at ON rncs(created_at)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_status ON rncs(status)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_user ON rncs(user_id)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_assigned_user ON rncs(assigned_user_id)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_client ON rncs(client)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_equipment ON rncs(equipment)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_priority ON rncs(priority)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_is_deleted ON rncs(is_deleted)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_rncs_finalized_at ON rncs(finalized_at)')
            conn.commit()
        except Exception as e:
            logger.error(f"Erro ao criar índices: {e}")
        conn.close()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def update_user(user_id, name, email, department, role, permissions, is_active):
    """Atualizar usuário"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    import json
    permissions_json = json.dumps(permissions)
    
    # Verificar se o grupo existe, se não, criar automaticamente
    group_id = None
    if department:
        cursor.execute('SELECT id FROM groups WHERE name = ?', (department,))
        existing_group = cursor.fetchone()
        
        if existing_group:
            group_id = existing_group[0]
        else:
            # Criar o grupo automaticamente
            cursor.execute('INSERT INTO groups (name, description) VALUES (?, ?)', 
                         (department, f'Grupo criado automaticamente para {department}'))
            group_id = cursor.lastrowid
    
    cursor.execute('''
        UPDATE users 
        SET name = ?, email = ?, department = ?, role = ?, permissions = ?, is_active = ?, group_id = ?
        WHERE id = ?
    ''', (name, email, department, role, permissions_json, is_active, group_id, user_id))
    
    conn.commit()
    conn.close()
    return True

def delete_user(user_id):
    """Desativar usuário (soft delete)"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET is_active = 0 WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return True

def restore_user(user_id: int) -> bool:
    """Reativar usuário previamente desativado."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET is_active = 1 WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return True

def update_user_password(user_id: int, new_password: str) -> bool:
    """Atualizar a senha de um usuário."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    new_hash = generate_password_hash(new_password)
    cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?', (new_hash, user_id))
    conn.commit()
    conn.close()
    return True

def get_all_groups():
    """Obter todos os grupos"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT g.id, g.name, g.description, COUNT(u.id) as user_count
            FROM groups g
            LEFT JOIN users u ON g.id = u.group_id AND u.is_active = 1
            GROUP BY g.id
            ORDER BY g.name
        ''')
        groups = cursor.fetchall()
        conn.close()
        return groups
    except Exception as e:
        logger.error(f"Erro ao buscar grupos: {e}")
        return []

def get_group_by_id(group_id):
    """Obter grupo por ID"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM groups WHERE id = ?', (group_id,))
        group = cursor.fetchone()
        conn.close()
        return group
    except Exception as e:
        logger.error(f"Erro ao buscar grupo: {e}")
        return None

def create_group(name, description):
    """Criar novo grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('INSERT INTO groups (name, description) VALUES (?, ?)', (name, description))
        group_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return group_id
    except Exception as e:
        logger.error(f"Erro ao criar grupo: {e}")
        return None

def update_group(group_id, name, description):
    """Atualizar grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('UPDATE groups SET name = ?, description = ? WHERE id = ?', (name, description, group_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Erro ao atualizar grupo: {e}")
        return False

def delete_group(group_id):
    """Deletar grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        # Primeiro, remover referências de usuários
        cursor.execute('UPDATE users SET group_id = NULL WHERE group_id = ?', (group_id,))
        # Depois deletar o grupo
        cursor.execute('DELETE FROM groups WHERE id = ?', (group_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Erro ao deletar grupo: {e}")
        return False

def get_users_by_group(group_id):
    """Obter usuários de um grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, name, email, department, role, is_active
            FROM users 
            WHERE group_id = ? AND is_active = 1
            ORDER BY name
        ''', (group_id,))
        users = cursor.fetchall()
        conn.close()
        return users
    except Exception as e:
        logger.error(f"Erro ao buscar usuários do grupo: {e}")
        return []
def share_rnc_with_user(rnc_id, shared_by_user_id, shared_with_user_id, permission_level='view'):
    """Compartilhar RNC com usuário com tolerância a lock do SQLite (retry/backoff)."""
    attempt = 0
    max_attempts = 5
    backoff_base = 0.2
    while attempt < max_attempts:
        conn = None
        try:
            conn = sqlite3.connect(DB_PATH, timeout=10, isolation_level=None)
            cursor = conn.cursor()
            cursor.execute('PRAGMA busy_timeout=5000')
            cursor.execute('BEGIN IMMEDIATE')
            cursor.execute(
                '''INSERT OR REPLACE INTO rnc_shares (rnc_id, shared_by_user_id, shared_with_user_id, permission_level)
                   VALUES (?, ?, ?, ?)''',
                (rnc_id, shared_by_user_id, shared_with_user_id, permission_level)
            )
            conn.commit()
            return True
        except sqlite3.OperationalError as e:
            msg = str(e).lower()
            if 'locked' in msg or 'busy' in msg:
                attempt += 1
                wait = backoff_base * attempt
                try:
                    logger.warning(f"Lock ao compartilhar RNC {rnc_id} (tentativa {attempt}/{max_attempts}) aguardando {wait:.2f}s: {e}")
                except Exception:
                    pass
                time.sleep(wait)
                continue
            else:
                try:
                    logger.error(f"Erro operacional ao compartilhar RNC {rnc_id}: {e}")
                except Exception:
                    pass
                return False
        except Exception as e:
            try:
                logger.error(f"Erro inesperado ao compartilhar RNC {rnc_id}: {e}")
            except Exception:
                pass
            return False
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass
    try:
        logger.warning("share_rnc_with_user: desistindo após múltiplas tentativas (database locked)")
    except Exception:
        pass
    return False

def get_rnc_shared_users(rnc_id):
    """Obter usuários com quem a RNC foi compartilhada"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT rs.shared_with_user_id, rs.permission_level, u.name, u.email
            FROM rnc_shares rs
            JOIN users u ON rs.shared_with_user_id = u.id
            WHERE rs.rnc_id = ?
        ''', (rnc_id,))
        shared_users = cursor.fetchall()
        conn.close()
        return shared_users
    except Exception as e:
        logger.error(f"Erro ao buscar usuários compartilhados da RNC {rnc_id}: {e}")
        return []

def can_user_access_rnc(user_id, rnc_id):
    """Verificar se usuário pode acessar RNC (criador, compartilhado ou admin)"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se é o criador
        cursor.execute('SELECT user_id, department FROM rncs WHERE id = ?', (rnc_id,))
        rnc_data = cursor.fetchone()
        if not rnc_data:
            conn.close()
            logger.warning(f"RNC {rnc_id} não encontrada na verificação de acesso")
            return False
        
        logger.info(f"Verificando acesso do usuário {user_id} à RNC {rnc_id}, criada por {rnc_data[0]}, departamento: {rnc_data[1]}")
        
        # Se é o criador
        if str(rnc_data[0]) == str(user_id):
            logger.info(f"Acesso permitido: usuário {user_id} é o criador da RNC {rnc_id}")
            conn.close()
            return True
        
        # Verificar permissões
        if has_permission(user_id, 'view_all_rncs') or has_permission(user_id, 'admin_access'):
            logger.info(f"Acesso permitido: usuário {user_id} tem permissão admin ou view_all_rncs")
            conn.close() 
            return True
            
        # Verificar acesso por departamento
        rnc_department = rnc_data[1]
        if rnc_department == 'Engenharia' and has_permission(user_id, 'view_engineering_rncs'):
            logger.info(f"Acesso permitido: usuário {user_id} tem permissão view_engineering_rncs e RNC é da Engenharia")
            conn.close()
            return True
            
        if has_permission(user_id, 'view_all_departments_rncs'):
            logger.info(f"Acesso permitido: usuário {user_id} tem permissão view_all_departments_rncs")
            conn.close()
            return True
        
        # Verificar se foi compartilhado com o usuário
        cursor.execute('''
            SELECT permission_level FROM rnc_shares 
            WHERE rnc_id = ? AND shared_with_user_id = ?
        ''', (rnc_id, user_id))
        share_data = cursor.fetchone()
        
        conn.close()
        has_access = share_data is not None
        if has_access:
            logger.info(f"Acesso permitido: RNC {rnc_id} foi compartilhada com o usuário {user_id}")
        else:
            logger.warning(f"Acesso negado: usuário {user_id} não tem permissões necessárias para RNC {rnc_id}")
        return has_access
        
    except Exception as e:
        logger.error(f"Erro ao verificar acesso à RNC: {e}")
        return False

# ==================== FUNÇÕES DE PERMISSÕES ====================

def get_group_permissions(group_id):
    """Obter todas as permissões de um grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT permission_name, permission_value
            FROM group_permissions
            WHERE group_id = ?
            ORDER BY permission_name
        ''', (group_id,))
        permissions = cursor.fetchall()
        conn.close()
        return permissions
    except Exception as e:
        logger.error(f"Erro ao buscar permissões do grupo: {e}")
        return []

def update_group_permissions(group_id, permissions):
    """Atualizar permissões de um grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Remover permissões existentes
        cursor.execute('DELETE FROM group_permissions WHERE group_id = ?', (group_id,))
        
        # Inserir novas permissões
        for permission_name, permission_value in permissions.items():
            cursor.execute('''
                INSERT INTO group_permissions (group_id, permission_name, permission_value)
                VALUES (?, ?, ?)
            ''', (group_id, permission_name, 1 if permission_value else 0))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Erro ao atualizar permissões do grupo: {e}")
        return False

def user_in_group(user_id, group_name):
    """Verificar se o usuário está em um grupo específico pelo nome do grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT COUNT(*) 
            FROM users u
            JOIN groups g ON u.group_id = g.id
            WHERE u.id = ? AND UPPER(g.name) = UPPER(?)
        ''', (user_id, group_name))
        
        result = cursor.fetchone()
        conn.close()
        return result and result[0] > 0
    except Exception as e:
        logger.error(f"Erro ao verificar grupo do usuário: {e}")
        return False

def get_user_group_permissions(user_id):
    """Obter permissões do usuário baseadas no seu grupo"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se é admin
        cursor.execute('SELECT role FROM users WHERE id = ?', (user_id,))
        result = cursor.fetchone()
        
        if result and result[0] == 'admin':
            conn.close()
            return ['all']  # Admin tem todas as permissões
        
        # Buscar permissões do grupo
        cursor.execute('''
            SELECT gp.permission_name
            FROM group_permissions gp
            JOIN users u ON u.group_id = gp.group_id
            WHERE u.id = ? AND gp.permission_value = 1
        ''', (user_id,))
        
        permissions = [row[0] for row in cursor.fetchall()]
        conn.close()
        return permissions
    except Exception as e:
        logger.error(f"Erro ao buscar permissões do usuário: {e}")
        return []

def clear_permissions_cache():
    """Limpar cache de permissões"""
    from services.cache import cache_lock, query_cache
    with cache_lock:
        keys_to_remove = [key for key in list(query_cache.keys()) if 'permissions' in key]
        for key in keys_to_remove:
            query_cache.pop(key, None)
    logger.info("Cache de permissões limpo")

def get_all_permissions():
    """Obter lista de todas as permissões disponíveis com descrições"""
    # Limpar qualquer cache de permissões para forçar dados atualizados
    clear_permissions_cache()
    
    permissions = [
        {
            'name': 'create_rnc',
            'display_name': 'Criar RNC',
            'description': 'Permite criar novas RNCs no sistema'
        },
        {
            'name': 'assign_rnc_to_group',
            'display_name': 'Atribuir RNC ao Grupo',
            'description': 'Permite atribuir RNCs visíveis apenas para o grupo/usuários selecionados na criação'
        },
        {
            'name': 'update_avatar',
            'display_name': 'Atualizar Avatar do Usuário',
            'description': 'Permite enviar e alterar a imagem de avatar do usuário'
        },
        {
            'name': 'edit_own_rnc',
            'display_name': 'Editar Próprias RNCs',
            'description': 'Permite editar RNCs criadas pelo próprio usuário'
        },
        {
            'name': 'view_own_rnc',
            'display_name': 'Visualizar Próprias RNCs',
            'description': 'Permite visualizar RNCs criadas pelo próprio usuário'
        },
        {
            'name': 'view_all_rncs',
            'display_name': 'Visualizar Todas as RNCs',
            'description': 'Permite visualizar todas as RNCs do sistema'
        },
        {
            'name': 'edit_all_rncs',
            'display_name': 'Editar Todas as RNCs',
            'description': 'Permite editar qualquer RNC do sistema'
        },
        {
            'name': 'delete_rnc',
            'display_name': 'Excluir RNCs',
            'description': 'Permite excluir RNCs do sistema'
        },
        {
            'name': 'view_finalized_rncs',
            'display_name': 'Visualizar RNCs Finalizadas',
            'description': 'Permite visualizar RNCs com status finalizado'
        },
        {
            'name': 'view_charts',
            'display_name': 'Visualizar Gráficos',
            'description': 'Permite acessar gráficos e estatísticas'
        },
        {
            'name': 'view_reports',
            'display_name': 'Visualizar Relatórios',
            'description': 'Permite acessar relatórios e indicadores'
        },
        {
            'name': 'admin_access',
            'display_name': 'Acesso Administrativo',
            'description': 'Acesso total ao sistema administrativo'
        },
        {
            'name': 'manage_users',
            'display_name': 'Gerenciar Usuários',
            'description': 'Permite criar, editar e excluir usuários'
        },
        {
            'name': 'view_engineering_rncs',
            'display_name': 'Visualizar RNCs da Engenharia',
            'description': 'Permite visualizar RNCs do departamento de engenharia'
        },
        {
            'name': 'view_all_departments_rncs',
            'display_name': 'Visualizar RNCs de Todos os Departamentos',
            'description': 'Permite visualizar RNCs de qualquer departamento'
        },
        {
            'name': 'view_levantamento_14_15',
            'display_name': 'Visualizar Levantamento 14/15',
            'description': 'Permite acessar dados do levantamento 14/15'
        },
        {
            'name': 'view_groups_for_assignment',
            'display_name': 'Visualizar Grupos para Atribuição',
            'description': 'Permite visualizar grupos para atribuição de RNCs'
        },
        {
            'name': 'view_users_for_assignment',
            'display_name': 'Visualizar Usuários para Atribuição',
            'description': 'Permite visualizar usuários para atribuição de RNCs'
        },
        {
            'name': 'reply_rncs',
            'display_name': 'Responder RNC',
            'description': 'Permite responder e reabrir qualquer RNC do sistema'
        }
    ]
    
    return permissions

@app.route('/')
def index():
    """Página principal com login sempre como entrada."""
    try:
        if 'user_id' in session:
            return redirect('/dashboard')
        # Tentar template personalizado de login; fallback para index.html
        try:
            return render_template('login.html')
        except Exception:
            return send_from_directory('.', 'index.html')
    except Exception:
        return send_from_directory('.', 'index.html')

@app.route('/teste-som-notificacoes')
def teste_som_notificacoes():
    """Página de teste de som e notificações"""
    return render_template('teste_som_notificacoes.html')

@app.route('/diagnostico-notificacoes')
def diagnostico_notificacoes():
    """Página de diagnóstico completo de notificações do Windows"""
    return render_template('diagnostico_notificacoes.html')

@app.route('/dashboard')
def dashboard():
    """Dashboard interativo protegido por sessão."""
    if 'user_id' not in session:
        return redirect('/')
    
    # Obter informações de permissões do usuário para o frontend
    user_permissions = {
        'canViewAllRncs': has_permission(session['user_id'], 'view_all_rncs'),
        'canViewFinalizedRncs': has_permission(session['user_id'], 'view_finalized_rncs'),
        'canViewCharts': has_permission(session['user_id'], 'view_charts'),
        'canViewReports': has_permission(session['user_id'], 'view_reports'),
    'hasAdminAccess': has_permission(session['user_id'], 'admin_access'),
    # Permissão correta: qualquer usuário com 'create_rnc' pode criar RNCs
    'canCreateRnc': has_permission(session['user_id'], 'create_rnc'),
        'canAssignRncToGroup': has_permission(session['user_id'], 'assign_rnc_to_group'),
        'canViewLevantamento1415': has_permission(session['user_id'], 'view_levantamento_14_15'),
        'canViewGroupsForAssignment': has_permission(session['user_id'], 'view_groups_for_assignment'),
        'canViewUsersForAssignment': has_permission(session['user_id'], 'view_users_for_assignment'),
        'canViewEngineeringRncs': has_permission(session['user_id'], 'view_engineering_rncs'),
        'canPrintReports': True,  # Forçando permissão de impressão para todos os usuários
        'department': get_user_department(session['user_id'])
    }
    
    return render_template('dashboard_improved.html', user_permissions=user_permissions)

@app.route('/dashboard/expenses')
def dashboard_expenses():
    """Dashboard com gastos por funcionário e setor."""
    if 'user_id' not in session:
        return redirect('/')
    
    user_id = session['user_id']
    user = get_user_by_id(user_id)
    
    if not user:
        return redirect('/')
    
    # Buscar todas as RNCs para calcular estatísticas
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Estatísticas gerais
    cursor.execute("SELECT COUNT(*) FROM rncs WHERE is_deleted = 0")
    total_rncs = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM rncs WHERE status = 'Finalizado' AND is_deleted = 0")
    finalized_rncs = cursor.fetchone()[0]
    
    cursor.execute("SELECT SUM(price) FROM rncs WHERE is_deleted = 0")
    total_value = cursor.fetchone()[0] or 0
    
    cursor.execute("SELECT SUM(price) FROM rncs WHERE status = 'Finalizado' AND is_deleted = 0")
    finalized_value = cursor.fetchone()[0] or 0
    
    # Organizar dados por departamento e responsável
    cursor.execute("""
        SELECT department, responsavel, SUM(price) as total_value, COUNT(*) as rnc_count
        FROM rncs 
        WHERE is_deleted = 0 AND responsavel IS NOT NULL AND responsavel != ''
        GROUP BY department, responsavel
        ORDER BY department, total_value DESC
    """)
    
    dept_employee_data = cursor.fetchall()
    
    # Organizar em estrutura hierárquica
    departments = {}
    
    for dept, responsavel, value, count in dept_employee_data:
        if dept not in departments:
            departments[dept] = {'employees': {}, 'total': 0}
        
        # Usar o responsável diretamente
        employee_name = responsavel if responsavel else "Sistema"
        
        departments[dept]['employees'][employee_name] = value
        departments[dept]['total'] += value
    
    conn.close()
    
    stats = {
        'total_rncs': total_rncs,
        'finalized_rncs': finalized_rncs,
        'total_value': total_value,
        'finalized_value': finalized_value
    }
    
    return render_template('dashboard_with_employee_expenses.html', 
                         user=user, 
                         stats=stats,
                         departments=departments)

# ===================== MONITORING DASHBOARD (Admin) =====================
@app.route('/admin/monitoring')
def monitoring_dashboard():
    if 'user_id' not in session:
        return redirect('/')
    try:
        from services.permissions import has_permission
        if not has_permission(session['user_id'], 'admin_access'):
            return redirect('/dashboard?error=access_denied&message=Acesso negado ao painel de monitoramento')
    except Exception:
        return redirect('/dashboard')
    return render_template('monitoring_dashboard.html')

@app.get('/api/monitoring/security-events')
def api_monitoring_security_events():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    try:
        from services.permissions import has_permission
        if not has_permission(session['user_id'], 'admin_access'):
            return jsonify({'success': False, 'message': 'Sem permissão'}), 403
        import os, json
        limit = int(request.args.get('limit', 200))
        if limit < 1:
            limit = 1
        if limit > 2000:
            limit = 2000
        # Localizar arquivo de log de segurança
        base_dir = os.path.dirname(os.path.abspath(__file__))
        logs_dir = os.path.join(os.path.dirname(base_dir), 'logs')
        log_path = os.path.join(logs_dir, 'security.log')
        events = []
        if os.path.exists(log_path):
            # Ler de trás para frente com consumo moderado
            with open(log_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()[-limit:]
            for ln in lines:
                try:
                    events.append(json.loads(ln.strip()))
                except Exception:
                    continue
        return jsonify({'success': True, 'events': events})
    except Exception as e:
        try: logger.error(f"Erro ao ler eventos de segurança: {e}")
        except Exception: pass
        return jsonify({'success': False, 'message': 'Erro ao carregar eventos'}), 500

@app.get('/api/monitoring/summary')
def api_monitoring_summary():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    try:
        from services.permissions import has_permission
        if not has_permission(session['user_id'], 'admin_access'):
            return jsonify({'success': False, 'message': 'Sem permissão'}), 403
        import os, json, time
        from datetime import datetime, timedelta
        base_dir = os.path.dirname(os.path.abspath(__file__))
        logs_dir = os.path.join(os.path.dirname(base_dir), 'logs')
        log_path = os.path.join(logs_dir, 'security.log')
        now = datetime.utcnow()
        window_hours = int(request.args.get('hours', 24))
        if window_hours < 1:
            window_hours = 1
        if window_hours > 168:
            window_hours = 168
        since = now - timedelta(hours=window_hours)

        counters = {
            'auth_success': 0,
            'auth_fail': 0,
            'auth_lockout': 0,
            'api_unauthorized': 0,
        }
        timeline = []  # [(iso, count)] aggregated per hour for failures
        bucket = {}

        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                for ln in f:
                    try:
                        ev = json.loads(ln)
                    except Exception:
                        continue
                    ts = ev.get('ts')
                    if not ts:
                        continue
                    try:
                        dt = datetime.fromisoformat(ts.replace('Z','').replace('z',''))
                    except Exception:
                        continue
                    if dt < since:
                        continue
                    cat = str(ev.get('cat') or '')
                    act = str(ev.get('act') or '')
                    status = str(ev.get('status') or '')
                    # Counters
                    if cat == 'auth' and act == 'login' and status == 'success':
                        counters['auth_success'] += 1
                    if cat == 'auth' and act == 'login' and status == 'fail':
                        counters['auth_fail'] += 1
                        key = dt.strftime('%Y-%m-%d %H:00')
                        bucket[key] = bucket.get(key, 0) + 1
                    if cat == 'auth' and act == 'lockout':
                        counters['auth_lockout'] += 1
                        key = dt.strftime('%Y-%m-%d %H:00')
                        bucket[key] = bucket.get(key, 0) + 1
                    if cat == 'api' and act == 'unauthorized':
                        counters['api_unauthorized'] += 1

        # Build ordered timeline
        for k in sorted(bucket.keys()):
            timeline.append({'bucket': k, 'count': bucket[k]})

        # Lockouts at the moment
        try:
            import sqlite3
            conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
            cur.execute('SELECT COUNT(*) FROM login_lockouts WHERE locked_until IS NOT NULL AND locked_until > strftime("%s","now")')
            active_lockouts = cur.fetchone()[0]
            conn.close()
        except Exception:
            active_lockouts = 0

        return jsonify({'success': True, 'window_hours': window_hours, 'counters': counters, 'timeline': timeline, 'active_lockouts': active_lockouts})
    except Exception as e:
        try: logger.error(f"Erro no summary de monitoramento: {e}")
        except Exception: pass
        return jsonify({'success': False, 'message': 'Erro ao carregar resumo'}), 500

@app.get('/api/monitoring/lockouts')
def api_monitoring_lockouts():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    try:
        from services.permissions import has_permission
        if not has_permission(session['user_id'], 'admin_access'):
            return jsonify({'success': False, 'message': 'Sem permissão'}), 403
        import sqlite3
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        cur.execute('''
            SELECT ll.user_id, u.name, u.email, ll.failed_count, ll.locked_until
              FROM login_lockouts ll
              LEFT JOIN users u ON u.id = ll.user_id
             WHERE ll.locked_until IS NOT NULL AND ll.locked_until > strftime('%s','now')
             ORDER BY ll.locked_until DESC
             LIMIT 100
        ''')
        rows = cur.fetchall(); conn.close()
        data = [
            {
                'user_id': r[0],
                'name': r[1],
                'email': r[2],
                'failed_count': r[3],
                'locked_until': int(r[4]) if r[4] is not None else None
            } for r in rows
        ]
        return jsonify({'success': True, 'lockouts': data})
    except Exception as e:
        try: logger.error(f"Erro ao listar lockouts: {e}")
        except Exception: pass
        return jsonify({'success': False, 'message': 'Erro ao carregar lockouts'}), 500

# Rota /indicadores-dashboard removida - funcionalidade descontinuada

# Rotas e APIs de indicadores removidas completamente - funcionalidade descontinuada

@app.route('/api/indicadores/engenharia')
def api_indicadores_engenharia():
    """API específica para dados da engenharia"""
    # Campos retornados:
    #  - stats: { total_rncs, finalized_rncs, active_rncs, total_value, avg_value }
    #  - monthly_trend: lista com { month (YYYY-MM), count, accumulated_count, value, accumulated_value }
    #  - rncs_count: número total de RNCs (mesmo que stats.total_rncs)
    #  - rncs: lista simplificada de RNCs para tabela
    # Observação: Frontend (dashboard_improved.html) usa rncs_count ou stats.finalized_rncs
    # para preencher badge de Engenharia; patch garante atualização imediata.
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    
    try:
        conn = sqlite3.connect('ippel_system.db')
        cursor = conn.cursor()
        
        # Calcular data limite (últimos 12 meses)
        twelve_months_ago = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        # Buscar RNCs da engenharia dos últimos 12 meses
        cursor.execute("""
            SELECT 
                id, rnc_number, title, equipment, client, priority, status,
                responsavel, setor, area_responsavel, finalized_at, created_at,
                price
            FROM rncs 
            WHERE (
                LOWER(TRIM(area_responsavel)) LIKE '%engenharia%'
                OR LOWER(TRIM(setor)) LIKE '%engenharia%'
                OR LOWER(TRIM(signature_engineering_name)) LIKE '%engenharia%'
            )
            AND (is_deleted = 0 OR is_deleted IS NULL)
            AND (
                (finalized_at IS NOT NULL AND DATE(finalized_at) >= ?)
                OR (finalized_at IS NULL AND DATE(created_at) >= ?)
            )
            ORDER BY COALESCE(finalized_at, created_at) DESC
        """, (twelve_months_ago, twelve_months_ago))
        
        rncs_raw = cursor.fetchall()
        
        # Dados mensais para gráficos (baseado em finalized_at)
        monthly_data = {}
        cumulative_data = {}
        total_value = 0
        
        for rnc in rncs_raw:
            finalized_at = rnc[10]  # finalized_at
            created_at = rnc[11]    # created_at
            price_str = rnc[12] or "0"  # price
            
            # Converter preço string para float
            try:
                if isinstance(price_str, str):
                    # Remover R$, espaços e vírgulas, substituir vírgula por ponto
                    price_clean = price_str.replace('R$', '').replace(' ', '').replace(',', '.')
                    price = float(price_clean) if price_clean else 0.0
                else:
                    price = float(price_str) if price_str else 0.0
            except (ValueError, TypeError):
                price = 0.0
            
            # Filtrar apenas RNCs finalizadas
            if not finalized_at:
                continue
                
            total_value += price
            
            # Usar apenas finalized_at para RNCs finalizadas
            date_to_use = finalized_at
            if date_to_use and date_to_use != '':
                try:
                    # Tentar parse com hora
                    if isinstance(date_to_use, str):
                        date_str = date_to_use.strip()
                        if ' ' in date_str:
                            # Formato com hora: 2023-01-02 14:30:00
                            date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            # Formato apenas data: 2023-01-02
                            date = datetime.strptime(date_str, '%Y-%m-%d')
                    else:
                        # Se já for datetime
                        date = date_to_use
                except Exception as parse_err:
                    # Fallback: tentar created_at se finalized_at falhar
                    try:
                        if created_at:
                            created_str = str(created_at).strip().split(' ')[0]
                            date = datetime.strptime(created_str, '%Y-%m-%d')
                        else:
                            continue
                    except:
                        # Se ambos falharem, pular esta RNC
                        continue
                        
                month_key = date.strftime('%Y-%m')
                month_label = date.strftime('%b/%Y')
                
                if month_key not in monthly_data:
                    monthly_data[month_key] = {'label': month_label, 'count': 0, 'value': 0, 'finalized': 0, 'active': 0}
                
                monthly_data[month_key]['count'] += 1
                monthly_data[month_key]['value'] += price
                
                # Classificar entre finalizadas e ativas
                status = rnc[6]  # status column
                if status == 'Finalizado' or finalized_at:
                    monthly_data[month_key]['finalized'] += 1
                else:
                    monthly_data[month_key]['active'] += 1
        
        # Ordenar por data e calcular acumulado
        sorted_months = sorted(monthly_data.keys())
        accumulated_count = 0
        accumulated_value = 0
        
        monthly_trend = []
        for month_key in sorted_months:
            data = monthly_data[month_key]
            accumulated_count += data['count']
            accumulated_value += data['value']
            
            monthly_trend.append({
                'month': month_key,
                'label': data['label'],
                'count': data['count'],
                'value': data['value'],
                'accumulated_count': accumulated_count,
                'accumulated_value': accumulated_value
            })
        
        # Estatísticas gerais - contar finalizadas vs ativas
        finalized_count = 0
        active_count = 0
        
        for rnc in rncs_raw:
            status = rnc[6]  # status column
            finalized_at = rnc[10]  # finalized_at column
            
            if status == 'Finalizado' or finalized_at:
                finalized_count += 1
            else:
                active_count += 1
        
        stats = {
            'total_rncs': len(rncs_raw),
            'finalized_rncs': finalized_count,
            'active_rncs': active_count,
            'total_value': total_value,
            'avg_value': total_value / max(len(rncs_raw), 1) if len(rncs_raw) > 0 else 0,
            'latest_month': monthly_trend[-1] if monthly_trend else None
        }
        
        conn.close()
        
        return jsonify({
            'success': True,
            'stats': stats,
            'monthly_trend': monthly_trend,
            'rncs_count': len(rncs_raw),
            'rncs': [{
                'id': r[0],
                'rnc_number': r[1],
                'title': r[2],
                'equipment': r[3],
                'client': r[4],
                'priority': r[5],
                'status': r[6],
                'responsavel': r[7],
                'setor': r[8],
                'area_responsavel': r[9],
                'finalized_at': r[10],
                'created_at': r[11],
                'price': r[12]
            } for r in rncs_raw]
        })
        
    except Exception as e:
        try:
            logger.error(f"Erro na API de engenharia: {e}")
        except:
            print(f"Erro na API de engenharia: {e}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@app.route('/api/indicadores/setor')
def api_indicadores_setor():
    """API genérica para dados de qualquer setor"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    
    setor = request.args.get('setor', '').lower()
    
    # Mapeamento de setor para nome no banco
    setor_mapping = {
        'engenharia': 'Engenharia',
        'producao': 'Produção',
        'pcp': 'PCP',
        'qualidade': 'Qualidade',
        'compras': 'Compras',
        'comercial': 'Comercial',
        'terceiros': 'Terceiros'
    }
    
    # Mapeamento adicional para setores específicos de produção
    setor_producao_mapping = {
        'usinagem_plana': 'Usinagem Plana',
        'usin_cilindrica_cnc': 'Usin. Cilíndrica CNC',
        'usin_cilindrica_convencional': 'Usin. Cilíndrica Convencional',
        'caldeiraria_carbono': 'Caldeiraria de Carbono',
        'caldeiraria_inox': 'Caldeiraria de Inox',
        'corte': 'Corte',
        'montagem': 'Montagem',
        'pintura': 'Pintura',
        'balanceamento': 'Balanceamento'
    }
    
    setor_nome = setor_mapping.get(setor) or setor_producao_mapping.get(setor)
    if not setor_nome:
        return jsonify({'success': False, 'message': 'Setor inválido'}), 400
    
    try:
        conn = sqlite3.connect('ippel_system.db')
        cursor = conn.cursor()
        
        # Calcular data limite (últimos 12 meses)
        twelve_months_ago = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        
        # Buscar RNCs do setor - busca mais precisa, limitada aos últimos 12 meses
        if setor in ['usinagem_plana', 'usin_cilindrica_cnc', 'usin_cilindrica_convencional', 
                     'caldeiraria_carbono', 'caldeiraria_inox', 'corte', 'montagem', 'pintura', 'balanceamento']:
            # Para setores específicos de produção, buscar apenas no campo 'setor'
            cursor.execute("""
                SELECT 
                    id, rnc_number, title, description, client, equipment, area_responsavel,
                    setor, status, priority, finalized_at, created_at, price
                FROM rncs 
                WHERE LOWER(TRIM(setor)) LIKE ?
                AND (is_deleted = 0 OR is_deleted IS NULL)
                AND (
                    (finalized_at IS NOT NULL AND DATE(finalized_at) >= ?)
                    OR (finalized_at IS NULL AND DATE(created_at) >= ?)
                )
                ORDER BY COALESCE(finalized_at, created_at) DESC
            """, (f'%{setor_nome.lower()}%', twelve_months_ago, twelve_months_ago))
        else:
            # Para setores gerais, buscar em ambos os campos
            cursor.execute("""
                SELECT 
                    id, rnc_number, title, description, client, equipment, area_responsavel,
                    setor, status, priority, finalized_at, created_at, price
                FROM rncs 
                WHERE (
                    LOWER(TRIM(area_responsavel)) LIKE ?
                    OR LOWER(TRIM(setor)) LIKE ?
                )
                AND (is_deleted = 0 OR is_deleted IS NULL)
                AND (
                    (finalized_at IS NOT NULL AND DATE(finalized_at) >= ?)
                    OR (finalized_at IS NULL AND DATE(created_at) >= ?)
                )
                ORDER BY COALESCE(finalized_at, created_at) DESC
            """, (f'%{setor_nome.lower()}%', f'%{setor_nome.lower()}%', twelve_months_ago, twelve_months_ago))
        
        rncs_raw = cursor.fetchall()
        
        # Dados mensais para gráficos
        monthly_data = {}
        cumulative_data = {}
        total_value = 0
        
        for rnc in rncs_raw:
            finalized_at = rnc[10]
            created_at = rnc[11]
            price_str = rnc[12] or "0"
            
            # Converter preço
            try:
                if isinstance(price_str, str):
                    price_clean = price_str.replace('R$', '').replace(' ', '').replace(',', '.')
                    price = float(price_clean) if price_clean else 0
                else:
                    price = float(price_str) if price_str else 0
            except:
                price = 0
            
            # Filtrar apenas RNCs finalizadas
            if not finalized_at:
                continue
            
            total_value += price
            
            # Usar apenas finalized_at para RNCs finalizadas
            date_to_use = finalized_at
            if date_to_use and date_to_use != '':
                try:
                    if isinstance(date_to_use, str):
                        date_str = date_to_use.strip()
                        if ' ' in date_str:
                            date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            date = datetime.strptime(date_str, '%Y-%m-%d')
                    else:
                        date = date_to_use
                except Exception as parse_err:
                    try:
                        if created_at:
                            created_str = str(created_at).strip().split(' ')[0]
                            date = datetime.strptime(created_str, '%Y-%m-%d')
                        else:
                            continue
                    except:
                        continue
            else:
                continue
            
            month_key = date.strftime('%Y-%m')
            day = date.day
            
            if month_key not in monthly_data:
                monthly_data[month_key] = {'count': 0, 'value': 0, 'daily': {}}
            
            monthly_data[month_key]['count'] += 1
            monthly_data[month_key]['value'] += price
            
            if day not in monthly_data[month_key]['daily']:
                monthly_data[month_key]['daily'][day] = 0
            monthly_data[month_key]['daily'][day] += 1
        
        # Ordenar meses e calcular acumulado
        sorted_months = sorted(monthly_data.keys())
        accumulated_count = 0
        accumulated_value = 0
        
        monthly_trend = []
        for month in sorted_months:
            data = monthly_data[month]
            accumulated_count += data['count']
            accumulated_value += data['value']
            
            daily_details = [
                {'day': day, 'count': count}
                for day, count in sorted(data['daily'].items())
            ]
            
            monthly_trend.append({
                'month': month,
                'count': data['count'],
                'accumulated_count': accumulated_count,
                'value': round(data['value'], 2),
                'accumulated_value': round(accumulated_value, 2),
                'daily_details': daily_details
            })
        
        # Estatísticas gerais
        total_rncs = len(rncs_raw)
        finalized_rncs = sum(1 for r in rncs_raw if r[10])
        active_rncs = total_rncs - finalized_rncs
        avg_value = total_value / total_rncs if total_rncs > 0 else 0
        
        conn.close()
        
        return jsonify({
            'success': True,
            'setor': setor_nome,
            'rncs_count': total_rncs,
            'stats': {
                'total_rncs': total_rncs,
                'finalized_rncs': finalized_rncs,
                'active_rncs': active_rncs,
                'total_value': round(total_value, 2),
                'avg_value': round(avg_value, 2)
            },
            'monthly_trend': monthly_trend,
            'rncs': [{
                'id': r[0],
                'rnc_number': r[1],
                'title': r[2],
                'description': r[3],
                'client': r[4],
                'equipment': r[5],
                'area_responsavel': r[6],
                'setor': r[7],
                'status': r[8],
                'priority': r[9],
                'finalized_at': r[10],
                'created_at': r[11],
                'price': r[12]
            } for r in rncs_raw]
        })
        
    except Exception as e:
        try:
            logger.error(f"Erro na API de setor: {e}")
        except:
            print(f"Erro na API de setor: {e}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@app.route('/api/indicadores')
def api_indicadores():
    """API para dados dos indicadores (formato compatível com dashboard_improved.html)"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    
    try:
        conn = sqlite3.connect('ippel_system.db')
        cursor = conn.cursor()
        
        # Buscar dados das tabelas principais
        cursor.execute("SELECT COUNT(*) FROM rncs WHERE is_deleted = 0")
        total_rncs = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM rncs WHERE status = 'Pendente' AND is_deleted = 0")
        pendentes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM rncs WHERE finalized_at IS NOT NULL AND is_deleted = 0")
        finalizadas = cursor.fetchone()[0]
        
        # Dados por setor/departamento
        cursor.execute("""
            SELECT u.department as setor, COUNT(r.id) as total
            FROM rncs r
            LEFT JOIN users u ON r.user_id = u.id
            WHERE r.is_deleted = 0 AND u.department IS NOT NULL
            GROUP BY u.department
            ORDER BY total DESC
        """)
        setores_raw = cursor.fetchall()
        
        # Dados por prioridade
        cursor.execute("""
            SELECT priority, COUNT(*) as total
            FROM rncs 
            WHERE is_deleted = 0 AND priority IS NOT NULL
            GROUP BY priority
            ORDER BY total DESC
        """)
        prioridades_raw = cursor.fetchall()
        
        # Dados mensais para tendência (últimos 6 meses)
        monthly_data = []
        for i in range(6):
            date = datetime.now() - timedelta(days=30*i)
            month_key = date.strftime('%Y-%m')
            cursor.execute("""
                SELECT COUNT(*) FROM rncs 
                WHERE strftime('%Y-%m', created_at) = ? AND is_deleted = 0
            """, (month_key,))
            count = cursor.fetchone()[0]
            monthly_data.append({
                'mes': date.strftime('%b'),
                'total': count
            })
        
        monthly_data.reverse()  # Ordem cronológica
        
        # Eficiência por departamento
        efficiency_data = []
        for setor, total in setores_raw:
            cursor.execute("""
                SELECT COUNT(*) FROM rncs r
                LEFT JOIN users u ON r.user_id = u.id
                WHERE r.is_deleted = 0 AND u.department = ? AND r.finalized_at IS NOT NULL
            """, (setor,))
            finalizadas_setor = cursor.fetchone()[0]
            
            efficiency = round((finalizadas_setor / max(total, 1)) * 100, 1)
            efficiency_data.append({
                'setor': setor,
                'eficiencia': efficiency,
                'meta': 85.0,  # Meta padrão
                'realizado': efficiency
            })
        
        conn.close()
        
        # Formatar dados compatíveis com o dashboard
        # Converter dados de eficiência em formato de departamentos esperado pelo dashboard
        departments_data = []
        for item in efficiency_data:
            departments_data.append({
                'department': item['setor'],
                'meta': item['meta'],
                'realizado': item['realizado'],
                'efficiency': item['eficiencia']
            })
        
        # Se não há dados de departamentos, usar dados fictícios
        if not departments_data:
            departments_data = [
                {'department': 'PRODUÇÃO', 'meta': 80, 'realizado': 60, 'efficiency': 75.0},
                {'department': 'ENGENHARIA', 'meta': 70, 'realizado': 55, 'efficiency': 78.6},
                {'department': 'QUALIDADE', 'meta': 50, 'realizado': 35, 'efficiency': 70.0}
            ]
        
        result = {
            'success': True,
            'kpis': {
                'total_rncs': total_rncs,
                'total_metas': total_rncs,  # Assumindo que todas as RNCs são metas
                'active_departments': len(setores_raw) if setores_raw else 3,
                'overall_efficiency': round((finalizadas / max(total_rncs, 1)) * 100, 1),
                'avg_rncs_per_dept': round(total_rncs / max(len(setores_raw), 1), 1) if setores_raw else 0
            },
            'totals': {
                'total': total_rncs,
                'pendentes': pendentes,
                'finalizadas': finalizadas,
                'resolvidas': finalizadas
            },
            'departments': departments_data,  # Nome correto esperado pelo dashboard
            'monthly_trends': monthly_data,   # Nome correto esperado pelo dashboard
            'setores': [{'setor': row[0], 'total': row[1]} for row in setores_raw] if setores_raw else [{'setor': 'Geral', 'total': total_rncs}],
            'prioridades': [{'prioridade': row[0], 'total': row[1]} for row in prioridades_raw] if prioridades_raw else [{'prioridade': 'Média', 'total': total_rncs}],
            'tendencia': monthly_data,
            'eficiencia_departamentos': efficiency_data,
            'eficiencia': round((finalizadas / max(total_rncs, 1)) * 100, 1)
        }
        
        # Aplicar formatação aos dados se o módulo estiver disponível
        if HAS_FORMATTING:
            result = format_data_for_dashboard(result)
        
        print(f" API Indicadores retornando: {result}")
        return jsonify(result)
        
    except Exception as e:
        print(f" Erro na API de indicadores: {e}")
        import traceback
        traceback.print_exc()
        
        # Dados de fallback em caso de erro
        return jsonify({
            'success': True,
            'kpis': {
                'total_rncs': 0,
                'total_metas': 0,
                'active_departments': 3,
                'overall_efficiency': 0,
                'avg_rncs_per_dept': 0
            },
            'totals': {'total': 0, 'pendentes': 0, 'finalizadas': 0, 'resolvidas': 0},
            'departments': [
                {'department': 'PRODUÇÃO', 'meta': 80, 'realizado': 0, 'efficiency': 0},
                {'department': 'ENGENHARIA', 'meta': 70, 'realizado': 0, 'efficiency': 0},
                {'department': 'QUALIDADE', 'meta': 50, 'realizado': 0, 'efficiency': 0}
            ],
            'monthly_trends': [
                {'mes': 'Jan', 'total': 0},
                {'mes': 'Fev', 'total': 0},
                {'mes': 'Mar', 'total': 0},
                {'mes': 'Abr', 'total': 0},
                {'mes': 'Mai', 'total': 0},
                {'mes': 'Jun', 'total': 0}
            ],
            'setores': [{'setor': 'Geral', 'total': 0}],
            'prioridades': [{'prioridade': 'Média', 'total': 0}],
            'tendencia': [],
            'eficiencia_departamentos': [],
            'eficiencia': 0
        })

@app.route('/dashboard/api/kpis')
def dashboard_api_kpis():
    """API de KPIs para o dashboard melhorado"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'}), 401
    
    try:
        conn = sqlite3.connect('ippel_system.db')
        cursor = conn.cursor()
        
        # KPIs básicos
        cursor.execute("SELECT COUNT(*) FROM rncs WHERE is_deleted = 0")
        total_rncs = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM rncs WHERE status = 'Pendente' AND is_deleted = 0")
        pendentes = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM rncs WHERE finalized_at IS NOT NULL AND is_deleted = 0")
        finalizadas = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(DISTINCT u.department) FROM rncs r LEFT JOIN users u ON r.user_id = u.id WHERE r.is_deleted = 0 AND u.department IS NOT NULL")
        departamentos_ativos = cursor.fetchone()[0]
        
        # Eficiência geral
        eficiencia_geral = round((finalizadas / max(total_rncs, 1)) * 100, 1)
        
        conn.close()
        
        kpis_data = {
            'success': True,
            'kpis': {
                'total_rncs': total_rncs,
                'pendentes': pendentes,
                'finalizadas': finalizadas,
                'departamentos_ativos': departamentos_ativos,
                'eficiencia_geral': eficiencia_geral
            }
        }
        
        # Aplicar formatação aos dados se o módulo estiver disponível
        if HAS_FORMATTING:
            kpis_data = format_data_for_dashboard(kpis_data)
        
        return jsonify(kpis_data)
        
    except Exception as e:
        print(f" Erro na API de KPIs: {e}")
        return jsonify({
            'success': True,
            'kpis': {
                'total_rncs': 0,
                'pendentes': 0,
                'finalizadas': 0,
                'departamentos_ativos': 0,
                'eficiencia_geral': 0
            }
        })

@app.route('/form')
def form():
    """Formulário RNC (apenas para administradores)"""
    if 'user_id' not in session:
        return redirect('/')
    
    # Verificar permissão correta para criar RNCs
    if not has_permission(session['user_id'], 'create_rnc'):
        return render_template('error.html', 
                             error_title='Acesso Negado', 
                             error_message='Você não tem permissão para criar RNCs.'), 403
    
    # Buscar clientes cadastrados na tabela clients
    try:
        ensure_clients_table()
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT name FROM clients ORDER BY name')
        clients = [row[0] for row in cursor.fetchall()]
        conn.close()
    except Exception as e:
        logger.error(f"Erro ao carregar clientes para formulário: {e}")
        clients = []
    
    return render_template('new_rnc.html', clients=clients)

    # rotas /api/login e /api/logout movidas para routes/auth.py (Blueprint)

@app.route('/api/employee-performance')
def get_employee_performance():
    """API para obter desempenho por funcionário - CORRIGIDA"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    conn = None
    try:
        year = request.args.get('year', '')
        month = request.args.get('month', '')
        print(f" API chamada - Ano: {year or 'Todos'}, Mês: {month or 'Todos'}")

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # NÃO buscar usuários da tabela users - usar apenas as assinaturas das RNCs
        print(f" Usando apenas assinaturas das RNCs (não usuários da tabela)")

        # Query corrigida para buscar RNCs de TODOS os usuários
        # Primeiro, vamos verificar a estrutura da tabela
        cursor.execute("PRAGMA table_info(rncs)")
        cols = {row[1] for row in cursor.fetchall()}
        print(f" Colunas da tabela RNCs: {cols}")
        
        # Query corrigida - usar assinatura de engenharia (segunda assinatura)
        base_query = """
            SELECT 
                CASE 
                    WHEN r.signature_engineering_name IS NOT NULL AND r.signature_engineering_name != '' 
                    THEN r.signature_engineering_name
                    ELSE r.user_id
                END as owner_id,
                COUNT(r.id) as rnc_count
            FROM rncs r
            WHERE r.status IN ('Finalizado','finalized')
              AND r.is_deleted = 0
        """

        params = []
        if year and year.lower() != 'todos':
            base_query += " AND strftime('%Y', r.created_at) = ?"
            params.append(year)
            print(f" Filtro ano aplicado: {year}")
        if month and month.lower() != 'todos':
            base_query += " AND strftime('%m', r.created_at) = ?"
            params.append(month.zfill(2))
            print(f" Filtro mês aplicado: {month}")

        base_query += " GROUP BY owner_id"
        print(f" Executando query: {base_query}")
        print(f" Parâmetros: {params}")

        cursor.execute(base_query, params)
        rnc_rows = cursor.fetchall()
        rnc_data = {row[0]: row[1] for row in rnc_rows}
        print(f" RNCs por usuário: {rnc_data}")
        
        # Debug: verificar algumas RNCs para entender a distribuição
        cursor.execute("""
            SELECT r.user_id, u.name, COUNT(*) as count
            FROM rncs r
            LEFT JOIN users u ON r.user_id = u.id
            WHERE r.status IN ('Finalizado','finalized')
              AND r.is_deleted = 0
            GROUP BY r.user_id, u.name
            ORDER BY count DESC
            LIMIT 10
        """)
        debug_rncs = cursor.fetchall()
        print(f" Debug - Top 10 usuários por RNCs: {debug_rncs}")
        
        # Verificar se há RNCs com user_id NULL ou inválido
        cursor.execute("""
            SELECT COUNT(*) as total_rncs,
                   COUNT(CASE WHEN user_id IS NULL THEN 1 END) as null_user_id,
                   COUNT(CASE WHEN user_id = 1 THEN 1 END) as admin_rncs
            FROM rncs 
            WHERE status IN ('Finalizado','finalized')
              AND is_deleted = 0
        """)
        debug_stats = cursor.fetchone()
        print(f" Debug - Estatísticas: Total={debug_stats[0]}, NULL user_id={debug_stats[1]}, Admin RNCs={debug_stats[2]}")

        meta_mensal = 5
        result = []
        
        # Criar lista de assinaturas únicas (SEM remover espaços para manter compatibilidade com rnc_data)
        unique_signatures = set()
        for owner_id, count in rnc_rows:
            if isinstance(owner_id, str) and owner_id:
                unique_signatures.add(owner_id)  # SEM .strip() para manter compatibilidade
            elif isinstance(owner_id, int):
                # Se for ID, buscar nome do usuário
                cursor.execute("SELECT name FROM users WHERE id = ?", (owner_id,))
                user_result = cursor.fetchone()
                if user_result:
                    unique_signatures.add(user_result[0])
        
        print(f" Assinaturas únicas encontradas: {len(unique_signatures)}")
        print(f" Exemplos: {list(unique_signatures)[:5]}")
        
        # Processar cada assinatura
        for signature in sorted(unique_signatures):
            rncs = rnc_data.get(signature, 0)
            percentage = (rncs / meta_mensal * 100) if meta_mensal > 0 else 0
            status = 'Acima da Meta' if rncs >= meta_mensal else 'Abaixo da Meta'
            
            # Adicionar logs para debug
            print(f"   Processando assinatura: '{signature}' - RNCs: {rncs}")
            
            result.append({
                'id': signature,
                'name': signature,
                'rncs': rncs,
                'meta': meta_mensal,
                'percentage': round(percentage, 1),
                'status': status,
                'department': 'Engenharia'
            })

        result.sort(key=lambda x: x['percentage'], reverse=True)
        print(f" Resultado final: {len(result)} funcionários processados")
        for emp in result[:3]:
            print(f"   {emp['name']}: {emp['rncs']} RNCs ({emp['percentage']}%)")

        performance_data = {
            'success': True,
            'data': result,
            'filters': {
                'year': year or 'todos',
                'month': month or 'todos'
            }
        }
        
        # Aplicar formatação aos dados se o módulo estiver disponível
        if HAS_FORMATTING:
            performance_data = format_data_for_dashboard(performance_data)

        return jsonify(performance_data)
    except Exception as e:
        print(f" Erro ao buscar desempenho de funcionários: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar dados de funcionários: {str(e)}'
        }), 500
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route('/api/dashboard/performance')
def get_dashboard_performance():
    """API específica para dashboard de desempenho - SEM VERIFICAÇÃO DE PERMISSÃO"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    conn = None
    try:
        year = request.args.get('year', '')
        month = request.args.get('month', '')
        print(f" Dashboard API chamada - Ano: {year or 'Todos'}, Mês: {month or 'Todos'}")

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # NÃO buscar usuários da tabela users - usar apenas as assinaturas das RNCs
        print(f" Dashboard - Usando apenas assinaturas das RNCs (não usuários da tabela)")

        # Query para buscar RNCs finalizadas
        # Usar assinatura de engenharia (segunda assinatura)
        base_query = """
            SELECT 
                CASE 
                    WHEN r.signature_engineering_name IS NOT NULL AND r.signature_engineering_name != '' 
                    THEN r.signature_engineering_name
                    ELSE r.user_id
                END as owner_id,
                COUNT(r.id) as rnc_count
            FROM rncs r
            WHERE r.status IN ('Finalizado','finalized')
              AND r.is_deleted = 0
        """

        params = []
        if year and year.lower() != 'todos':
            base_query += " AND strftime('%Y', r.created_at) = ?"
            params.append(year)
        if month and month.lower() != 'todos':
            base_query += " AND strftime('%m', r.created_at) = ?"
            params.append(month.zfill(2))

        base_query += " GROUP BY owner_id"
        cursor.execute(base_query, params)
        rnc_rows = cursor.fetchall()
        rnc_data = {row[0]: row[1] for row in rnc_rows}
        
        # Debug: verificar algumas RNCs para entender a distribuição
        cursor.execute("""
            SELECT r.user_id, u.name, COUNT(*) as count
            FROM rncs r
            LEFT JOIN users u ON r.user_id = u.id
            WHERE r.status IN ('Finalizado','finalized')
              AND r.is_deleted = 0
            GROUP BY r.user_id, u.name
            ORDER BY count DESC
            LIMIT 10
        """)
        debug_rncs = cursor.fetchall()
        print(f" Dashboard Debug - Top 10 usuários por RNCs: {debug_rncs}")
        
        # Verificar se há RNCs com user_id NULL ou inválido
        cursor.execute("""
            SELECT COUNT(*) as total_rncs,
                   COUNT(CASE WHEN user_id IS NULL THEN 1 END) as null_user_id,
                   COUNT(CASE WHEN user_id = 1 THEN 1 END) as admin_rncs
            FROM rncs 
            WHERE status IN ('Finalizado','finalized')
              AND is_deleted = 0
        """)
        debug_stats = cursor.fetchone()
        print(f" Dashboard Debug - Estatísticas: Total={debug_stats[0]}, NULL user_id={debug_stats[1]}, Admin RNCs={debug_stats[2]}")

        meta_mensal = 5
        result = []
        
        # Criar lista de assinaturas únicas (SEM remover espaços para manter compatibilidade com rnc_data)
        unique_signatures = set()
        for owner_id, count in rnc_rows:
            if isinstance(owner_id, str) and owner_id:
                unique_signatures.add(owner_id)  # SEM .strip() para manter compatibilidade
            elif isinstance(owner_id, int):
                # Se for ID, buscar nome do usuário
                cursor.execute("SELECT name FROM users WHERE id = ?", (owner_id,))
                user_result = cursor.fetchone()
                if user_result:
                    unique_signatures.add(user_result[0])
        
        print(f" Dashboard - Assinaturas únicas encontradas: {len(unique_signatures)}")
        print(f" Dashboard - Exemplos: {list(unique_signatures)[:5]}")
        
        # Processar cada assinatura
        for signature in sorted(unique_signatures):
            rncs = rnc_data.get(signature, 0)
            percentage = (rncs / meta_mensal * 100) if meta_mensal > 0 else 0
            status = 'Acima da Meta' if rncs >= meta_mensal else 'Abaixo da Meta'
            
            # Adicionar logs para debug
            print(f"   Dashboard - Processando assinatura: '{signature}' - RNCs: {rncs}")
            
            result.append({
                'id': signature,
                'name': signature,
                'rncs': rncs,
                'meta': meta_mensal,
                'percentage': round(percentage, 1),
                'status': status,
                'department': 'Engenharia'
            })

        result.sort(key=lambda x: x['percentage'], reverse=True)
        print(f" Dashboard: {len(result)} funcionários processados")

        dashboard_data = {
            'success': True,
            'data': result,
            'filters': {
                'year': year or 'todos',
                'month': month or 'todos'
            }
        }
        
        # Aplicar formatação aos dados se o módulo estiver disponível
        if HAS_FORMATTING:
            dashboard_data = format_data_for_dashboard(dashboard_data)

        return jsonify(dashboard_data)
    except Exception as e:
        print(f" Erro no dashboard: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar dados do dashboard: {str(e)}'
        }), 500
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route('/api/user/info')
def get_user_info():
    """API para obter informações do usuário logado"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        # Buscar dados básicos do usuário
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT u.permissions, g.name as group_name, u.avatar_key, u.avatar_prefs, u.group_id
            FROM users u
            LEFT JOIN groups g ON u.group_id = g.id
            WHERE u.id = ?
        ''', (session['user_id'],))
        user_data = cursor.fetchone()
        
        # Buscar permissões específicas do grupo
        cursor.execute('''
            SELECT gp.permission_name, gp.permission_value
            FROM group_permissions gp
            JOIN users u ON u.group_id = gp.group_id
            WHERE u.id = ? AND gp.permission_value = 1
        ''', (session['user_id'],))
        group_permissions = cursor.fetchall()
        conn.close()
        
        # Permissions do usuário (compatibilidade)
        permissions = []
        if user_data and user_data[0]:
            try:
                # Tentar parsear como JSON
                permissions = json.loads(user_data[0])
            except json.JSONDecodeError:
                # Se falhar, assumir que é string separada por vírgulas
                permissions = [p.strip() for p in user_data[0].split(',')]
        
        # Adicionar permissões do grupo
        group_perms_list = [perm[0] for perm in group_permissions]
        
        # Adicionar permissões baseadas em departamento
        department_permissions = {
            'canViewAllRncs': has_permission(session['user_id'], 'view_all_rncs'),
            'canViewOwnRncs': has_permission(session['user_id'], 'view_own_rncs'),
            'canViewFinalizedRncs': has_permission(session['user_id'], 'view_finalized_rncs'),
            'canViewCharts': has_permission(session['user_id'], 'view_charts'),
            'canViewReports': has_permission(session['user_id'], 'view_reports'),
            'hasAdminAccess': has_permission(session['user_id'], 'admin_access'),
            'canManageUsers': has_permission(session['user_id'], 'manage_users'),
            'canEditRncs': has_permission(session['user_id'], 'edit_all_rncs'),
            'canEditOwnRncs': has_permission(session['user_id'], 'edit_own_rnc'),
            'canCreateRnc': has_permission(session['user_id'], 'create_rnc'),
            'canAssignRncToGroup': has_permission(session['user_id'], 'assign_rnc_to_group'),
            'canDeleteRncs': has_permission(session['user_id'], 'delete_rncs'),
            'canReplyRncs': has_permission(session['user_id'], 'reply_rncs'),
            'canFinalizeRnc': has_permission(session['user_id'], 'finalize_rnc'),
            'canViewGroupsForAssignment': True,  # Todos podem ver grupos para atribuição
            'canViewUsersForAssignment': True    # Todos podem ver usuários para atribuição
        }
        
        # Preparar preferências de avatar
        avatar_prefs = None
        try:
            if user_data and len(user_data) > 3 and user_data[3]:
                avatar_prefs = json.loads(user_data[3])
        except Exception:
            avatar_prefs = None

        user_info = {
            'id': session['user_id'],
            'name': session['user_name'],
            'email': session['user_email'],
            'department': session.get('user_department') or get_user_department(session['user_id']),
            'group': user_data[1] if user_data else None,
            'group_id': user_data[4] if user_data and len(user_data) > 4 else None,
            'role': session['user_role'],
            'permissions': permissions,
            'groupPermissions': group_perms_list,
            'departmentPermissions': department_permissions,
            'avatar': (user_data[2] if user_data and len(user_data) > 2 else None),
            'avatarPrefs': avatar_prefs
        }
        
        logger.info(f"Informações do usuário: {user_info}")
        
        resp = jsonify({'success': True, 'user': user_info})
        resp.headers['Cache-Control'] = 'no-store'
        return resp
    except Exception as e:
        logger.error(f"Erro ao buscar informações do usuário: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro ao buscar informações do usuário'
        }), 500

    # rota /api/user/avatar movida para routes/api.py (Blueprint)

## Rota movida: /api/rnc/create agora está em routes/rnc.py (Blueprint)
## Rota movida: /api/rnc/<id>/update (variante 1) agora está em routes/rnc.py

    except Exception as e:
        logger.error(f"Erro ao criar RNC: {e}")
        try:
            return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500
        except Exception:
            # fallback caso jsonify falhe
            return ('{"success": false, "message": "Erro interno"}', 500, {'Content-Type': 'application/json'})
## Rota movida: /api/rnc/list agora está em routes/rnc.py
    """API para listar RNCs por abas ('active' e 'finalized').

    Observação: o sistema de lixeira foi removido; não há mais aba 'deleted'.
    """
    if 'user_id' not in session:
        return jsonify({
            'success': False,
            'message': 'Usuário não autenticado'
        }), 401
    
    conn = None
    try:
        # Obter parâmetros de filtro
        tab = request.args.get('tab', 'active')  # active, finalized
        user_id = session['user_id']
        
        # Verificar se deve ignorar cache (parâmetro _t indica force refresh)
        force_refresh = request.args.get('_t') is not None
        
        # Cache key baseada no usuário e aba
        cache_key = f"rncs_list_{user_id}_{tab}"
        if not force_refresh:
            cached_result = get_cached_query(cache_key)
            if cached_result:
                logger.info(f"Retornando cache para {cache_key}")
                return jsonify(cached_result)
        else:
            logger.info(f"Force refresh solicitado - ignorando cache para {cache_key}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Construir query otimizada - incluindo campos de atribuição
        base_query = '''
                SELECT 
                    r.id,
                    r.rnc_number,
                    r.title,
                    r.equipment,
                    r.client,
                    r.priority,
                    r.status,
                    r.user_id,
                    r.assigned_user_id,
                    r.created_at,
                    r.updated_at,
                    r.finalized_at,
                    u.name AS user_name,
                    u.department AS user_department,
                    au.name AS assigned_user_name
                FROM rncs r 
                LEFT JOIN users u ON r.user_id = u.id
                LEFT JOIN users au ON r.assigned_user_id = au.id
        '''
        
        # Aplicar filtros baseados na aba - SEM LIMITES para mostrar todos os RNCs
        if tab == 'active':
            # RNCs ativos - mostrar todos
            if has_permission(session['user_id'], 'view_all_rncs'):
                cursor.execute(base_query + '''
                    WHERE (r.is_deleted = 0 OR r.is_deleted IS NULL) 
                    AND r.status NOT IN ('Finalizado') 
                    ORDER BY r.id DESC
                ''')
            else:
                # Query para usuário específico - RNCs criados, atribuídos OU compartilhados via rnc_shares
                cursor.execute('''
                    SELECT DISTINCT
                        r.id,
                        r.rnc_number,
                        r.title,
                        r.equipment,
                        r.client,
                        r.priority,
                        r.status,
                        r.user_id,
                        r.assigned_user_id,
                        r.created_at,
                        r.updated_at,
                        r.finalized_at,
                        u.name AS user_name,
                        u.department AS user_department,
                        au.name AS assigned_user_name
                    FROM rncs r 
                    LEFT JOIN users u ON r.user_id = u.id
                    LEFT JOIN users au ON r.assigned_user_id = au.id
                    LEFT JOIN rnc_shares rs ON rs.rnc_id = r.id
                    WHERE (r.is_deleted = 0 OR r.is_deleted IS NULL) 
                    AND r.status NOT IN ('Finalizado') 
                    AND (r.user_id = ? OR r.assigned_user_id = ? OR rs.shared_with_user_id = ?)
                    ORDER BY r.id DESC
                ''', (session['user_id'], session['user_id'], session['user_id']))

        elif tab == 'finalized':
            # RNCs finalizados - mostrar todos
            if has_permission(session['user_id'], 'view_finalized_rncs'):
                cursor.execute(base_query + '''
                    WHERE (r.is_deleted = 0 OR r.is_deleted IS NULL) 
                    AND r.status = 'Finalizado'
                    ORDER BY r.id DESC
                ''')
            else:
                # RNCs finalizados - criados, atribuídos OU compartilhados via rnc_shares
                cursor.execute('''
                    SELECT DISTINCT
                        r.id,
                        r.rnc_number,
                        r.title,
                        r.equipment,
                        r.client,
                        r.priority,
                        r.status,
                        r.user_id,
                        r.assigned_user_id,
                        r.created_at,
                        r.updated_at,
                        r.finalized_at,
                        u.name AS user_name,
                        u.department AS user_department,
                        au.name AS assigned_user_name
                    FROM rncs r 
                    LEFT JOIN users u ON r.user_id = u.id
                    LEFT JOIN users au ON r.assigned_user_id = au.id
                    LEFT JOIN rnc_shares rs ON rs.rnc_id = r.id
                    WHERE (r.is_deleted = 0 OR r.is_deleted IS NULL) 
                    AND r.status = 'Finalizado' 
                    AND (r.user_id = ? OR r.assigned_user_id = ? OR rs.shared_with_user_id = ?)
                    ORDER BY r.id DESC
                ''', (session['user_id'], session['user_id'], session['user_id']))

        else:
            # Fallback para aba ativa - todos os RNCs
            cursor.execute(base_query + '''
                WHERE (r.is_deleted = 0 OR r.is_deleted IS NULL) 
                AND r.status NOT IN ('Finalizado')
                ORDER BY r.id DESC
            ''')
        
        rncs = cursor.fetchall()
        logger.info(f" Query executada para {tab}: {len(rncs)} RNCs encontrados no banco")
        
        # Formatar dados
        formatted_rncs = []
        current_user_id = session['user_id']
        
        # Usar list comprehension para melhor performance
        formatted_rncs = [
            {
                'id': rnc[0],
                'rnc_number': rnc[1],
                'title': rnc[2],
                'equipment': rnc[3],
                'client': rnc[4],
                'priority': rnc[5],
                'status': rnc[6],
                'user_id': rnc[7],
                'assigned_user_id': rnc[8],
                'created_at': rnc[9],
                'updated_at': rnc[10],
                'finalized_at': rnc[11],
                'user_name': rnc[12],
                'user_department': rnc[13] or 'N/A',
                'assigned_user_name': rnc[14],
                'department': rnc[13] or 'N/A',
                'setor': rnc[13] or 'N/A',
                'is_creator': (current_user_id == rnc[7]),
                'is_assigned': (current_user_id == rnc[8])  # Novo campo para identificar se está atribuído
            }
            for rnc in rncs
        ]
        
        result = {
            'success': True,
            'rncs': formatted_rncs,
            'tab': tab
        }
        
        logger.info(f" Resultado final para {tab}: {len(formatted_rncs)} RNCs serão retornados")
        
        # Cache moderado - 2 minutos para balancear performance e atualização
        cache_query(cache_key, result, ttl=120)
        
        # Retornar com headers otimizados
        response = jsonify(result)
        response.headers['Cache-Control'] = 'public, max-age=120' if not force_refresh else 'no-cache'
        response.headers['X-Content-Type-Options'] = 'nosniff'
        return response
        
    except Exception as e:
        logger.error(f"Erro ao listar RNCs: {e}")
        return jsonify({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }), 500
    finally:
        if conn:
            return_db_connection(conn)

## Rota movida: /api/rnc/get/<id> agora está em routes/rnc.py
    """Retorna os dados completos de uma RNC para visualização/edição."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('PRAGMA table_info(rncs)')
        columns = [row[1] for row in cursor.fetchall()]
        cursor.execute('SELECT * FROM rncs WHERE id = ?', (rnc_id,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            return jsonify({'success': False, 'message': 'RNC não encontrada'}), 404
        rnc_dict = dict(zip(columns, row))
        # Normalizar checkboxes para booleano
        for key in rnc_dict:
            if key.startswith('disposition_') or key.startswith('inspection_'):
                rnc_dict[key] = bool(rnc_dict[key])
        return jsonify({'success': True, 'rnc': rnc_dict})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

## Rota movida: /rnc/<id> agora está em routes/rnc.py
    """Visualizar RNC específico"""
    if 'user_id' not in session:
        return redirect('/')
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Buscar RNC com informações do usuário, dados de disposição/inspeção e assinaturas
        cursor.execute('''
            SELECT r.*, u.name as user_name, au.name as assigned_user_name
            FROM rncs r 
            LEFT JOIN users u ON r.user_id = u.id 
            LEFT JOIN users au ON r.assigned_user_id = au.id
            WHERE r.id = ?
        ''', (rnc_id,))
        
        rnc_data = cursor.fetchone()
        conn.close()
        
        if not rnc_data:
            return render_template('error.html', message='RNC não encontrado')
        
        # Verificar se rnc_data é uma tupla/lista antes de acessar índices
        if not isinstance(rnc_data, (tuple, list)):
            logger.error(f"Erro: rnc_data não é uma tupla/lista: {type(rnc_data)} - {rnc_data}")
            return render_template('error.html', message='Erro interno do sistema')
        
        # Verificar permissão de acesso
        is_creator = (str(rnc_data[8]) == str(session['user_id']))  # user_id está na posição 8
        is_finalized = (rnc_data[7] == 'Finalizado')  # status está na posição 7
        
        logger.info(f"Verificando acesso do usuário {session['user_id']} à RNC {rnc_id}")
        logger.info(f"É criador: {is_creator}, Está finalizado: {is_finalized}")
        
        if not can_user_access_rnc(session['user_id'], rnc_id):
            logger.warning(f"Acesso negado para usuário {session['user_id']} à RNC {rnc_id}")
            return render_template('error.html', message='Acesso negado')
        
        # Converter para dicionário segundo a ordem real das colunas da tabela 'rncs'
        # Consultamos PRAGMA para garantir alinhamento (inclui 'price' e assinaturas na posição correta)
        try:
            conn_cols = sqlite3.connect(DB_PATH)
            cur_cols = conn_cols.cursor()
            cur_cols.execute('PRAGMA table_info(rncs)')
            base_columns = [row[1] for row in cur_cols.fetchall()]
            conn_cols.close()
        except Exception:
            base_columns = [
                'id','rnc_number','title','description','equipment','client','priority','status','user_id','assigned_user_id',
                'is_deleted','deleted_at','finalized_at','created_at','updated_at','disposition_usar','disposition_retrabalhar',
                'disposition_rejeitar','disposition_sucata','disposition_devolver_estoque','disposition_devolver_fornecedor',
                'inspection_aprovado','inspection_reprovado','inspection_ver_rnc','signature_inspection_date','signature_engineering_date',
                'signature_inspection2_date','signature_inspection_name','signature_engineering_name','signature_inspection2_name','price',
                'department','instruction_retrabalho','cause_rnc','action_rnc'
            ]
        columns = base_columns + ['user_name', 'assigned_user_name']
        
        # Ajustar o número de colunas baseado na estrutura atual
        if len(rnc_data) < len(columns):
            # Adicionar valores padrão para colunas que podem não existir
            rnc_data = list(rnc_data) + [None] * (len(columns) - len(rnc_data))
        
        rnc_dict = dict(zip(columns, rnc_data))

        # Extrair campos rotulados do description (provenientes do TXT)
        def parse_label_map(text: str):
            if not text:
                return {}
            mapping = {}
            lines = [ln.strip() for ln in str(text).split('\n') if ln.strip()]
            for ln in lines:
                if ':' in ln:
                    parts = ln.split(':', 1)
                    if len(parts) == 2:
                        label = parts[0].strip()
                        val = parts[1].strip()
                        normalized_label = label.lower().replace(' ', '').replace('ã', 'a').replace('ç', 'c')
                        if 'desenho' in normalized_label:
                            mapping['Desenho'] = val
                        elif 'mp' in normalized_label:
                            mapping['MP'] = val
                        elif 'revisao' in normalized_label or 'revisão' in normalized_label:
                            mapping['Revisão'] = val
                        elif 'cv' in normalized_label:
                            mapping['CV'] = val
                        elif 'pos' in normalized_label:
                            mapping['POS'] = val
                        elif 'conjunto' in normalized_label:
                            mapping['Conjunto'] = val
                        elif 'modelo' in normalized_label:
                            mapping['Modelo'] = val
                        elif 'quantidade' in normalized_label:
                            mapping['Quantidade'] = val
                        elif 'area' in normalized_label and 'responsavel' in normalized_label:
                            mapping['Área responsável'] = val
                        elif 'descricao' in normalized_label and 'rnc' in normalized_label:
                            mapping['Descrição da RNC'] = val
                        elif 'instrucao' in normalized_label and 'retrabalho' in normalized_label:
                            mapping['Instrução para retrabalho'] = val
                        elif 'valor' in normalized_label:
                            mapping['Valor'] = val
                        else:
                            mapping[label] = val
            return mapping

        txt_fields = parse_label_map(rnc_dict.get('description') or '')
        
        # Verificar se o usuário é o criador do RNC
        is_creator = (session['user_id'] == rnc_data[8])  # user_id do RNC
        
        # Nova visualização completa, preservando estilo das abas do formulário
        return render_template('view_rnc_full.html', rnc=rnc_dict, is_creator=is_creator, txt_fields=txt_fields)
        
    except Exception as e:
        logger.error(f"Erro ao visualizar RNC {rnc_id}: {e}")
        return render_template('error.html', message='Erro interno do sistema')

## Rota movida: /rnc/<id>/reply agora está em routes/rnc.py
    """Abrir modo Responder (edição simplificada) para a RNC.

    - Permite criador, admin ou quem tenha 'reply_rncs'
    - Reaproveita o formulário de edição, sinalizando is_reply=True (UI pode ajustar rótulos)
    """
    if 'user_id' not in session:
        return redirect('/')

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT r.*, u.name as user_name, au.name as assigned_user_name
              FROM rncs r
              LEFT JOIN users u ON r.user_id = u.id
              LEFT JOIN users au ON r.assigned_user_id = au.id
             WHERE r.id = ?
        ''', (rnc_id,))
        rnc_data = cursor.fetchone()
        conn.close()

        if not rnc_data:
            return render_template('error.html', message='RNC não encontrado')

        # Verificar permissão para responder
        owner_id = rnc_data[8]  # user_id
        is_creator = str(session['user_id']) == str(owner_id)
        is_admin = has_permission(session['user_id'], 'admin_access')
        can_reply = has_permission(session['user_id'], 'reply_rncs')
        if not (is_creator or is_admin or can_reply):
            return render_template('error.html', message='Acesso negado: você não tem permissão para responder este RNC')

        # Mapear colunas dinamicamente conforme a estrutura real
        try:
            conn_cols = sqlite3.connect(DB_PATH)
            cur_cols = conn_cols.cursor()
            cur_cols.execute('PRAGMA table_info(rncs)')
            base_columns = [row[1] for row in cur_cols.fetchall()]
            conn_cols.close()
        except Exception:
            base_columns = [
                'id','rnc_number','title','description','equipment','client','priority','status','user_id','assigned_user_id',
                'is_deleted','deleted_at','finalized_at','created_at','updated_at','disposition_usar','disposition_retrabalhar',
                'disposition_rejeitar','disposition_sucata','disposition_devolver_estoque','disposition_devolver_fornecedor',
                'inspection_aprovado','inspection_reprovado','inspection_ver_rnc','signature_inspection_date','signature_engineering_date',
                'signature_inspection2_date','signature_inspection_name','signature_engineering_name','signature_inspection2_name','price',
                'department','instruction_retrabalho','cause_rnc','action_rnc'
            ]
        columns = base_columns + ['user_name', 'assigned_user_name']

        if len(rnc_data) < len(columns):
            rnc_data = list(rnc_data) + [None] * (len(columns) - len(rnc_data))

        rnc_dict = dict(zip(columns, rnc_data))

        # Sinalizar modo de resposta para o template (pode ajustar labels e lógica)
        return render_template('edit_rnc_form.html', rnc=rnc_dict, is_editing=True, is_reply=True)
    except Exception as e:
        logger.error(f"Erro ao abrir modo Responder para RNC {rnc_id}: {e}")
        return render_template('error.html', message='Erro interno do sistema')

## Rota movida: /rnc/<id>/print agora está em routes/rnc.py
    """Página de impressão do RNC com template idêntico à visualização"""
    if 'user_id' not in session:
        return redirect('/')
    
    try:
        # Usar função segura para buscar dados do RNC
        rnc_data, error_message = get_rnc_data_safe(rnc_id)
        
        if rnc_data is None:
            logger.error(f"Erro ao buscar RNC {rnc_id}: {error_message}")
            return render_template('error.html', message=error_message)
        
        # Reexecutar a consulta para capturar os nomes das colunas e mapear corretamente,
        # evitando erros de índice quando novas colunas são adicionadas (ex.: price)
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT r.*, u.name as user_name, au.name as assigned_user_name
            FROM rncs r 
            LEFT JOIN users u ON r.user_id = u.id 
            LEFT JOIN users au ON r.assigned_user_id = au.id
            WHERE r.id = ?
        ''', (rnc_id,))
        row = cursor.fetchone()
        columns = [d[0] for d in cursor.description]
        conn.close()

        rnc_dict = dict(zip(columns, row)) if row else {}
        # Normalizar datas para string simples (evita .strftime no template)
        for date_key in ['created_at', 'updated_at', 'finalized_at']:
            if isinstance(rnc_dict.get(date_key), (tuple, list)):
                rnc_dict[date_key] = rnc_dict.get(date_key)[0]

        # Garantias de campos esperados pelo template
        if 'price' not in rnc_dict:
            rnc_dict['price'] = 0
        if 'user_name' not in rnc_dict:
            rnc_dict['user_name'] = 'Sistema'
        
        # Extrair campos rotulados do description (provenientes do TXT)
        def parse_label_map(text: str):
            if not text:
                return {}
            mapping = {}
            lines = [ln.strip() for ln in str(text).split('\n') if ln.strip()]
            for ln in lines:
                if ':' in ln:
                    parts = ln.split(':', 1)
                    if len(parts) == 2:
                        label = parts[0].strip()
                        val = parts[1].strip()
                        normalized_label = label.lower().replace(' ', '').replace('ã', 'a').replace('ç', 'c')
                        if 'desenho' in normalized_label:
                            mapping['Desenho'] = val
                        elif 'mp' in normalized_label:
                            mapping['MP'] = val
                        elif 'revisao' in normalized_label or 'revisão' in normalized_label:
                            mapping['Revisão'] = val
                        elif 'cv' in normalized_label:
                            mapping['CV'] = val
                        elif 'pos' in normalized_label:
                            mapping['POS'] = val
                        elif 'conjunto' in normalized_label:
                            mapping['Conjunto'] = val
                        elif 'modelo' in normalized_label:
                            mapping['Modelo'] = val
                        elif 'quantidade' in normalized_label:
                            mapping['Quantidade'] = val
                        elif 'area' in normalized_label and 'responsavel' in normalized_label:
                            mapping['Área responsável'] = val
                        elif 'descricao' in normalized_label and 'rnc' in normalized_label:
                            mapping['Descrição da RNC'] = val
                        elif 'instrucao' in normalized_label and 'retrabalho' in normalized_label:
                            mapping['Instrução para retrabalho'] = val
                        elif 'valor' in normalized_label:
                            mapping['Valor'] = val
                        else:
                            mapping[label] = val
            return mapping
        
        # Parse dos campos de texto importados
        txt_fields = parse_label_map(rnc_dict.get('description') or '')
        
        return render_template('view_rnc_print.html', rnc=rnc_dict, txt_fields=txt_fields)
        
    except Exception as e:
        logger.error(f"Erro ao gerar página de impressão para RNC {rnc_id}: {e}")
        return render_template('error.html', message='Erro interno do sistema')

## Rota movida: /rnc/<id>/pdf-generator agora está em routes/rnc.py
    """Gerador de PDF com múltiplas opções"""
    if 'user_id' not in session:
        return redirect('/')
    
    try:
        logger.info(f"Tentando acessar gerador de PDF para RNC {rnc_id} para usuário {session['user_id']}")
        
        # Usar função segura para buscar dados do RNC
        rnc_data, error_message = get_rnc_data_safe(rnc_id)
        
        if rnc_data is None:
            logger.error(f"Erro ao buscar RNC {rnc_id}: {error_message}")
            return render_template('error.html', message=error_message)
        
        logger.info(f"RNC {rnc_id} encontrado com {len(rnc_data)} colunas")
        
        # Verificar permissão de acesso - usar índice correto para user_id
        user_id_index = 8  # user_id está na posição 8
        
        # Verificação adicional de segurança
        try:
            if len(rnc_data) <= user_id_index:
                logger.error(f"RNC {rnc_id} não tem dados suficientes: {len(rnc_data)} colunas")
                return render_template('error.html', message='Dados do RNC incompletos')
            
            user_id_from_rnc = rnc_data[user_id_index]
            logger.info(f"User ID do RNC: {user_id_from_rnc}")
            logger.info(f"User ID da sessão: {session['user_id']}")
            
            # Verificar se o usuário tem permissão para ver todos os RNCs ou se é o criador
            user_has_permission = has_permission(session['user_id'], 'view_all_rncs')
            is_creator = (user_id_from_rnc == session['user_id'])
            
            logger.info(f"Usuário tem permissão para ver todos: {user_has_permission}")
            logger.info(f"Usuário é criador: {is_creator}")
            
            if not user_has_permission and not is_creator:
                logger.warning(f"Usuário {session['user_id']} tentou acessar RNC {rnc_id} sem permissão")
                return render_template('error.html', message='Acesso negado')
                
        except Exception as access_error:
            logger.error(f"Erro ao verificar permissões para RNC {rnc_id}: {access_error}")
            return render_template('error.html', message='Erro ao verificar permissões')
        
        # Converter para dicionário incluindo os novos campos e assinaturas
        columns = [
            'id', 'rnc_number', 'title', 'description', 'equipment', 'client', 
            'priority', 'status', 'user_id', 'created_at', 'updated_at', 
            'assigned_user_id', 'disposition_usar', 'disposition_retrabalhar', 
            'disposition_rejeitar', 'disposition_sucata', 'disposition_devolver_estoque', 
            'disposition_devolver_fornecedor', 'inspection_aprovado', 'inspection_reprovado', 
            'inspection_ver_rnc', 'signature_inspection_date', 'signature_engineering_date', 
            'signature_inspection2_date', 'signature_inspection_name', 'signature_engineering_name', 
            'signature_inspection2_name', 'is_deleted', 'deleted_at', 'finalized_at',
            'user_name', 'assigned_user_name'
        ]
        
        # Ajustar o número de colunas baseado na estrutura atual
        if len(rnc_data) < len(columns):
            logger.info(f"Ajustando RNC {rnc_id}: {len(rnc_data)} colunas para {len(columns)}")
            # Adicionar valores padrão para colunas que podem não existir
            rnc_data = list(rnc_data) + [None] * (len(columns) - len(rnc_data))
        
        rnc_dict = dict(zip(columns, rnc_data))
        
        logger.info(f"RNC {rnc_id} preparado para gerador de PDF com sucesso")
        
        try:
            return render_template('view_rnc_pdf_js.html', rnc=rnc_dict)
        except Exception as template_error:
            logger.error(f"Erro ao renderizar template para RNC {rnc_id}: {template_error}")
            import traceback
            logger.error(f"Traceback do template: {traceback.format_exc()}")
            return render_template('error.html', message='Erro ao gerar página')
        
    except Exception as e:
        logger.error(f"Erro ao acessar gerador de PDF para RNC {rnc_id}: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return render_template('error.html', message='Erro interno do sistema')

## Rota movida: /rnc/<id>/edit agora está em routes/rnc.py
    """Editar RNC existente - Usa o mesmo template de criação"""
    if 'user_id' not in session:
        return redirect('/')
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Buscar RNC com assinaturas
        cursor.execute('''
            SELECT r.*, u.name as user_name, au.name as assigned_user_name
            FROM rncs r 
            LEFT JOIN users u ON r.user_id = u.id 
            LEFT JOIN users au ON r.assigned_user_id = au.id
            WHERE r.id = ?
        ''', (rnc_id,))
        
        rnc_data = cursor.fetchone()
        conn.close()
                
        if not rnc_data:
            return render_template('error.html', message='RNC não encontrado')
        
        # Verificar se rnc_data é uma tupla/lista antes de acessar índices
        if not isinstance(rnc_data, (tuple, list)):
            logger.error(f"Erro: rnc_data não é uma tupla/lista: {type(rnc_data)} - {rnc_data}")
            return render_template('error.html', message='Erro interno do sistema')
        
        # Verificar permissão de edição
        user_is_creator = rnc_data[8] == session['user_id']
        can_edit_all = has_permission(session['user_id'], 'edit_all_rncs')
        can_edit_own = has_permission(session['user_id'], 'edit_own_rnc')
        
        if not (can_edit_all or (can_edit_own and user_is_creator)):
            return render_template('error.html', message='Acesso negado: você não tem permissão para editar este RNC')
        
        # Mapear colunas dinamicamente conforme a estrutura real da tabela (inclui price e demais campos)
        try:
            conn_cols = sqlite3.connect(DB_PATH)
            cur_cols = conn_cols.cursor()
            cur_cols.execute('PRAGMA table_info(rncs)')
            base_columns = [row[1] for row in cur_cols.fetchall()]
            conn_cols.close()
        except Exception:
            # Fallback com conjunto amplo conhecido
            base_columns = [
                'id','rnc_number','title','description','equipment','client','priority','status','user_id','assigned_user_id',
                'is_deleted','deleted_at','finalized_at','created_at','updated_at','disposition_usar','disposition_retrabalhar',
                'disposition_rejeitar','disposition_sucata','disposition_devolver_estoque','disposition_devolver_fornecedor',
                'inspection_aprovado','inspection_reprovado','inspection_ver_rnc','signature_inspection_date','signature_engineering_date',
                'signature_inspection2_date','signature_inspection_name','signature_engineering_name','signature_inspection2_name','price',
                'department','instruction_retrabalho','cause_rnc','action_rnc'
            ]
        columns = base_columns + ['user_name', 'assigned_user_name']

        # Ajustar o tamanho para evitar perdas quando a estrutura tiver mais/menos colunas
        if len(rnc_data) < len(columns):
            rnc_data = list(rnc_data) + [None] * (len(columns) - len(rnc_data))
        
        rnc_dict = dict(zip(columns, rnc_data))
        
        # Retornar o mesmo template de criação, mas com dados preenchidos
        return render_template('edit_rnc_form.html', rnc=rnc_dict, is_editing=True)
        
    except Exception as e:
        logger.error(f"Erro ao editar RNC {rnc_id}: {e}")
        return render_template('error.html', message='Erro interno do sistema')
## Rota movida: /api/rnc/<id>/update (variante 2) agora está em routes/rnc.py
    """API para atualizar RNC"""
    logger.info(f"Iniciando atualização da RNC {rnc_id}")
    
    if 'user_id' not in session:
        logger.warning("Tentativa de atualização sem autenticação")
        return jsonify({
            'success': False,
            'message': 'Usuário não autenticado'
        }), 401
    
    try:
        # Verificar se RNC existe e capturar estado atual com nomes de colunas
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM rncs WHERE id = ?', (rnc_id,))
        rnc_data = cursor.fetchone()
        
        if not rnc_data:
            return jsonify({
                'success': False,
                'message': 'RNC não encontrado'
            }), 404
        
        # Verificar se rnc_data é uma tupla/lista antes de acessar índices
        if not isinstance(rnc_data, (tuple, list)):
            logger.error(f"Erro: rnc_data não é uma tupla/lista: {type(rnc_data)} - {rnc_data}")
            return jsonify({
                'success': False,
                'message': 'Erro interno do sistema'
            }), 500
        
        # Verificar permissão de edição
        user_is_creator = str(rnc_data[8]) == str(session['user_id'])  # rnc_data[8] é user_id
        can_edit_all = has_permission(session['user_id'], 'edit_all_rncs')
        can_edit_own = has_permission(session['user_id'], 'edit_own_rnc')
        has_admin = has_permission(session['user_id'], 'admin_access')
        department_match = False
        
        # Verificar se tem permissão para ver todos os departamentos
        if has_permission(session['user_id'], 'view_all_departments_rncs'):
            department_match = True
        # Verificar se é RNC do departamento da engenharia
        else:
            cursor.execute('SELECT department FROM rncs WHERE id = ?', (rnc_id,))
            rnc_dept = cursor.fetchone()
            if rnc_dept and rnc_dept[0] == 'Engenharia' and has_permission(session['user_id'], 'view_engineering_rncs'):
                department_match = True
        
        logger.info(f"Verificação de permissão para edição RNC {rnc_id}:")
        logger.info(f"  Usuário: {session['user_id']} ({session.get('user_name', 'N/A')})")
        logger.info(f"  Criador do RNC: {rnc_data[8]}")
        logger.info(f"  É criador: {user_is_creator}")
        logger.info(f"  Pode editar todos: {can_edit_all}")
        logger.info(f"  Pode editar próprios: {can_edit_own}")
        logger.info(f"  É admin: {has_admin}")
        logger.info(f"  Match departamento: {department_match}")
        
        # Permitir também usuários com permissão de resposta (reply_rncs)
        can_reply = has_permission(session['user_id'], 'reply_rncs')
        if not (can_edit_all or (can_edit_own and user_is_creator) or has_admin or department_match or can_reply):
            logger.warning(f"Acesso negado para edição do RNC {rnc_id}")
            return jsonify({
                'success': False,
                'message': 'Acesso negado: você não tem permissão para editar este RNC'
            }), 403
        
        data = request.get_json() or {}
        # Mapear colunas para dict atual
        try:
            cur_cols = conn.cursor()
            cur_cols.execute('PRAGMA table_info(rncs)')
            col_names = [row[1] for row in cur_cols.fetchall()]
        except Exception:
            col_names = []
        current = {}
        try:
            if col_names and isinstance(rnc_data, (tuple, list)):
                current = dict(zip(col_names, rnc_data))
        except Exception:
            current = {}

        def get_bool(key):
            v = data.get(key, current.get(key))
            if isinstance(v, bool):
                return 1 if v else 0
            try:
                return 1 if str(v).strip().lower() in ('1','true','on','yes','y') else 0
            except Exception:
                return 0
        logger.info(f"Dados recebidos para atualização do RNC {rnc_id}: {data}")
        
        # Processar assinaturas - permitir atualização se não estiver travada
        cursor.execute('SELECT signature_inspection_name, signature_engineering_name, signature_inspection2_name FROM rncs WHERE id = ?', (rnc_id,))
        current_sign = cursor.fetchone() or (None, None, None)
        
        # Log para debug
        logger.info(f" Assinaturas atuais: {current_sign}")
        logger.info(f" Assinaturas recebidas: {data.get('signature_inspection_name')}, {data.get('signature_engineering_name')}, {data.get('signature_inspection2_name')}")
        
        # Usar novas assinaturas se fornecidas, senão manter as atuais
        new_sign = (
            data.get('signature_inspection_name', current_sign[0] or ''),
            data.get('signature_engineering_name', current_sign[1] or ''),
            data.get('signature_inspection2_name', current_sign[2] or '')
        )
        
        logger.info(f" Assinaturas finais para salvar: {new_sign}")
        if not any(s and s != 'NOME' for s in new_sign):
            return jsonify({'success': False, 'message': 'É obrigatório preencher pelo menos uma assinatura!'}), 400
        
        logger.info(f"Executando UPDATE para RNC {rnc_id}")
        # Preparar campos adicionais (disposição/inspeção + seções textuais) com fallback ao valor atual
        disposition_usar = get_bool('disposition_usar')
        disposition_retrabalhar = get_bool('disposition_retrabalhar')
        disposition_rejeitar = get_bool('disposition_rejeitar')
        disposition_sucata = get_bool('disposition_sucata')
        disposition_devolver_estoque = get_bool('disposition_devolver_estoque')
        disposition_devolver_fornecedor = get_bool('disposition_devolver_fornecedor')
        inspection_aprovado = get_bool('inspection_aprovado')
        inspection_reprovado = get_bool('inspection_reprovado')
        inspection_ver_rnc = data.get('inspection_ver_rnc', current.get('inspection_ver_rnc', ''))
        instruction_retrabalho = data.get('instruction_retrabalho', current.get('instruction_retrabalho', ''))
        cause_rnc = data.get('cause_rnc', current.get('cause_rnc', ''))
        action_rnc = data.get('action_rnc', current.get('action_rnc', ''))

        cursor.execute('''
            UPDATE rncs 
            SET title = ?, description = ?, equipment = ?, client = ?, 
                priority = ?, status = ?, updated_at = CURRENT_TIMESTAMP,
                signature_inspection_name = ?, signature_engineering_name = ?, signature_inspection2_name = ?,
                signature_inspection_date = COALESCE(NULLIF(?, ''), signature_inspection_date),
                signature_engineering_date = COALESCE(NULLIF(?, ''), signature_engineering_date),
                signature_inspection2_date = COALESCE(NULLIF(?, ''), signature_inspection2_date),
                disposition_usar = ?, disposition_retrabalhar = ?, disposition_rejeitar = ?, disposition_sucata = ?,
                disposition_devolver_estoque = ?, disposition_devolver_fornecedor = ?,
                inspection_aprovado = ?, inspection_reprovado = ?, inspection_ver_rnc = ?,
                instruction_retrabalho = ?, cause_rnc = ?, action_rnc = ?
            WHERE id = ?
        ''', (
            data.get('title', current.get('title','')),
            data.get('description', current.get('description','')),
            data.get('equipment', current.get('equipment','')),
            data.get('client', current.get('client','')),
            data.get('priority', current.get('priority','Média')),
            data.get('status', current.get('status','Pendente')),
            new_sign[0],
            new_sign[1],
            new_sign[2],
            data.get('signature_inspection_date',''),
            data.get('signature_engineering_date',''),
            data.get('signature_inspection2_date',''),
            disposition_usar,
            disposition_retrabalhar,
            disposition_rejeitar,
            disposition_sucata,
            disposition_devolver_estoque,
            disposition_devolver_fornecedor,
            inspection_aprovado,
            inspection_reprovado,
            inspection_ver_rnc,
            instruction_retrabalho,
            cause_rnc,
            action_rnc,
            rnc_id
        ))
        
        affected_rows = cursor.rowcount
        logger.info(f"UPDATE executado. Linhas afetadas: {affected_rows}")
        
        conn.commit()
        conn.close()
        
        # Limpar cache para todos os usuários e invalidar caches do dashboard
        clear_rnc_cache()
        try:
            keys_to_remove = []
            with cache_lock:
                for key in list(query_cache.keys()):
                    if key.startswith('rncs_list_') or key.startswith('charts_'):
                        keys_to_remove.append(key)
                for key in keys_to_remove:
                    del query_cache[key]
        except Exception:
            pass
        logger.info(f"Cache limpo após atualização do RNC {rnc_id}")
        
        return jsonify({
            'success': True,
            'message': 'RNC atualizado com sucesso!',
            'affected_rows': affected_rows
        })
        
    except Exception as e:
        logger.error(f"Erro ao atualizar RNC {rnc_id}: {e}")
        return jsonify({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }), 500



## Rota movida: /api/rnc/<id>/finalize agora está em routes/rnc.py
    """API para finalizar RNC"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se o RNC existe e não está deletado
        cursor.execute('SELECT * FROM rncs WHERE id = ? AND is_deleted = 0', (rnc_id,))
        rnc = cursor.fetchone()
        
        if not rnc:
            conn.close()
            return jsonify({'success': False, 'message': 'RNC não encontrado'}), 404
        
        # Verificar se rnc é uma tupla/lista antes de acessar índices
        if not isinstance(rnc, (tuple, list)):
            logger.error(f"Erro: rnc não é uma tupla/lista: {type(rnc)} - {rnc}")
            conn.close()
            return jsonify({'success': False, 'message': 'Erro interno do sistema'}), 500
        
        # Verificar permissões
        user_id = session['user_id']
        cursor.execute('SELECT role FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        
        if not user:
            conn.close()
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
        
        user_role = user[0]
        rnc_creator_id = rnc[8]  # user_id do RNC (ajustado para nova estrutura)
        is_creator = (user_id == rnc_creator_id)
        
        # Apenas o criador ou admin pode finalizar o RNC
        if not is_creator and user_role != 'admin':
            conn.close()
            return jsonify({'success': False, 'message': 'Apenas o criador do RNC pode finalizá-lo'}), 403
        
        # Atualizar status para Finalizado com timestamp
        cursor.execute('''
            UPDATE rncs 
            SET status = 'Finalizado', finalized_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (rnc_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Erro ao finalizar RNC'}), 500
        
        conn.commit()
        conn.close()
        
        # Limpar cache de RNCs para todos os usuários
        clear_rnc_cache()
        
        return jsonify({
            'success': True,
            'message': 'RNC finalizado com sucesso!'
        })
        
    except Exception as e:
        logger.error(f"Erro ao finalizar RNC: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

## Rota movida: /api/rnc/<id>/reply agora está em routes/rnc.py
    """Reabrir/reenviar uma RNC para tratamento (status volta para 'Pendente').

    Regras:
    - Requer sessão ativa
    - Permitido ao criador, admin ou quem tenha permissão 'reply_rncs'
    - Define status='Pendente', finalized_at=NULL, updated_at=CURRENT_TIMESTAMP
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Buscar RNC
        cursor.execute('SELECT id, user_id, assigned_user_id, status FROM rncs WHERE id = ? AND is_deleted = 0', (rnc_id,))
        rnc = cursor.fetchone()
        if not rnc:
            conn.close()
            return jsonify({'success': False, 'message': 'RNC não encontrada'}), 404

        rnc_creator_id = rnc[1]
        user_id = session['user_id']
        is_creator = str(user_id) == str(rnc_creator_id)
        is_admin = has_permission(user_id, 'admin_access')
        can_reply = has_permission(user_id, 'reply_rncs')
        in_responder_group = user_in_group(user_id, 'RESPONDER RNC')

        if not (is_creator or is_admin or can_reply or in_responder_group):
            conn.close()
            return jsonify({'success': False, 'message': 'Sem permissão para responder esta RNC'}), 403

        # Reabrir RNC
        cursor.execute('''
            UPDATE rncs
               SET status = 'Pendente',
                   finalized_at = NULL,
                   updated_at = CURRENT_TIMESTAMP,
                   assigned_user_id = user_id -- atribui novamente ao criador
             WHERE id = ?
        ''', (rnc_id,))

        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'Nenhuma alteração realizada'}), 400

        conn.commit()
        conn.close()

        # Limpar cache
        clear_rnc_cache()

        return jsonify({'success': True, 'message': 'RNC reenviada com sucesso'})
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

## Mantido: /api/debug/session permanece neste módulo

@app.route('/api/private-chat/messages/<int:contact_id>')
def get_private_messages(contact_id):
    """API para obter mensagens privadas com um contato específico"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        limit = request.args.get('limit', type=int)
        user_id = session['user_id']
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        query = '''
            SELECT pm.id, pm.sender_id, pm.recipient_id, pm.message, pm.message_type,
                   pm.is_read, pm.created_at, u.name as user_name, u.department
            FROM private_messages pm
            JOIN users u ON pm.sender_id = u.id
            WHERE (pm.sender_id = ? AND pm.recipient_id = ?) 
               OR (pm.sender_id = ? AND pm.recipient_id = ?)
            ORDER BY pm.created_at DESC
        '''
        
        if limit:
            query += f' LIMIT {limit}'
            
        cursor.execute(query, (user_id, contact_id, contact_id, user_id))
        messages = cursor.fetchall()
        conn.close()
        
        # Converter para formato JSON
        messages_list = []
        for msg in messages:
            messages_list.append({
                'id': msg[0],
                'sender_id': msg[1],
                'recipient_id': msg[2],
                'message': msg[3],
                'message_type': msg[4],
                'is_read': msg[5],
                'created_at': msg[6],
                'user_name': msg[7],
                'department': msg[8],
                'user_id': msg[1]  # Para compatibilidade
            })
        
        # Inverter ordem se não for limit=1 (para mostrar mais antigas primeiro no chat)
        if not limit or limit > 1:
            messages_list.reverse()
        
        return jsonify({'success': True, 'messages': messages_list})
        
    except Exception as e:
        logger.error(f"Erro ao buscar mensagens privadas: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/private-chat/unread-count/<int:contact_id>')
def get_unread_count(contact_id):
    """API para obter contagem de mensagens não lidas de um contato"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        user_id = session['user_id']
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT COUNT(*) FROM private_messages
            WHERE sender_id = ? AND recipient_id = ? AND is_read = 0
        ''', (contact_id, user_id))
        
        count = cursor.fetchone()[0]
        conn.close()
        
        return jsonify({'success': True, 'count': count})
        
    except Exception as e:
        logger.error(f"Erro ao contar mensagens não lidas: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/cache/clear', methods=['POST'])
def clear_cache_api():
    """API para limpar cache do servidor"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        # LIMPEZA COMPLETA E FORÇADA DO CACHE
        user_id = session['user_id']
        is_admin = has_permission(user_id, 'admin_access')
        
        # Limpar TODOS os caches independentemente de permissão
        from services.cache import cache_lock, query_cache
        with cache_lock:
            cache_count = len(query_cache)
            query_cache.clear()
            logger.info(f" LIMPEZA FORÇADA: {cache_count} entradas de cache removidas")
        
        # Também limpar usando a função específica
        clear_rnc_cache()
        
        if is_admin:
            logger.info(f"Admin {user_id} executou limpeza completa do cache")
            message = f"Cache completo limpo com sucesso ({cache_count} entradas removidas)"
        else:
            logger.info(f"Usuário {user_id} executou limpeza de cache")
            message = f"Cache limpo com sucesso ({cache_count} entradas removidas)"
        
        return jsonify({
            'success': True,
            'message': message,
            'cache_cleared_count': cache_count
        })
        
    except Exception as e:
        logger.error(f"Erro ao limpar cache: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/debug/rnc-count')
def debug_rnc_count():
    """Rota de debug para verificar contagem real de RNCs"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Contar todos os RNCs
        cursor.execute('SELECT COUNT(*) FROM rncs')
        total = cursor.fetchone()[0]
        
        # Contar finalizados
        cursor.execute('SELECT COUNT(*) FROM rncs WHERE status = "Finalizado"')
        finalizados = cursor.fetchone()[0]
        
        # Contar finalizados sem deletados
        cursor.execute('SELECT COUNT(*) FROM rncs WHERE (is_deleted = 0 OR is_deleted IS NULL) AND status = "Finalizado"')
        finalizados_ativos = cursor.fetchone()[0]
        
        # Contar ativos
        cursor.execute('SELECT COUNT(*) FROM rncs WHERE (is_deleted = 0 OR is_deleted IS NULL) AND status != "Finalizado"')
        ativos = cursor.fetchone()[0]
        
        return_db_connection(conn)
        
        return jsonify({
            'success': True,
            'counts': {
                'total': total,
                'finalizados': finalizados,
                'finalizados_ativos': finalizados_ativos,
                'ativos': ativos
            }
        })
        
    except Exception as e:
        logger.error(f"Erro no debug: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@app.route('/api/debug/user-rncs')
def debug_user_rncs():
    """Debug para verificar RNCs do usuário atual"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        user_id = session['user_id']
        
        # RNCs criados pelo usuário
        cursor.execute('SELECT COUNT(*) FROM rncs WHERE user_id = ?', (user_id,))
        created_by_user = cursor.fetchone()[0]
        
        # RNCs atribuídos ao usuário
        cursor.execute('SELECT COUNT(*) FROM rncs WHERE assigned_user_id = ?', (user_id,))
        assigned_to_user = cursor.fetchone()[0]
        
        # RNCs ativos (criados OU atribuídos)
        cursor.execute('''SELECT COUNT(*) FROM rncs 
                         WHERE (is_deleted = 0 OR is_deleted IS NULL) 
                         AND status != "Finalizado" 
                         AND (user_id = ? OR assigned_user_id = ?)''', (user_id, user_id))
        active_total = cursor.fetchone()[0]
        
        # RNCs finalizados (criados OU atribuídos)
        cursor.execute('''SELECT COUNT(*) FROM rncs 
                         WHERE (is_deleted = 0 OR is_deleted IS NULL) 
                         AND status = "Finalizado" 
                         AND (user_id = ? OR assigned_user_id = ?)''', (user_id, user_id))
        finalized_total = cursor.fetchone()[0]
        
        # Obter alguns exemplos
        cursor.execute('''SELECT id, rnc_number, title, status, user_id, assigned_user_id 
                         FROM rncs 
                         WHERE (user_id = ? OR assigned_user_id = ?) 
                         ORDER BY id DESC LIMIT 5''', (user_id, user_id))
        examples = cursor.fetchall()
        
        return_db_connection(conn)
        
        return jsonify({
            'success': True,
            'user_id': user_id,
            'user_name': session.get('user_name', 'N/A'),
            'counts': {
                'created_by_user': created_by_user,
                'assigned_to_user': assigned_to_user,
                'active_total': active_total,
                'finalized_total': finalized_total
            },
            'examples': [
                {
                    'id': ex[0],
                    'rnc_number': ex[1],
                    'title': ex[2],
                    'status': ex[3],
                    'is_creator': ex[4] == user_id,
                    'is_assigned': ex[5] == user_id
                }
                for ex in examples
            ]
        })
        
    except Exception as e:
        logger.error(f"Erro no debug de usuário: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@app.route('/api/debug/user-shares')
def debug_user_shares():
    """Debug para verificar compartilhamentos do usuário atual"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        user_id = session['user_id']
        
        # RNCs compartilhadas COM o usuário
        cursor.execute('''
            SELECT rs.rnc_id, r.rnc_number, r.title, r.status, 
                   u.name as shared_by, rs.permission_level, rs.created_at
            FROM rnc_shares rs
            JOIN rncs r ON rs.rnc_id = r.id
            LEFT JOIN users u ON rs.shared_by_user_id = u.id
            WHERE rs.shared_with_user_id = ?
            ORDER BY rs.created_at DESC LIMIT 10
        ''', (user_id,))
        shared_with_user = cursor.fetchall()
        
        # RNCs compartilhadas PELO usuário
        cursor.execute('''
            SELECT rs.rnc_id, r.rnc_number, r.title, r.status, 
                   u.name as shared_with, rs.permission_level, rs.created_at
            FROM rnc_shares rs
            JOIN rncs r ON rs.rnc_id = r.id
            LEFT JOIN users u ON rs.shared_with_user_id = u.id
            WHERE rs.shared_by_user_id = ?
            ORDER BY rs.created_at DESC LIMIT 10
        ''', (user_id,))
        shared_by_user = cursor.fetchall()
        
        # Contar total de compartilhamentos
        cursor.execute('SELECT COUNT(*) FROM rnc_shares WHERE shared_with_user_id = ?', (user_id,))
        total_shared_with = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM rnc_shares WHERE shared_by_user_id = ?', (user_id,))
        total_shared_by = cursor.fetchone()[0]
        
        return_db_connection(conn)
        
        return jsonify({
            'success': True,
            'user_id': user_id,
            'user_name': session.get('user_name', 'N/A'),
            'totals': {
                'shared_with_me': total_shared_with,
                'shared_by_me': total_shared_by
            },
            'shared_with_me': [
                {
                    'rnc_id': share[0],
                    'rnc_number': share[1],
                    'title': share[2],
                    'status': share[3],
                    'shared_by': share[4],
                    'permission': share[5],
                    'shared_at': share[6]
                }
                for share in shared_with_user
            ],
            'shared_by_me': [
                {
                    'rnc_id': share[0],
                    'rnc_number': share[1],
                    'title': share[2],
                    'status': share[3],
                    'shared_with': share[4],
                    'permission': share[5],
                    'shared_at': share[6]
                }
                for share in shared_by_user
            ]
        })
        
    except Exception as e:
        logger.error(f"Erro no debug de compartilhamentos: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@app.route('/api/debug/rnc-signatures/<int:rnc_id>')
def debug_rnc_signatures(rnc_id):
    """Debug específico para verificar assinaturas de uma RNC"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Buscar todas as colunas de assinatura
        cursor.execute('''
            SELECT id, rnc_number, title,
                   signature_inspection_name, signature_engineering_name, signature_inspection2_name,
                   signature_inspection_date, signature_engineering_date, signature_inspection2_date
            FROM rncs 
            WHERE id = ?
        ''', (rnc_id,))
        
        rnc_data = cursor.fetchone()
        return_db_connection(conn)
        
        if not rnc_data:
            return jsonify({'success': False, 'message': 'RNC não encontrada'}), 404
        
        return jsonify({
            'success': True,
            'rnc_id': rnc_id,
            'rnc_number': rnc_data[1],
            'title': rnc_data[2],
            'signatures': {
                'inspection_name': rnc_data[3],
                'engineering_name': rnc_data[4],
                'inspection2_name': rnc_data[5],
                'inspection_date': rnc_data[6],
                'engineering_date': rnc_data[7],
                'inspection2_date': rnc_data[8]
            },
            'debug_info': {
                'inspection_empty': not rnc_data[3] or rnc_data[3] == 'NOME',
                'engineering_empty': not rnc_data[4] or rnc_data[4] == 'NOME',
                'inspection2_empty': not rnc_data[5] or rnc_data[5] == 'NOME'
            }
        })
        
    except Exception as e:
        logger.error(f"Erro no debug de assinaturas: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

## Rota movida: /api/rnc/<id>/delete agora está em routes/rnc.py
    """API para deletar RNC definitivamente (lixeira removida)"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se RNC existe
        cursor.execute('SELECT * FROM rncs WHERE id = ?', (rnc_id,))
        rnc = cursor.fetchone()
        
        if not rnc:
            conn.close()
            return jsonify({'success': False, 'message': 'RNC não encontrado'}), 404
        
        # Exclusão definitiva
        cursor.execute('DELETE FROM rncs WHERE id = ?', (rnc_id,))
        
        # Também excluir dados relacionados
        cursor.execute('DELETE FROM rnc_shares WHERE rnc_id = ?', (rnc_id,))
        cursor.execute('DELETE FROM chat_messages WHERE rnc_id = ?', (rnc_id,))
        
        conn.commit()
        conn.close()
        
        # LIMPAR CACHE IMEDIATAMENTE e forçar reload
        with cache_lock:
            # Limpar TODOS os caches relacionados a RNCs
            keys_to_remove = [
                key for key in query_cache.keys()
                if 'rncs_list_' in key or 'rnc_' in key or 'charts_' in key
            ]
            for key in keys_to_remove:
                del query_cache[key]
        
        logger.info(f"RNC {rnc_id} excluído definitivamente por usuário {session['user_id']}")
        
        return jsonify({
            'success': True,
            'message': 'RNC excluído definitivamente.',
            'cache_cleared': True  # Indicar que cache foi limpo
        })
        
    except Exception as e:
        logger.error(f"Erro ao deletar RNC: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

## Rota movida: /api/rnc/<id>/share agora está em routes/rnc.py
    """Compartilhar RNC com usuários"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        shared_with_user_ids = data.get('shared_with_user_ids', [])
        permission_level = data.get('permission_level', 'view')
        
        # Verificar se usuário pode compartilhar esta RNC
        rnc_data, error_message = get_rnc_data_safe(rnc_id)
        if rnc_data is None:
            return jsonify({'success': False, 'message': error_message}), 404
        
        # Verificar se é o criador ou tem permissão de admin
        user_id_index = 8  # user_id
        if len(rnc_data) <= user_id_index:
            return jsonify({'success': False, 'message': 'Dados do RNC incompletos'}), 400
        
        is_creator = (rnc_data[user_id_index] == session['user_id'])
        has_admin_permission = has_permission(session['user_id'], 'view_all_rncs')
        
        if not is_creator and not has_admin_permission:
            return jsonify({'success': False, 'message': 'Sem permissão para compartilhar esta RNC'}), 403
        
        # Compartilhar com cada usuário
        success_count = 0
        for user_id in shared_with_user_ids:
            if share_rnc_with_user(rnc_id, session['user_id'], user_id, permission_level):
                success_count += 1
        
        if success_count > 0:
            return jsonify({
                'success': True,
                'message': f'RNC compartilhada com {success_count} usuário(s) com sucesso!'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Erro ao compartilhar RNC'
            }), 500
        
    except Exception as e:
        logger.error(f"Erro ao compartilhar RNC {rnc_id}: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

## Rota movida: /api/rnc/<id>/shared-users agora está em routes/rnc.py
    """Obter usuários com quem a RNC foi compartilhada"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        # Verificar se usuário pode ver esta RNC
        if not can_user_access_rnc(session['user_id'], rnc_id):
            return jsonify({'success': False, 'message': 'Sem permissão para acessar esta RNC'}), 403
        
        shared_users = get_rnc_shared_users(rnc_id)
        
        shared_users_list = []
        for user_data in shared_users:
            shared_users_list.append({
                'user_id': user_data[0],
                'permission_level': user_data[1],
                'name': user_data[2],
                'email': user_data[3]
            })
        
        return jsonify({
            'success': True,
            'shared_users': shared_users_list
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar usuários compartilhados da RNC {rnc_id}: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/grant-permission', methods=['POST'])
def grant_permission():
    """API para admins concederem permissões a usuários"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    # Verificar se usuário é admin
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão para gerenciar usuários'}), 403
    
    try:
        data = request.get_json()
        target_user_id = data.get('user_id')
        permission = data.get('permission')
        
        if not target_user_id or not permission:
            return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Buscar permissões atuais do usuário
        cursor.execute('SELECT permissions FROM users WHERE id = ?', (target_user_id,))
        user_data = cursor.fetchone()
        
        if not user_data:
            conn.close()
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
        
        # Processar permissões atuais
        current_permissions_str = user_data[0] or '[]'
        try:
            # Tentar parsear como JSON
            current_permissions = json.loads(current_permissions_str)
        except json.JSONDecodeError:
            # Se falhar, assumir que é string separada por vírgulas
            if current_permissions_str.strip():
                current_permissions = [p.strip() for p in current_permissions_str.split(',')]
            else:
                current_permissions = []
        
        # Adicionar nova permissão se não existir
        if permission not in current_permissions:
            current_permissions.append(permission)
            permissions_json = json.dumps(current_permissions)
            
            cursor.execute('UPDATE users SET permissions = ? WHERE id = ?', (permissions_json, target_user_id))
            conn.commit()
            conn.close()
            
            logger.info(f"Permissão '{permission}' concedida para usuário {target_user_id}")
            
            return jsonify({
                'success': True,
                'message': f'Permissão "{permission}" concedida com sucesso!'
            })
        else:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Usuário já possui a permissão "{permission}"'
            }), 400
        
    except Exception as e:
        logger.error(f"Erro ao conceder permissão: {e}")
        return jsonify({
            'success': False,
            'message': f'Erro interno do sistema: {str(e)}'
        }), 500

@app.route('/api/admin/revoke-permission', methods=['POST'])
def revoke_permission():
    """API para admins removerem permissões de usuários"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    # Verificar se usuário é admin
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão para gerenciar usuários'}), 403
    
    try:
        data = request.get_json()
        target_user_id = data.get('user_id')
        permission = data.get('permission')
        
        if not target_user_id or not permission:
            return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Buscar permissões atuais do usuário
        cursor.execute('SELECT permissions FROM users WHERE id = ?', (target_user_id,))
        user_data = cursor.fetchone()
        
        if not user_data:
            conn.close()
            return jsonify({'success': False, 'message': 'Usuário não encontrado'}), 404
        
        # Processar permissões atuais
        current_permissions_str = user_data[0] or '[]'
        try:
            # Tentar parsear como JSON
            current_permissions = json.loads(current_permissions_str)
        except json.JSONDecodeError:
            # Se falhar, assumir que é string separada por vírgulas
            if current_permissions_str.strip():
                current_permissions = [p.strip() for p in current_permissions_str.split(',')]
            else:
                current_permissions = []
        
        # Remover permissão se existir
        if permission in current_permissions:
            current_permissions.remove(permission)
            permissions_json = json.dumps(current_permissions)
            
            cursor.execute('UPDATE users SET permissions = ? WHERE id = ?', (permissions_json, target_user_id))
            conn.commit()
            conn.close()
            
            logger.info(f"Permissão '{permission}' removida do usuário {target_user_id}")
            
            return jsonify({
                'success': True,
                'message': f'Permissão "{permission}" removida com sucesso!'
            })
        else:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Usuário não possui a permissão "{permission}"'
            }), 400
        
    except Exception as e:
        logger.error(f"Erro ao remover permissão: {e}")
        return jsonify({
            'success': False,
            'message': f'Erro interno do sistema: {str(e)}'
        }), 500

# Rotas de gerenciamento de grupos (apenas para admins)
@app.route('/api/admin/groups', methods=['GET'])
@app.route('/api/groups', methods=['GET'])
def api_get_groups():
    """API para obter todos os grupos.
    Disponível para qualquer usuário autenticado - permite visualizar grupos
    para atribuição/compartilhamento de RNCs independente do departamento."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        groups = get_all_groups()
        groups_list = []
        for group in groups:
            groups_list.append({
                'id': group[0],
                'name': group[1],
                'description': group[2],
                'user_count': group[3]
            })
        
        return jsonify({'success': True, 'groups': groups_list})
    except Exception as e:
        logger.error(f"Erro ao buscar grupos: {e}")
        return jsonify({'success': False, 'message': 'Erro interno do sistema'}), 500

@app.route('/api/admin/groups', methods=['POST'])
def api_create_group():
    """API para criar novo grupo"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão para gerenciar grupos'}), 403
    
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description', '')
        
        if not name:
            return jsonify({'success': False, 'message': 'Nome do grupo é obrigatório'}), 400
        
        group_id = create_group(name, description)
        if group_id:
            return jsonify({
                'success': True,
                'message': f'Grupo "{name}" criado com sucesso!',
                'group_id': group_id
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Erro ao criar grupo'
            }), 500
    except Exception as e:
        logger.error(f"Erro ao criar grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/groups/<int:group_id>', methods=['PUT'])
def api_update_group(group_id):
    """API para atualizar grupo"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão para gerenciar grupos'}), 403
    
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description', '')
        
        if not name:
            return jsonify({'success': False, 'message': 'Nome do grupo é obrigatório'}), 400
        
        if update_group(group_id, name, description):
            return jsonify({
                'success': True,
                'message': f'Grupo atualizado com sucesso!'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Erro ao atualizar grupo'
            }), 500
    except Exception as e:
        logger.error(f"Erro ao atualizar grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/groups/<int:group_id>', methods=['DELETE'])
def api_delete_group(group_id):
    """API para deletar grupo"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão para gerenciar grupos'}), 403
    
    try:
        if delete_group(group_id):
            return jsonify({
                'success': True,
                'message': 'Grupo deletado com sucesso!'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Erro ao deletar grupo'
            }), 500
    except Exception as e:
        logger.error(f"Erro ao deletar grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/groups/<int:group_id>/users', methods=['GET'])
def api_get_group_users(group_id):
    """API para obter usuários de um grupo"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        users = get_users_by_group(group_id)
        users_list = []
        for user in users:
            users_list.append({
                'id': user[0],
                'name': user[1],
                'email': user[2],
                'department': user[3],
                'role': user[4]
            })
        
        return jsonify({
            'success': True,
            'users': users_list
        })
    except Exception as e:
        logger.error(f"Erro ao buscar usuários do grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/groups/<int:group_id>/manager', methods=['GET'])
def api_get_group_manager(group_id):
    """API para obter o gerente de um grupo específico
    
    Mapeamento de gerentes por grupo:
    - Comercial: Sandro
    - Compras: Marcelo
    - Engenharia: Guilherme
    - PCP: Fernando
    - Qualidade: Alan
    - Usinagem Plana: Ronaldo
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Mapeamento de gerentes por nome do grupo
        GROUP_MANAGERS = {
            'Comercial': 'Sandro',
            'Compras': 'Marcelo',
            'Engenharia': 'Guilherme',
            'PCP': 'Fernando',
            'Qualidade': 'Alan',
            'Usinagem Plana': 'Ronaldo'
        }
        
        # Buscar nome do grupo
        cursor.execute("SELECT name FROM groups WHERE id = ?", (group_id,))
        group_row = cursor.fetchone()
        
        if not group_row:
            conn.close()
            return jsonify({'success': False, 'message': 'Grupo não encontrado'}), 404
        
        group_name = group_row[0]
        manager_name = GROUP_MANAGERS.get(group_name)
        
        if not manager_name:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Nenhum gerente definido para o grupo {group_name}'
            }), 404
        
        # Buscar dados do gerente
        cursor.execute("""
            SELECT id, name, email, department
            FROM users
            WHERE name = ? AND group_id = ?
        """, (manager_name, group_id))
        
        manager_row = cursor.fetchone()
        conn.close()
        
        if not manager_row:
            return jsonify({
                'success': False,
                'message': f'Gerente {manager_name} não encontrado no grupo {group_name}'
            }), 404
        
        return jsonify({
            'success': True,
            'manager': {
                'id': manager_row[0],
                'name': manager_row[1],
                'email': manager_row[2],
                'department': manager_row[3]
            }
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar gerente do grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/groups/suggestions', methods=['GET'])
def api_get_group_suggestions():
    """API para obter sugestões de grupos existentes"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT name FROM groups ORDER BY name')
        groups = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({
            'success': True,
            'groups': groups
        })
    except Exception as e:
        logger.error(f"Erro ao buscar sugestões de grupos: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/users/<int:user_id>/groups', methods=['GET'])
def api_get_user_groups(user_id):
    """API para obter grupos de um usuário específico"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Obter grupo principal do usuário
        cursor.execute('''
            SELECT g.id, g.name, g.description 
            FROM groups g 
            JOIN users u ON g.id = u.group_id 
            WHERE u.id = ?
        ''', (user_id,))
        main_group = cursor.fetchone()
        
        # Obter todos os grupos disponíveis
        cursor.execute('SELECT id, name, description FROM groups ORDER BY name')
        all_groups = cursor.fetchall()
        
        conn.close()
        
        groups_list = []
        for group in all_groups:
            groups_list.append({
                'id': group[0],
                'name': group[1],
                'description': group[2],
                'is_main': main_group and main_group[0] == group[0]
            })
        
        return jsonify({
            'success': True,
            'groups': groups_list,
            'main_group': main_group[1] if main_group else None
        })
    except Exception as e:
        logger.error(f"Erro ao buscar grupos do usuário: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/users/<int:user_id>/add-to-group', methods=['POST'])
def api_add_user_to_group(user_id):
    """API para adicionar usuário a um grupo"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão para gerenciar usuários'}), 403
    
    try:
        data = request.get_json()
        group_id = data.get('group_id')
        
        if not group_id:
            return jsonify({'success': False, 'message': 'ID do grupo é obrigatório'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se o grupo existe
        cursor.execute('SELECT name FROM groups WHERE id = ?', (group_id,))
        group = cursor.fetchone()
        
        if not group:
            conn.close()
            return jsonify({'success': False, 'message': 'Grupo não encontrado'}), 404
        
        # Atualizar o grupo do usuário
        cursor.execute('UPDATE users SET group_id = ? WHERE id = ?', (group_id, user_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Usuário adicionado ao grupo "{group[0]}" com sucesso!'
        })
    except Exception as e:
        logger.error(f"Erro ao adicionar usuário ao grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

# ==================== APIs DE PERMISSÕES ====================

@app.route('/api/admin/groups/<int:group_id>/permissions', methods=['GET'])
def api_get_group_permissions(group_id):
    """API para obter permissões de um grupo"""
    if 'user_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    # Permitir acesso para admin_access OU manage_users
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        permissions = get_group_permissions(group_id)  # lista de tuplas (name, value)
        # Filtrar apenas permissões concedidas (valor truthy)
        granted_permissions = [name for name, value in permissions if value]
        return jsonify({
            'success': True,
            'group_id': group_id,
            'permissions': granted_permissions
        })
    except Exception as e:
        logger.error(f"Erro ao buscar permissões do grupo: {e}")
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500

@app.route('/api/admin/groups/<int:group_id>/permissions', methods=['PUT'])
def api_update_group_permissions(group_id):
    """API para atualizar permissões de um grupo"""
    if 'user_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    # Permitir acesso para admin_access OU manage_users
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        data = request.get_json() or {}
        raw_permissions = data.get('permissions', [])

        # Aceitar tanto lista (['perm_a', 'perm_b']) quanto dict ({'perm_a': True})
        if isinstance(raw_permissions, list):
            # Validar nomes contra lista global de permissões
            valid_names = {p['name'] for p in get_all_permissions()}
            invalid = [p for p in raw_permissions if p not in valid_names]
            # Ignorar inválidas ao invés de falhar completamente
            if invalid:
                logger.warning(f"Ignorando permissões desconhecidas para grupo {group_id}: {invalid}")
            # Converter somente válidas para dict com valor True
            permissions_dict = {p: True for p in raw_permissions if p in valid_names}
        elif isinstance(raw_permissions, dict):
            # Filtrar por nomes válidos e normalizar valores para boolean
            valid_names = {p['name'] for p in get_all_permissions()}
            permissions_dict = {name: bool(val) for name, val in raw_permissions.items() if name in valid_names}
        else:
            return jsonify({'success': False, 'error': 'Formato de permissões não suportado'}), 400

        if update_group_permissions(group_id, permissions_dict):
            logger.info(f"Permissões atualizadas para grupo {group_id}: {sorted(list(permissions_dict.keys()))}")
            return jsonify({'success': True, 'message': 'Permissões atualizadas com sucesso', 'applied': list(permissions_dict.keys())})
        else:
            return jsonify({'success': False, 'error': 'Erro ao atualizar permissões'}), 500
    except Exception as e:
        logger.error(f"Erro ao atualizar permissões do grupo: {e}")
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500

@app.route('/api/admin/permissions/list', methods=['GET'])
def api_get_all_permissions():
    """API para obter lista de todas as permissões disponíveis"""
    if 'user_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    # Permitir acesso para admin_access OU manage_users
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Limpar qualquer cache de permissões para garantir dados atualizados
        clear_permissions_cache()
        permissions = get_all_permissions()
        logger.info(f"API: Retornando {len(permissions)} permissões atualizadas")
        return jsonify({'success': True, 'permissions': permissions})
    except Exception as e:
        logger.error(f"Erro ao buscar permissões: {e}")
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500

@app.route('/api/user/permissions', methods=['GET'])
def api_get_user_permissions():
    """API para obter permissões do usuário logado"""
    if 'user_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    try:
        permissions = get_user_group_permissions(session['user_id'])
        return jsonify({'permissions': permissions})
    except Exception as e:
        logger.error(f"Erro ao buscar permissões do usuário: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# Rotas de gerenciamento de usuários (apenas para admins)
@app.route('/admin/users')
def admin_users():
    """Página de gerenciamento de usuários"""
    if 'user_id' not in session:
        return redirect('/')
    
    if not has_permission(session['user_id'], 'manage_users'):
        return redirect('/dashboard')
    
    return render_template('admin_users.html')

@app.route('/admin/groups')
def admin_groups():
    """Página de gerenciamento de grupos"""
    if 'user_id' not in session:
        return redirect('/')
    
    if not has_permission(session['user_id'], 'manage_users'):
        return redirect('/dashboard')
    
    return render_template('admin_groups.html')

@app.route('/admin/permissions')
def admin_permissions():
    """Página de gerenciamento de permissões dos grupos"""
    if 'user_id' not in session:
        return redirect('/')
    
    # Permitir acesso para admin_access OU manage_users
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return redirect('/dashboard')
    
    return render_template('admin_permissions.html')

@app.route('/admin/managers')
def admin_managers():
    """Página de gerenciamento de gerentes por grupo"""
    if 'user_id' not in session:
        return redirect('/')
    
    # Apenas admins podem acessar
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return redirect('/dashboard')
    
    return render_template('admin_managers.html')

@app.route('/api/admin/groups/with-managers', methods=['GET'])
def api_get_groups_with_managers():
    """API para obter todos os grupos com informações de gerentes"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                g.id,
                g.name,
                g.description,
                g.manager_user_id,
                g.sub_manager_user_id,
                m.name as manager_name,
                m.email as manager_email,
                sm.name as sub_manager_name,
                sm.email as sub_manager_email,
                (SELECT COUNT(*) FROM users WHERE group_id = g.id AND is_active = 1) as user_count
            FROM groups g
            LEFT JOIN users m ON g.manager_user_id = m.id
            LEFT JOIN users sm ON g.sub_manager_user_id = sm.id
            ORDER BY g.name
        ''')
        
        groups = []
        for row in cursor.fetchall():
            groups.append({
                'id': row[0],
                'name': row[1],
                'description': row[2] or '',
                'manager_user_id': row[3],
                'sub_manager_user_id': row[4],
                'manager_name': row[5],
                'manager_email': row[6],
                'sub_manager_name': row[7],
                'sub_manager_email': row[8],
                'user_count': row[9]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'groups': groups
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar grupos com gerentes: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/api/admin/groups/<int:group_id>/managers', methods=['PUT'])
def api_update_group_managers(group_id):
    """API para atualizar gerente e sub-gerente de um grupo"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_users')):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403
    
    try:
        data = request.get_json()
        manager_user_id = data.get('manager_user_id')
        sub_manager_user_id = data.get('sub_manager_user_id')
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se o grupo existe
        cursor.execute('SELECT name FROM groups WHERE id = ?', (group_id,))
        group = cursor.fetchone()
        
        if not group:
            conn.close()
            return jsonify({'success': False, 'message': 'Grupo não encontrado'}), 404
        
        # Atualizar gerentes
        cursor.execute('''
            UPDATE groups 
            SET manager_user_id = ?, sub_manager_user_id = ?
            WHERE id = ?
        ''', (manager_user_id, sub_manager_user_id, group_id))
        
        conn.commit()
        conn.close()
        
        logger.info(f"Gerentes do grupo {group[0]} (ID: {group_id}) atualizados por usuário {session['user_id']}")
        
        return jsonify({
            'success': True,
            'message': 'Gerentes atualizados com sucesso'
        })
        
    except Exception as e:
        logger.error(f"Erro ao atualizar gerentes do grupo: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno do sistema'
        }), 500

@app.route('/admin/clients')
def admin_clients():
    """Cadastro de Clientes (somente admins)."""
    if 'user_id' not in session:
        return redirect('/')
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_clients_admin')):
        return redirect('/dashboard')
    return render_template('admin_client.html')

@app.route('/admin/areas')
def admin_areas():
    """Cadastro de Áreas (UI)."""
    if 'user_id' not in session:
        return redirect('/')
    # Permissão: admin ou manage_areas ou view_areas_admin
    if not (has_permission(session['user_id'], 'admin_access') or
            has_permission(session['user_id'], 'manage_areas') or
            has_permission(session['user_id'], 'view_areas_admin')):
        return redirect('/dashboard')
    # Garantir tabela e dados padrão
    ensure_areas_table()
    return render_template('admin_areas.html')

# ---------------------- SECTORS (SETOR) ----------------------
def ensure_sectors_table():
    """Garante a tabela de setores (semelhante a áreas)."""
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS sectors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            description TEXT DEFAULT '',
            color TEXT DEFAULT '#5b21b6',
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        # Inserir setores padrão (idempotente)
        default_sectors = [
            ('Engenharia',                'Setor de engenharia',                 '#1f77b4'),
            ('Cliente',                   'Cliente/Comercial',                   '#2ca02c'),
            ('Montagem',                  'Montagem mecânica',                   '#6f42c1'),
            ('Corte',                     'Corte de chapas/tubos',               '#e83e8c'),
            ('Conformação',               'Dobra/conformação',                   '#20c997'),
            ('Caldeiraria de Carbono',   'Solda e caldeiraria - carbono',       '#fd7e14'),
            ('Caldeiraria de Inox',      'Solda e caldeiraria - inox',          '#17a2b8'),
            ('Jato de Granalha',         'Preparação de superfície',            '#7952b3'),
            ('Pintura',                   'Pintura industrial',                  '#dc3545'),
            ('Usin. Cilíndrica Convencional', 'Usinagem cilíndrica convencional', '#6610f2'),
            ('Usin. Cilíndrica CNC',     'Usinagem cilíndrica CNC',             '#28a745'),
            ('Usinagem Plana',           'Retífica/usinagem plana',             '#0d6efd'),
            ('Furação',                   'Furação e mandrilhamento',            '#198754'),
            ('Célula de Secadores',      'Célula de montagem de secadores',     '#0dcaf0'),
            ('Balanceamento',            'Balanceamento de rotores',            '#ffc107'),
            ('Embalagem',                'Embalagem/expedição',                 '#343a40'),
        ]
        try:
            cur.executemany(
                'INSERT OR IGNORE INTO sectors (name, description, color) VALUES (?, ?, ?)',
                default_sectors
            )
        except Exception:
            pass
        conn.commit(); conn.close()
    except Exception as e:
        logger.error(f"Erro ao garantir tabela sectors: {e}")

@app.route('/admin/sectors')
def admin_sectors():
    """Cadastro de Setores (UI)."""
    if 'user_id' not in session:
        return redirect('/')
    if not (has_permission(session['user_id'], 'admin_access') or
            has_permission(session['user_id'], 'manage_sectors') or
            has_permission(session['user_id'], 'view_sectors_admin')):
        return redirect('/dashboard')
    ensure_sectors_table()
    return render_template('admin_sectors.html')

def ensure_clients_table():
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.commit(); conn.close()
    except Exception as e:
        logger.error(f"Erro ao garantir tabela clients: {e}")

@app.route('/api/clients/import', methods=['POST'])
def api_import_clients():
    """Importa clientes a partir da pasta 'DADOS PUXAR RNC'.

    Procura, nesta ordem:
      - 'clientes.xlsx' (aba 'clientes')
      - 'Cliente.xlsx' (aba 'clientes')
      - 'DADOS PUXAR RNC.xlsx' (aba 'clientes')
      - 'DADOS PUXAR RNC.txt' (coluna CLIENTE)
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_clients_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403

    ensure_clients_table()
    base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'DADOS PUXAR RNC')
    candidates = [
        ('clientes.xlsx', 'clientes'),
        ('Cliente.xlsx', 'clientes'),
        ('DADOS PUXAR RNC.xlsx', 'clientes'),
    ]

    clients = set()
    used_source = None

    # Tentar XLSX usando openpyxl se disponível
    try:
        import openpyxl  # type: ignore
        for fname, sheet in candidates:
            fpath = os.path.join(base_dir, fname)
            if not os.path.exists(fpath):
                continue
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = None
                for name in wb.sheetnames:
                    if name.strip().lower() == sheet:
                        ws = wb[name]
                        break
                if ws is None:
                    ws = wb.active
                # Detectar coluna com 'cliente'
                header = None
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    header = [ (str(c).strip().lower() if c is not None else '') for c in row ]
                    break
                col_idx = 0
                if header:
                    for i, h in enumerate(header):
                        if 'cliente' in h:
                            col_idx = i
                            break
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is None:
                        continue
                    name = str(val).strip()
                    if not name or name.lower() in {'cliente', 'clientes'}:
                        continue
                    name = ' '.join(name.split())
                    if len(name) > 1:
                        clients.add(name)
                if clients:
                    used_source = fpath
                    break
            except Exception:
                continue
    except Exception:
        # openpyxl indisponível; seguirá para TXT
        pass

    # Fallback: TXT tabulado
    if not clients:
        txt_path = os.path.join(base_dir, 'DADOS PUXAR RNC.txt')
        if os.path.exists(txt_path):
            try:
                with open(txt_path, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()
                for line in lines[1:]:
                    if not line.strip():
                        continue
                    parts = line.split('\t')
                    if len(parts) >= 12:
                        raw = parts[11].strip()
                        if raw and raw.lower() != 'cliente':
                            name = ' '.join(raw.split())
                            if len(name) > 1:
                                clients.add(name)
                if clients:
                    used_source = txt_path
            except Exception as e:
                logger.error(f"Erro lendo TXT de clientes: {e}")

    if not clients:
        return jsonify({'success': False, 'message': 'Nenhum cliente encontrado nos arquivos em DADOS PUXAR RNC'}), 404

    # Inserir/atualizar no banco
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        inserted = 0; updated = 0
        for name in sorted(clients):
            try:
                cur.execute('SELECT id FROM clients WHERE name = ?', (name,))
                row = cur.fetchone()
                if row:
                    cur.execute('UPDATE clients SET updated_at = CURRENT_TIMESTAMP WHERE id = ?', (row[0],))
                    updated += 1
                else:
                    cur.execute('INSERT INTO clients (name) VALUES (?)', (name,))
                    inserted += 1
            except sqlite3.IntegrityError:
                updated += 1
        conn.commit(); conn.close()
    except Exception as e:
        logger.error(f"Erro ao salvar clientes: {e}")
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Erro ao salvar clientes no banco'}), 500

    return jsonify({'success': True, 'message': f'Importação concluída ({inserted} novos, {updated} atualizados).', 'inserted': inserted, 'updated': updated, 'source': used_source})

def ensure_operators_table():
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS operators (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            number TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        # Garantir coluna number caso a tabela já exista
        try:
            cur.execute('ALTER TABLE operators ADD COLUMN number TEXT')
        except sqlite3.OperationalError:
            pass
        # Índice único para número (apenas quando não nulo)
        try:
            cur.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_operators_number ON operators(number) WHERE number IS NOT NULL')
        except sqlite3.OperationalError:
            pass
        conn.commit(); conn.close()
    except Exception as e:
        logger.error(f"Erro ao garantir tabela operators: {e}")

@app.route('/admin/operators')
def admin_operators():
    if 'user_id' not in session:
        return redirect('/')
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_operators_admin')):
        return redirect('/dashboard')
    return render_template('admin_operator.html')

@app.route('/api/admin/operators', methods=['GET'])
def api_list_operators():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_operators_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_operators_table()
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('SELECT id, name, number FROM operators ORDER BY name')
    rows = [{'id': r[0], 'name': r[1], 'number': r[2]} for r in cur.fetchall()]
    conn.close()
    return jsonify({'success': True, 'operators': rows})

@app.route('/api/admin/operators', methods=['POST'])
def api_create_operator():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_operators_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_operators_table()
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    number = (str(data.get('number')).strip() if data.get('number') is not None else None)
    if not name:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    try:
        cur.execute('INSERT INTO operators (name, number) VALUES (?, ?)', (name, number))
        conn.commit()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Operador já existe'}), 400
    finally:
        conn.close()

@app.route('/api/admin/operators/<int:operator_id>', methods=['PUT'])
def api_update_operator(operator_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_operators_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_operators_table()
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    number = (str(data.get('number')).strip() if data.get('number') is not None else None)
    if not name:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('UPDATE operators SET name = ?, number = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (name, number, operator_id))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/operators/<int:operator_id>', methods=['DELETE'])
def api_delete_operator(operator_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_operators_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_operators_table()
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('DELETE FROM operators WHERE id = ?', (operator_id,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

def ensure_areas_table():
    """Garante a tabela de áreas para cadastro administrativo."""
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS areas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            description TEXT DEFAULT '',
            color TEXT DEFAULT '#6f42c1',
            is_active BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        # Inserir áreas padrão (idempotente)
        default_areas = [
            ('Fundibem',            '', '#6f42c1'),
            ('Engenharia',          '', '#007bff'),
            ('Terceiros',           '', '#17a2b8'),
            ('Transporte',          '', '#fd7e14'),
            ('Não Definidos',       '', '#6c757d'),
            ('Qualidade',           '', '#dc3545'),
            ('Compras',             '', '#20c997'),
            ('Comercial',           '', '#28a745'),
            ('PCP',                 '', '#6610f2'),
            ('Almoxarifado',        '', '#ffc107'),
            ('Produção',            '', '#343a40'),
            ('Filial',              '', '#4b2ca1'),
        ]
        cur.executemany('INSERT OR IGNORE INTO areas (name, description, color) VALUES (?, ?, ?)', default_areas)
        conn.commit(); conn.close()
    except Exception as e:
        logger.error(f"Erro ao garantir tabela areas: {e}")

@app.route('/api/admin/clients', methods=['GET'])
def api_list_clients():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_clients_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_clients_table()
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('SELECT id, name FROM clients ORDER BY name')
    rows = [{'id': r[0], 'name': r[1]} for r in cur.fetchall()]
    conn.close()
    return jsonify({'success': True, 'clients': rows})

@app.route('/api/admin/areas', methods=['GET'])
def api_list_areas():
    """Lista áreas cadastradas (somente admins/áreas)."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or
            has_permission(session['user_id'], 'manage_areas') or
            has_permission(session['user_id'], 'view_areas_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_areas_table()
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('SELECT id, name, description, color, is_active, created_at FROM areas ORDER BY name')
    rows = [
        {
            'id': r[0], 'name': r[1], 'description': r[2] or '', 'color': r[3] or '#6f42c1',
            'is_active': bool(r[4]), 'created_at': r[5]
        } for r in cur.fetchall()
    ]
    conn.close()
    return jsonify({'success': True, 'areas': rows})

@app.route('/api/admin/sectors', methods=['GET'])
def api_list_sectors():
    """Lista setores cadastrados (somente admins/setores)."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or
            has_permission(session['user_id'], 'manage_sectors') or
            has_permission(session['user_id'], 'view_sectors_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_sectors_table()
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('SELECT id, name, description, color, is_active, created_at FROM sectors ORDER BY name')
    rows = [
        {
            'id': r[0], 'name': r[1], 'description': r[2] or '', 'color': r[3] or '#5b21b6',
            'is_active': bool(r[4]), 'created_at': r[5]
        } for r in cur.fetchall()
    ]
    conn.close()
    return jsonify({'success': True, 'sectors': rows})
@app.route('/api/admin/areas', methods=['POST'])
def api_create_area():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_areas')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_areas_table()
    data = request.get_json(force=True) if request.is_json else request.form
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    color = (data.get('color') or '#6f42c1').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        cur.execute('INSERT INTO areas (name, description, color) VALUES (?, ?, ?)', (name, description, color))
        conn.commit(); area_id = cur.lastrowid; conn.close()
        return jsonify({'success': True, 'area': {'id': area_id, 'name': name, 'description': description, 'color': color}})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Já existe uma área com este nome'}), 409
    except Exception as e:
        logger.error(f"Erro ao criar área: {e}")
        return jsonify({'success': False, 'message': 'Erro interno'}), 500

@app.route('/api/admin/sectors', methods=['POST'])
def api_create_sector():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_sectors')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_sectors_table()
    data = request.get_json(force=True) if request.is_json else request.form
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    color = (data.get('color') or '#5b21b6').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        cur.execute('INSERT INTO sectors (name, description, color) VALUES (?, ?, ?)', (name, description, color))
        conn.commit(); sector_id = cur.lastrowid; conn.close()
        return jsonify({'success': True, 'sector': {'id': sector_id, 'name': name, 'description': description, 'color': color}})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Já existe um setor com este nome'}), 409
    except Exception as e:
        logger.error(f"Erro ao criar setor: {e}")
        return jsonify({'success': False, 'message': 'Erro interno'}), 500

@app.route('/api/admin/areas/<int:area_id>', methods=['PUT'])
def api_update_area(area_id: int):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_areas')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_areas_table()
    data = request.get_json(force=True) if request.is_json else request.form
    name = data.get('name')
    description = data.get('description')
    color = data.get('color')
    is_active = data.get('is_active')
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        fields = []; params = []
        if name is not None:
            fields.append('name = ?'); params.append(name)
        if description is not None:
            fields.append('description = ?'); params.append(description)
        if color is not None:
            fields.append('color = ?'); params.append(color)
        if is_active is not None:
            fields.append('is_active = ?'); params.append(1 if str(is_active).lower() in ('1','true','yes') else 0)
        if not fields:
            return jsonify({'success': False, 'message': 'Nada para atualizar'}), 400
        params.append(area_id)
        cur.execute(f"UPDATE areas SET {', '.join(fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", params)
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Nome já em uso'}), 409
    except Exception as e:
        logger.error(f"Erro ao atualizar área {area_id}: {e}")
        return jsonify({'success': False, 'message': 'Erro interno'}), 500

@app.route('/api/admin/sectors/<int:sector_id>', methods=['PUT'])
def api_update_sector(sector_id: int):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_sectors')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_sectors_table()
    data = request.get_json(force=True) if request.is_json else request.form
    name = data.get('name')
    description = data.get('description')
    color = data.get('color')
    is_active = data.get('is_active')
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        fields = []; params = []
        if name is not None:
            fields.append('name = ?'); params.append(name)
        if description is not None:
            fields.append('description = ?'); params.append(description)
        if color is not None:
            fields.append('color = ?'); params.append(color)
        if is_active is not None:
            fields.append('is_active = ?'); params.append(1 if str(is_active).lower() in ('1','true','yes') else 0)
        if not fields:
            return jsonify({'success': False, 'message': 'Nada para atualizar'}), 400
        params.append(sector_id)
        cur.execute(f"UPDATE sectors SET {', '.join(fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", params)
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Nome já em uso'}), 409
    except Exception as e:
        logger.error(f"Erro ao atualizar setor {sector_id}: {e}")
        return jsonify({'success': False, 'message': 'Erro interno'}), 500

@app.route('/api/admin/areas/<int:area_id>', methods=['DELETE'])
def api_delete_area(area_id: int):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_areas')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_areas_table()
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        cur.execute('UPDATE areas SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (area_id,))
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Erro ao desativar área {area_id}: {e}")
        return jsonify({'success': False, 'message': 'Erro interno'}), 500

@app.route('/api/admin/sectors/<int:sector_id>', methods=['DELETE'])
def api_delete_sector(sector_id: int):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'manage_sectors')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_sectors_table()
    try:
        conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
        cur.execute('UPDATE sectors SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (sector_id,))
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Erro ao desativar setor {sector_id}: {e}")
        return jsonify({'success': False, 'message': 'Erro interno'}), 500

# Lista de clientes para uso geral (formulários), sem exigir permissão de admin
@app.route('/api/clients', methods=['GET'])
def api_list_clients_public():
    """Retorna somente os nomes dos clientes para autocomplete/seleção no formulário de RNC.
    Requer usuário autenticado, mas não exige permissão administrativa.
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    try:
        ensure_clients_table()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('SELECT name FROM clients ORDER BY name')
        names = [r[0] for r in cur.fetchall()]
        conn.close()
        return jsonify({'success': True, 'clients': names})
    except Exception as e:
        logger.error(f"Erro ao listar clientes públicos: {e}")
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@app.route('/api/admin/clients', methods=['POST'])
def api_create_client():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_clients_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_clients_table()
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    try:
        cur.execute('INSERT INTO clients (name) VALUES (?)', (name,))
        conn.commit()
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Cliente já existe'}), 400
    finally:
        conn.close()

@app.route('/api/admin/clients/<int:client_id>', methods=['PUT'])
def api_update_client(client_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not (has_permission(session['user_id'], 'admin_access') or has_permission(session['user_id'], 'view_clients_admin')):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_clients_table()
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Nome é obrigatório'}), 400
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('UPDATE clients SET name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (name, client_id))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/clients/<int:client_id>', methods=['DELETE'])
def api_delete_client(client_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    if not has_permission(session['user_id'], 'admin_access') and not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    ensure_clients_table()
    conn = sqlite3.connect(DB_PATH); cur = conn.cursor()
    cur.execute('DELETE FROM clients WHERE id = ?', (client_id,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/users/list')
def api_list_users():
    """API para listar usuários ativos - disponível para todos os usuários
    autenticados para permitir atribuição/compartilhamento de RNCs
    Parâmetros opcionais:
    - group_id: filtra usuários por grupo específico
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se foi passado filtro de grupo
        group_id = request.args.get('group_id')
        
        if group_id:
            # Filtrar apenas usuários do grupo especificado
            cursor.execute('''
                SELECT id, name, email, department, group_id 
                FROM users 
                WHERE is_active = 1 AND group_id = ? 
                ORDER BY name
            ''', (group_id,))
        else:
            # Retornar todos os usuários ativos
            cursor.execute('SELECT id, name, email, department, group_id FROM users WHERE is_active = 1 ORDER BY name')
        
        users = cursor.fetchall()
        conn.close()
        
        formatted_users = []
        for user in users:
            formatted_users.append({
                'id': user[0],
                'name': user[1],
                'email': user[2],
                'department': user[3],
                'group_id': user[4]
            })
        
        return jsonify({'success': True, 'users': formatted_users})
        
    except Exception as e:
        logger.error(f"Erro ao listar usuários: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/admin/users', methods=['GET'])
def api_get_users():
    """Listar usuários com filtros (q, role, department, include_inactive)."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403
    
    # Coletar filtros
    q = request.args.get('q', '').strip()
    role = request.args.get('role')
    department = request.args.get('department')
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'

    # Montar query dinâmica
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    query = '''
        SELECT u.id, u.name, u.email, u.department, u.role, u.permissions,
               u.created_at, u.is_active, g.name as group_name
        FROM users u
        LEFT JOIN groups g ON u.group_id = g.id
        WHERE 1=1
    '''
    params = []
    if not include_inactive:
        query += ' AND u.is_active = 1'
    if q:
        like = f"%{q}%"
        query += ' AND (u.name LIKE ? OR u.email LIKE ? OR u.department LIKE ?)'
        params.extend([like, like, like])
    if role:
        query += ' AND u.role = ?'
        params.append(role)
    if department:
        query += ' AND u.department = ?'
        params.append(department)
    query += ' ORDER BY u.name'

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    formatted_users = []
    import json
    for user in rows:
        permissions = []
        try:
            permissions = json.loads(user[5] or '[]')
        except Exception:
            permissions = []
        formatted_users.append({
            'id': user[0],
            'name': user[1],
            'email': user[2],
            'department': user[3],
            'role': user[4],
            'permissions': permissions,
            'created_at': user[6],
            'is_active': bool(user[7]),
            'group_name': user[8]
        })
    
    return jsonify({'success': True, 'users': formatted_users})

@app.route('/api/admin/users', methods=['POST'])
def api_create_user():
    """API para criar usuário (apenas admin)"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403
    
    try:
        data = request.get_json()
        
        success = create_user(
            name=data.get('name'),
            email=data.get('email'),
            password=data.get('password'),
            department=data.get('department'),
            role=data.get('role', 'user'),
            permissions=data.get('permissions', [])
        )
        
        if success:
            return jsonify({'success': True, 'message': 'Usuário criado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Email já existe!'}), 400
            
    except Exception as e:
        logger.error(f"Erro ao criar usuário: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
def api_update_user(user_id):
    """API para atualizar usuário (apenas admin)"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403
    
    try:
        data = request.get_json()
        
        success = update_user(
            user_id=user_id,
            name=data.get('name'),
            email=data.get('email'),
            department=data.get('department'),
            role=data.get('role'),
            permissions=data.get('permissions', []),
            is_active=data.get('is_active', True)
        )
        
        if success:
            return jsonify({'success': True, 'message': 'Usuário atualizado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao atualizar usuário!'}), 400
            
    except Exception as e:
        logger.error(f"Erro ao atualizar usuário: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
def api_delete_user(user_id):
    """API para deletar usuário (apenas admin)"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403
    
    try:
        success = delete_user(user_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Usuário desativado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao deletar usuário!'}), 400
            
    except Exception as e:
        logger.error(f"Erro ao deletar usuário: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>/restore', methods=['POST'])
def api_restore_user(user_id):
    """Reativar usuário desativado."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401

    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403

    try:
        if restore_user(user_id):
            return jsonify({'success': True, 'message': 'Usuário restaurado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao restaurar usuário!'}), 400
    except Exception as e:
        logger.error(f"Erro ao restaurar usuário: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
def api_reset_user_password(user_id):
    """Definir uma nova senha para o usuário."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401

    if not has_permission(session['user_id'], 'manage_users'):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403

    data = request.get_json() or {}
    new_password = data.get('new_password')
    if not new_password or len(new_password) < 6:
        return jsonify({'success': False, 'message': 'Senha inválida (mínimo 6 caracteres)'}), 400

    try:
        if update_user_password(user_id, new_password):
            return jsonify({'success': True, 'message': 'Senha redefinida com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao redefinir senha!'}), 400
    except Exception as e:
        logger.error(f"Erro ao redefinir senha: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/admin/users/import-from-rncs', methods=['POST'])
def api_import_users_from_rncs():
    """Cria usuários automaticamente a partir das RNCs existentes.

    Regras:
    - Nome base: rnc.responsavel, ou alguma assinatura (inspection/engineering/inspection2)
    - Departamento: area_responsavel, senão setor, senão department, senão 'Geral'
    - E-mail sintético único: <slug>@ippel.local
    - Grupo é criado automaticamente a partir do departamento (reuso de create_user)
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    if not (has_permission(session['user_id'], 'manage_users') or has_permission(session['user_id'], 'admin_access')):
        return jsonify({'success': False, 'message': 'Sem permissão'}), 403

    def _slug(s: str) -> str:
        try:
            import re, unicodedata
            s = unicodedata.normalize('NFKD', s)
            s = ''.join(ch for ch in s if not unicodedata.combining(ch))
            s = re.sub(r'[^a-zA-Z0-9]+', '-', s).strip('-').lower()
            return s or 'user'
        except Exception:
            return 'user'

    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        # Usuários já existentes
        cur.execute('SELECT name, email FROM users')
        existing = cur.fetchall()
        existing_names = { (n or '').strip().lower() for n, _ in existing }
        existing_emails = { (e or '').strip().lower() for _, e in existing }

        # Coletar candidatos nas RNCs
        cur.execute('''
            SELECT 
                COALESCE(NULLIF(TRIM(responsavel), ''),
                         NULLIF(TRIM(signature_inspection_name), ''),
                         NULLIF(TRIM(signature_engineering_name), ''),
                         NULLIF(TRIM(signature_inspection2_name), '')) AS nome,
                COALESCE(NULLIF(TRIM(area_responsavel), ''),
                         NULLIF(TRIM(setor), ''),
                         NULLIF(TRIM(department), ''), 'Geral') AS dept
              FROM rncs
        ''')
        rows = cur.fetchall()
        conn.close()

        created = 0
        skipped = 0
        for nome, dept in rows:
            name = (nome or '').strip()
            dept = (dept or 'Geral').strip()
            if not name or name.lower() in ('n/a', 'nome', '-', 'sem nome'):
                skipped += 1
                continue
            low = name.lower()
            if low in existing_names:
                skipped += 1
                continue
            base = _slug(name)
            email = f"{base}@ippel.local"
            i = 1
            while email.lower() in existing_emails:
                i += 1
                email = f"{base}{i}@ippel.local"
            ok = create_user(
                name=name,
                email=email,
                password='ippel123',
                department=dept or 'Geral',
                role='user',
                permissions=['create_rnc','view_own_rnc']
            )
            if ok:
                created += 1
                existing_names.add(low)
                existing_emails.add(email.lower())
            else:
                skipped += 1

        return jsonify({'success': True, 'created': created, 'skipped': skipped})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/chat')
def general_chat():
    """Página do chat geral"""
    if 'user_id' not in session:
        return redirect('/')
    
    return render_template('general_chat.html')

@app.route('/api/chat/<int:rnc_id>/messages')
def get_chat_messages(rnc_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se a coluna file_path existe
        cursor.execute("PRAGMA table_info(chat_messages)")
        columns = [col[1] for col in cursor.fetchall()]
        has_file_path = 'file_path' in columns
        
        if has_file_path:
            cursor.execute('''
                SELECT cm.id, cm.rnc_id, cm.user_id, cm.message, cm.message_type, 
                       cm.created_at, cm.file_path, u.name as user_name, u.department
                FROM chat_messages cm
                LEFT JOIN users u ON cm.user_id = u.id
                WHERE cm.rnc_id = ?
                ORDER BY cm.created_at ASC
            ''', (rnc_id,))
        else:
            cursor.execute('''
                SELECT cm.id, cm.rnc_id, cm.user_id, cm.message, cm.message_type, 
                       cm.created_at, u.name as user_name, u.department
                FROM chat_messages cm
                LEFT JOIN users u ON cm.user_id = u.id
                WHERE cm.rnc_id = ?
                ORDER BY cm.created_at ASC
            ''', (rnc_id,))
        
        messages = cursor.fetchall()
        conn.close()
        
        formatted_messages = []
        for msg in messages:
            if has_file_path:
                formatted_messages.append({
                    'id': msg[0],
                    'rnc_id': msg[1],
                    'user_id': msg[2],
                    'message': msg[3],
                    'message_type': msg[4],
                    'created_at': msg[5],
                    'file_path': msg[6],
                    'user_name': msg[7],
                    'department': msg[8]
                })
            else:
                formatted_messages.append({
                    'id': msg[0],
                    'rnc_id': msg[1],
                    'user_id': msg[2],
                    'message': msg[3],
                    'message_type': msg[4],
                    'created_at': msg[5],
                    'file_path': None,
                    'user_name': msg[6],
                    'department': msg[7]
                })
        
        logger.info(f" API retornando {len(formatted_messages)} mensagens para RNC {rnc_id}")
        return jsonify({'success': True, 'messages': formatted_messages})
        
    except Exception as e:
        logger.error(f"Erro ao buscar mensagens: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500


@app.route('/api/chat/<int:rnc_id>/messages', methods=['POST'])
def post_chat_message(rnc_id):
    """Criar nova mensagem de chat via HTTP.
    Fallback para quando o Socket.IO não estiver disponível.
    Suporta texto, imagens e arquivos.
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    try:
        # VERIFICAR SE A RNC ESTÁ FINALIZADA
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT status FROM rncs WHERE id = ?', (rnc_id,))
        rnc = cursor.fetchone()
        conn.close()
        
        if rnc and rnc[0] == 'Finalizado':
            return jsonify({
                'success': False, 
                'message': 'Esta RNC foi finalizada. Não é possível enviar mensagens.'
            }), 403
        # Verificar se é FormData (com arquivo) ou JSON
        is_multipart = request.content_type and 'multipart/form-data' in request.content_type
        
        if is_multipart:
            # Upload de imagem/arquivo
            message = (request.form.get('message') or '').strip()
            message_type = request.form.get('message_type', 'text')
            message_id_from_client = request.form.get('message_id')
            
            # Verificar se há arquivo
            file_data = None
            file_path = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename:
                    # Criar pasta de uploads se não existir
                    import os
                    from werkzeug.utils import secure_filename
                    upload_folder = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'chat_images')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    # Salvar arquivo com nome único
                    file_ext = os.path.splitext(file.filename)[1]
                    unique_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{session['user_id']}{file_ext}"
                    file_path = os.path.join(upload_folder, unique_filename)
                    file.save(file_path)
                    
                    # Caminho relativo para o banco de dados
                    file_data = f"/static/uploads/chat_images/{unique_filename}"
                    message_type = 'image'
            elif 'file' in request.files:
                file = request.files['file']
                if file and file.filename:
                    import os
                    from werkzeug.utils import secure_filename
                    upload_folder = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'chat_files')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    file_ext = os.path.splitext(file.filename)[1]
                    unique_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{session['user_id']}_{secure_filename(file.filename)}"
                    file_path = os.path.join(upload_folder, unique_filename)
                    file.save(file_path)
                    
                    file_data = f"/static/uploads/chat_files/{unique_filename}"
                    message_type = 'file'
            
            if not message and not file_data:
                return jsonify({'success': False, 'message': 'Mensagem ou arquivo não pode estar vazio'}), 400
                
        else:
            # Mensagem de texto normal (JSON)
            payload = request.get_json(silent=True) or {}
            message = (payload.get('message') or '').strip()
            message_type = 'text'
            message_id_from_client = payload.get('message_id')
            file_data = None
            
            if not message:
                return jsonify({'success': False, 'message': 'Mensagem não pode estar vazia'}), 400

        user_id = session['user_id']
        
        # Verificar cache para evitar duplicação
        if message_id_from_client:
            cache_key = f"unique_{message_id_from_client}"
            if cache_key in message_cache:
                logger.info(f" Mensagem HTTP já processada recentemente, retornando sucesso: {cache_key}")
                # Retornar sucesso com dados da mensagem (para o cliente adicionar à interface)
                return jsonify({
                    'success': True, 
                    'message': 'Mensagem já processada', 
                    'already_processed': True,
                    'message_data': {
                        'id': message_id_from_client,
                        'rnc_id': rnc_id,
                        'user_id': user_id,
                        'message': message or '',
                        'message_type': message_type,
                        'file_path': file_data,
                        'created_at': datetime.now().isoformat()
                    }
                })
            message_cache[cache_key] = True
            import threading
            threading.Timer(10.0, lambda: message_cache.pop(cache_key, None)).start()
            logger.info(f" Processando nova mensagem HTTP: {cache_key}")

        # Salvar no banco
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se a coluna file_path existe na tabela
        cursor.execute("PRAGMA table_info(chat_messages)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'file_path' in columns:
            cursor.execute(
                'INSERT INTO chat_messages (rnc_id, user_id, message, message_type, file_path) VALUES (?, ?, ?, ?, ?)',
                (rnc_id, user_id, message or '', message_type, file_data)
            )
        else:
            # Fallback se a coluna não existir
            cursor.execute(
                'INSERT INTO chat_messages (rnc_id, user_id, message, message_type) VALUES (?, ?, ?, ?)',
                (rnc_id, user_id, message or '', message_type)
            )
        
        message_id = cursor.lastrowid

        # Buscar infos do usuário
        cursor.execute('SELECT name, department FROM users WHERE id = ?', (user_id,))
        user_info = cursor.fetchone()

        conn.commit()
        conn.close()

        message_data = {
            'id': message_id,
            'rnc_id': rnc_id,
            'user_id': user_id,
            'message': message or '',
            'message_type': message_type,
            'file_path': file_data,
            'created_at': datetime.now().isoformat(),
            'user_name': user_info[0] if user_info else 'Usuário',
            'department': user_info[1] if user_info else ''
        }

        # Notificar em tempo real para os OUTROS usuários
        # O remetente já adiciona a mensagem via resposta HTTP
        try:
            if 'socketio' in globals() and socketio is not None:
                # Emitir para todos na sala
                # O remetente vai ignorar a duplicação via verificação de message-id no frontend
                socketio.emit('new_message', message_data, room=f'rnc_{rnc_id}')
        except Exception:
            pass

        # Criar notificações para outros participantes do RNC
        try:
            from services.enhanced_notifications import enhanced_notification_service, NotificationType, NotificationChannel
            
            conn_notif = sqlite3.connect(DB_PATH)
            cursor_notif = conn_notif.cursor()
            
            # Buscar informações do RNC
            cursor_notif.execute('SELECT rnc_number, title, user_id, assigned_user_id FROM rncs WHERE id = ?', (rnc_id,))
            rnc_info = cursor_notif.fetchone()
            
            if rnc_info:
                rnc_number, rnc_title, creator_id, assigned_user_id = rnc_info
                
                # Buscar todos os usuários que participaram do chat
                cursor_notif.execute('''
                    SELECT DISTINCT user_id 
                    FROM chat_messages 
                    WHERE rnc_id = ? AND user_id != ?
                ''', (rnc_id, user_id))
                
                participants = [row[0] for row in cursor_notif.fetchall()]
                
                # Adicionar criador e responsável aos participantes
                if creator_id and creator_id != user_id and creator_id not in participants:
                    participants.append(creator_id)
                if assigned_user_id and assigned_user_id != user_id and assigned_user_id not in participants:
                    participants.append(assigned_user_id)
                
                # Criar notificação para cada participante
                for participant_id in participants:
                    try:
                        enhanced_notification_service.create_notification(
                            notification_type=NotificationType.RNC_COMMENTED,
                            to_user_id=participant_id,
                            data={
                                'rnc_number': rnc_number,
                                'commenter_name': message_data['user_name'],
                                'action_url': f'/rnc/{rnc_id}/chat',
                                'message_preview': message[:50] + ('...' if len(message or '') > 50 else '')
                            },
                            from_user_id=user_id,
                            rnc_id=rnc_id,
                            channels=[NotificationChannel.IN_APP, NotificationChannel.BROWSER]
                        )
                        logger.info(f" Notificação criada para usuário {participant_id}")
                    except Exception as e:
                        logger.error(f" Erro ao criar notificação para usuário {participant_id}: {e}")
                
                logger.info(f" Notificações criadas para {len(participants)} participantes")
            
            conn_notif.close()
        except Exception as e:
            logger.error(f" Erro ao criar notificações: {e}")
            if 'conn_notif' in locals():
                conn_notif.close()

        # CORREÇÃO: Retornar dados da mensagem com chave correta
        return jsonify({
            'success': True, 
            'message': message_data,  # Dados completos da mensagem
            'text': 'Mensagem enviada com sucesso'  # Texto de confirmação
        })

    except Exception as e:
        try:
            if 'conn' in locals():
                conn.close()
        except Exception:
            pass
        logger.error(f'Erro ao salvar mensagem via API: {e}')
        return jsonify({'success': False, 'message': 'Erro interno ao salvar mensagem'}), 500

@app.route('/api/chat/<int:rnc_id>/messages/<int:message_id>', methods=['DELETE'])
def delete_chat_message(rnc_id, message_id):
    """Excluir mensagem do chat de um RNC.
    Regras:
    - Usuário deve estar autenticado
    - Pode excluir se for autor da mensagem ou se for admin
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401

    user_id = session['user_id']
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Buscar mensagem
        cursor.execute('SELECT id, rnc_id, user_id FROM chat_messages WHERE id = ? AND rnc_id = ?', (message_id, rnc_id))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'message': 'Mensagem não encontrada'}), 404

        _, msg_rnc_id, author_id = row

        # Buscar role do usuário atual
        cursor.execute('SELECT role FROM users WHERE id = ?', (user_id,))
        user_row = cursor.fetchone()
        current_role = user_row[0] if user_row else 'user'

        # Verificar permissão (autor ou admin)
        if user_id != author_id and current_role != 'admin':
            conn.close()
            return jsonify({'success': False, 'message': 'Sem permissão para excluir esta mensagem'}), 403

        # Excluir mensagem
        cursor.execute('DELETE FROM chat_messages WHERE id = ?', (message_id,))
        conn.commit()
        conn.close()

        # Notificar clientes na sala para remover a mensagem
        try:
            socketio.emit('message_deleted', {'id': message_id, 'rnc_id': msg_rnc_id}, room=f'rnc_{msg_rnc_id}')
        except Exception as e:
            logger.error(f'Erro ao emitir evento de exclusão de mensagem: {e}')

        return jsonify({'success': True, 'message': 'Mensagem excluída com sucesso'})

    except Exception as e:
        logger.error(f"Erro ao excluir mensagem: {e}")
        if 'conn' in locals():
            conn.close()
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/notifications/unread')
def get_unread_notifications():
    """API para buscar notificações não lidas - SEM RATE LIMIT para polling"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Buscar notificações não lidas do usuário
        cursor.execute('''
            SELECT n.*, r.rnc_number, r.title as rnc_title
            FROM notifications n
            LEFT JOIN rncs r ON n.rnc_id = r.id
            WHERE n.user_id = ? AND n.is_read = 0
            ORDER BY n.created_at DESC
            LIMIT 10
        ''', (session['user_id'],))
        
        notifications = cursor.fetchall()
        conn.close()
        
        formatted_notifications = []
        for notif in notifications:
            formatted_notifications.append({
                'id': notif[0],
                'user_id': notif[1],
                'rnc_id': notif[2],
                'type': notif[3],
                'title': notif[4],
                'message': notif[5],
                'is_read': notif[6],
                'created_at': notif[7],
                'rnc_number': notif[8],
                'rnc_title': notif[9]
            })
        
        return jsonify({
            'success': True,
            'notifications': formatted_notifications,
            'count': len(formatted_notifications)
        })
        
    except Exception as e:
        logger.error(f"Erro ao buscar notificações: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/notifications/mark-read', methods=['POST'])
def mark_notifications_read():
    """API para marcar notificações como lidas"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        data = request.get_json()
        notification_ids = data.get('notification_ids', [])
        
        if not notification_ids:
            return jsonify({'success': False, 'message': 'IDs de notificação não fornecidos'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Marcar notificações como lidas
        placeholders = ','.join(['?' for _ in notification_ids])
        cursor.execute(f'''
            UPDATE notifications 
            SET is_read = 1 
            WHERE id IN ({placeholders}) AND user_id = ?
        ''', notification_ids + [session['user_id']])
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{cursor.rowcount} notificação(ões) marcada(s) como lida(s)'
        })
        
    except Exception as e:
        logger.error(f"Erro ao marcar notificações como lidas: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/api/messages/mark_viewed', methods=['POST'])
def mark_messages_viewed():
    """API para marcar mensagens de um RNC como visualizadas (estilo WhatsApp)"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    user_id = session['user_id']
    data = request.get_json()
    rnc_id = data.get('rnc_id')
    
    if not rnc_id:
        return jsonify({'success': False, 'message': 'RNC ID não fornecido'}), 400
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Verificar se as tabelas e colunas existem
        cursor.execute("PRAGMA table_info(message_views)")
        message_views_cols = [row[1] for row in cursor.fetchall()]
        
        cursor.execute("PRAGMA table_info(chat_messages)")
        chat_messages_cols = [row[1] for row in cursor.fetchall()]
        
        # Marcar mensagens como visualizadas na tabela message_views
        if 'user_id' in message_views_cols and 'rnc_id' in message_views_cols:
            cursor.execute('''
                INSERT OR REPLACE INTO message_views (user_id, rnc_id, last_viewed_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', (user_id, rnc_id))
        
        marked_count = 0
        
        # Marcar mensagens individuais como visualizadas (estilo WhatsApp)
        # Apenas se a coluna viewed_at existir
        if 'viewed_at' in chat_messages_cols and 'user_id' in chat_messages_cols:
            cursor.execute('''
                UPDATE chat_messages 
                SET viewed_at = CURRENT_TIMESTAMP 
                WHERE rnc_id = ? 
                AND user_id != ? 
                AND viewed_at IS NULL
            ''', (rnc_id, user_id))
            marked_count = cursor.rowcount
        else:
            logger.warning(f" Coluna viewed_at não existe em chat_messages. Colunas disponíveis: {chat_messages_cols}")
            # Garantir que a coluna seja adicionada usando a conexão existente
            ensure_chat_viewed_at_column(conn)
            marked_count = 0
            
            # Tentar novamente após adicionar a coluna
            try:
                cursor.execute('''
                    UPDATE chat_messages 
                    SET viewed_at = CURRENT_TIMESTAMP 
                    WHERE rnc_id = ? 
                    AND user_id != ? 
                    AND viewed_at IS NULL
                ''', (rnc_id, user_id))
                marked_count = cursor.rowcount
                logger.info(" Conseguiu marcar mensagens após adicionar coluna viewed_at")
            except Exception as retry_error:
                logger.error(f" Ainda não conseguiu marcar mensagens após adicionar coluna: {retry_error}")
                marked_count = 0
        
        # Marcar notificações relacionadas como lidas
        cursor.execute('''
            UPDATE notifications 
            SET is_read = 1 
            WHERE user_id = ? AND rnc_id = ? AND type = 'new_message'
        ''', (user_id, rnc_id))
        
        conn.commit()
        conn.close()
        
        logger.info(f" {marked_count} mensagens do RNC {rnc_id} marcadas como visualizadas (viewed_at) pelo usuário {user_id}")
        
        return jsonify({
            'success': True, 
            'message': 'Mensagens marcadas como visualizadas',
            'marked_count': marked_count
        })
        
    except Exception as e:
        logger.error(f"Erro ao marcar mensagens como visualizadas: {e}")
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

# Eventos WebSocket
@socketio.on('connect')
def handle_connect():
    try:
        if 'user_id' not in session:
            # Tentar reidratar sessão a partir do cookie auxiliar
            uid = request.cookies.get('IPPEL_UID')
            if uid and uid.isdigit():
                try:
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute('SELECT id, name, email, department, role FROM users WHERE id = ?', (int(uid),))
                    row = cursor.fetchone()
                    conn.close()
                    if row:
                        session['user_id'] = row[0]
                        session['user_name'] = row[1]
                        session['user_email'] = row[2]
                        session['user_department'] = row[3]
                        session['user_role'] = row[4]
                        session.permanent = True
                        logger.info(f"Sessão reidratada via cookie para user {row[0]}")
                except Exception as e:
                    logger.error(f"Falha ao reidratar sessão socket: {e}")
        if 'user_id' in session:
            user_id = session['user_id']
            online_users[request.sid] = user_id
            logger.info(f"Usuário {user_id} conectado (socket)")
        else:
            logger.warning("Conexão socket sem sessão válida")
    except Exception as e:
        logger.error(f"Erro em handle_connect: {e}")

@socketio.on('disconnect')
def handle_disconnect():
    if request.sid in online_users:
        user_id = online_users.pop(request.sid)
        logger.info(f"Usuário {user_id} desconectado")

@socketio.on('join_room')
def handle_join_room(data):
    room = data.get('room')
    user_id = session.get('user_id', 'unknown')
    user_name = session.get('user_name', 'unknown')
    
    if not room:
        logger.error(" Nome da sala não fornecido")
        emit('error', {'message': 'Nome da sala não fornecido'})
        return {'success': False, 'error': 'Nome da sala não fornecido'}
    
    join_room(room)
    logger.info(f" Usuário {user_name} (ID: {user_id}) entrou na sala: {room}")
    logger.info(f" Socket ID: {request.sid}")
    
    # Verificar quantos usuários estão na sala após entrar
    from flask_socketio import rooms
    room_users = rooms(room)
    logger.info(f" Usuários na sala {room} após entrada: {room_users}")
    
    # Log adicional: notificar outros na sala
    try:
        # Enviar notificação para outros na sala que alguém entrou
        emit('user_joined_room', {
            'user_id': user_id,
            'user_name': user_name,
            'room': room
        }, room=room, include_self=False)
        logger.info(f" Notificado outros usuários na sala {room}")
    except Exception as e:
        logger.error(f" Erro ao notificar entrada na sala: {e}")
    
    # Enviar confirmação de volta ao cliente
    emit('room_joined', {'room': room, 'success': True, 'user_id': user_id, 'user_name': user_name})
    
    # Log adicional: verificar quantos usuários estão na room após entrada
    try:
        from flask_socketio import rooms
        room_users = rooms(room)
        logger.info(f" Usuários na room {room} após entrada: {room_users}")
    except Exception as e:
        logger.error(f" Erro ao verificar usuários na room: {e}")
    
    return {'success': True, 'room': room, 'user_id': user_id}

@socketio.on('leave_room')
def handle_leave_room(data):
    room = data['room']
    leave_room(room)
    logger.info(f"Usuário saiu da sala: {room}")

@socketio.on('send_message')
def handle_send_message(data):
    # DEBUGGING: Log da sessão
    logger.info(f" SESSION DEBUG: {dict(session)}")
    logger.info(f" USER_ID in session: {'user_id' in session}")
    if 'user_id' in session:
        logger.info(f" USER_ID value: {session['user_id']}")
    
    # Verificar se o usuário está logado
    if 'user_id' not in session:
        logger.error(" Usuário não autenticado ao tentar enviar mensagem")
        logger.error(f" Sessão atual: {dict(session)}")
        emit('error', {'message': 'Usuário não autenticado'})
        return
    
    logger.info(f" Tentativa de envio de mensagem do usuário {session['user_id']}")
    logger.info(f" Dados recebidos: {data}")
    
    try:
        # Verificar tipo de chat
        chat_type = data.get('chat_type', 'rnc')  # Default para 'rnc' se não especificado
        logger.info(f" Tipo de chat identificado: {chat_type}")
        
        if chat_type == 'private':
            logger.info(" Processando mensagem privada")
            handle_private_message(data)
        elif chat_type == 'general':
            logger.info(" Processando mensagem de chat geral")
            handle_general_chat_message(data)
        elif chat_type == 'rnc':
            logger.info(" Processando mensagem de chat RNC")
            handle_rnc_chat_message(data)
        else:
            logger.warning(f" Tipo de chat desconhecido: {chat_type}, usando RNC como padrão")
            handle_rnc_chat_message(data)
            
        logger.info(" Mensagem processada com sucesso")
        # Retornar sucesso
        return {'success': True}
    except Exception as e:
        logger.error(f" Erro ao enviar mensagem: {e}")
        import traceback
        logger.error(traceback.format_exc())
        # Não emitir erro - deixar o cliente usar fallback HTTP
        logger.info(" Socket.IO falhou, cliente usará fallback HTTP")
        return {'success': False, 'error': str(e)}
def handle_private_message(data):
    """Processar mensagem privada"""
    try:
        message = data['message']
        recipient_id = data['recipient_id']
        sender_id = session['user_id']
        
        # Verificar cache para evitar duplicação
        message_key = f"private_{sender_id}_{recipient_id}_{hash(message)}"
        if message_key in message_cache:
            logger.info(f" Mensagem privada já processada recentemente, ignorando")
            return
        message_cache[message_key] = True
        import threading
        threading.Timer(10.0, lambda: message_cache.pop(message_key, None)).start()
        
        # Salvar mensagem no banco
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO private_messages (sender_id, recipient_id, message, message_type)
            VALUES (?, ?, ?, ?)
        ''', (sender_id, recipient_id, message, 'text'))
        
        message_id = cursor.lastrowid
        
        # Buscar informações do usuário
        cursor.execute('SELECT name, department FROM users WHERE id = ?', (sender_id,))
        user_info = cursor.fetchone()
        
        conn.commit()
        conn.close()
        
        # Preparar dados da mensagem
        message_data = {
            'id': message_id,
            'sender_id': sender_id,
            'recipient_id': recipient_id,
            'message': message,
            'message_type': 'text',
            'is_read': False,
            'created_at': datetime.now().isoformat(),
            'user_name': user_info[0] if user_info else 'Usuário',
            'department': user_info[1] if user_info else ''
        }
        
        # Criar sala única para a conversa
        room = f"private_{min(sender_id, recipient_id)}_{max(sender_id, recipient_id)}"
        
        # Enviar mensagem para a sala privada
        emit('private_message', message_data, room=room)
        
        # Criar notificação para o destinatário
        notification_data = {
            'type': 'new_private_message',
            'title': f' Nova mensagem de {user_info[0] if user_info else "Usuário"}',
            'message': f'{message[:50]}{"..." if len(message) > 50 else ""}',
            'sender_id': sender_id,
            'sender_name': user_info[0] if user_info else 'Usuário',
            'timestamp': datetime.now().isoformat()
        }
        
        # Salvar notificação no banco
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO notifications (user_id, type, title, message, is_read)
                VALUES (?, ?, ?, ?, ?)
            ''', (recipient_id, 'new_private_message', notification_data['title'], notification_data['message'], 0))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Erro ao salvar notificação: {e}")
        
        # Enviar notificação em tempo real
        emit('notification', notification_data, room=f'user_{recipient_id}')
        
        logger.info(f"Nova mensagem privada de {sender_id} para {recipient_id}: {message}")
        
    except Exception as e:
        logger.error(f"Erro ao processar mensagem privada: {e}")
        # Não emitir erro - deixar o cliente usar fallback HTTP

def handle_general_chat_message(data):
    """Processar mensagem do chat geral (mantido para compatibilidade)"""
    try:
        message = data['message']
        user_id = session['user_id']
        
        # Verificar cache para evitar duplicação
        message_key = f"general_{user_id}_{hash(message)}"
        if message_key in message_cache:
            logger.info(f" Mensagem geral já processada recentemente, ignorando")
            return
        message_cache[message_key] = True
        import threading
        threading.Timer(10.0, lambda: message_cache.pop(message_key, None)).start()
        
        # Buscar informações do usuário
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT name, department FROM users WHERE id = ?', (user_id,))
        user_info = cursor.fetchone()
        conn.close()
        
        # Preparar dados da mensagem
        message_data = {
            'id': 0,
            'user_id': user_id,
            'message': message,
            'message_type': 'text',
            'created_at': datetime.now().isoformat(),
            'user_name': user_info[0] if user_info else 'Usuário',
            'department': user_info[1] if user_info else ''
        }
        
        # Enviar mensagem para todos os usuários online
        emit('general_chat_message', message_data, broadcast=True)
        
        logger.info(f"Nova mensagem no chat geral: {message}")
        
    except Exception as e:
        logger.error(f"Erro ao processar mensagem do chat geral: {e}")
        # Não emitir erro - deixar o cliente usar fallback HTTP

# Cache para evitar processamento duplicado de mensagens
message_cache = {}

def handle_rnc_chat_message(data):
    """Processar mensagem do chat de RNC"""
    logger.info(f" ===== INICIANDO handle_rnc_chat_message =====")
    logger.info(f" Dados recebidos: {data}")
    logger.info(f" Sessão atual: {dict(session)}")
    
    try:
        rnc_id = data.get('rnc_id')
        message = data.get('message', '').strip()
        user_id = session.get('user_id')
        
        # Usar ID único da mensagem se fornecido, senão criar chave baseada no conteúdo
        message_id = data.get('message_id')
        if message_id:
            message_key = f"unique_{message_id}"
        else:
            message_key = f"{user_id}_{rnc_id}_{hash(message)}"
        
        # Verificar se já processamos esta mensagem recentemente
        if message_key in message_cache:
            logger.info(f" Mensagem já processada recentemente, ignorando: {message_key}")
            logger.info(f" Cache atual tem {len(message_cache)} entradas")
            # Emitir erro para que o cliente use HTTP
            emit('error', {'message': 'Mensagem já processada, use HTTP'})
            return
        
        # Adicionar ao cache (expira em 10 segundos)
        message_cache[message_key] = True
        import threading
        threading.Timer(10.0, lambda: message_cache.pop(message_key, None)).start()
        
        logger.info(f" Processando nova mensagem: {message_key}")
        logger.info(f" Cache atual tem {len(message_cache)} entradas")
        
        logger.info(f" Validando dados - rnc_id: {rnc_id}, message_length: {len(message)}, user_id: {user_id}")
        
        if not rnc_id:
            logger.error(f" RNC ID não fornecido")
            emit('error', {'message': 'ID do RNC não fornecido'})
            return
            
        if not message:
            logger.error(f" Mensagem vazia")
            emit('error', {'message': 'Mensagem não pode estar vazia'})
            return
            
        if not user_id:
            logger.error(f" Usuário não autenticado")
            emit('error', {'message': 'Usuário não autenticado'})
            return
        
        logger.info(f" Dados validados - Processando mensagem do RNC {rnc_id} do usuário {user_id}")
        
        # Salvar mensagem no banco
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO chat_messages (rnc_id, user_id, message, message_type)
            VALUES (?, ?, ?, ?)
        ''', (rnc_id, user_id, message, 'text'))
        
        message_id = cursor.lastrowid
        logger.info(f"Mensagem salva no banco com ID {message_id}")
        
        # Buscar informações do usuário
        cursor.execute('SELECT name, department FROM users WHERE id = ?', (user_id,))
        user_info = cursor.fetchone()
        
        # Buscar informações do RNC para notificação
        cursor.execute('SELECT rnc_number, title FROM rncs WHERE id = ?', (rnc_id,))
        rnc_info = cursor.fetchone()
        
        conn.commit()
        conn.close()
        
        # Preparar dados da mensagem
        message_data = {
            'id': message_id,
            'rnc_id': rnc_id,
            'user_id': user_id,
            'message': message,
            'message_type': 'text',
            'created_at': datetime.now().isoformat(),
            'user_name': user_info[0] if user_info else 'Usuário',
            'department': user_info[1] if user_info else ''
        }
        
        # Preparar dados da notificação
        notification_data = {
            'type': 'new_message',
            'title': f' Nova mensagem no RNC {rnc_info[0] if rnc_info else rnc_id}',
            'message': f'{user_info[0] if user_info else "Usuário"}: {message[:50]}{"..." if len(message) > 50 else ""}',
            'rnc_id': rnc_id,
            'rnc_number': rnc_info[0] if rnc_info else f'RNC-{rnc_id}',
            'rnc_title': rnc_info[1] if rnc_info else 'RNC',
            'user_name': user_info[0] if user_info else 'Usuário',
            'department': user_info[1] if user_info else '',
            'timestamp': datetime.now().isoformat()
        }
        
        # Enviar mensagem para todos na sala (EXCETO o remetente)
        room_name = f'rnc_{rnc_id}'
        logger.info(f" ========================================")
        logger.info(f" ENVIANDO MENSAGEM PARA SALA: {room_name}")
        logger.info(f" Dados da mensagem: {message_data}")
        logger.info(f" Remetente user_id: {user_id}, Socket ID: {request.sid}")
        logger.info(f" Emitindo evento 'new_message' com broadcast=True, include_self=False")
        
        # Verificar quantos usuários estão na sala
        from flask_socketio import rooms
        room_users = rooms(room_name)
        logger.info(f" Usuários na sala {room_name}: {room_users}")
        
        # Enviar para TODOS na sala, EXCETO o remetente (include_self=False)
        # O remetente já viu a mensagem ser adicionada localmente após confirmação
        emit('new_message', message_data, room=room_name, include_self=False)
        
        logger.info(f" ========================================")
        logger.info(f" MENSAGEM EMITIDA COM BROADCAST PARA SALA {room_name}")
        logger.info(f" Evento: 'new_message', Room: '{room_name}', Broadcast: True")
        logger.info(f" Usuários na sala: {room_users}")
        logger.info(f" ========================================")
        
        # Enviar confirmação de sucesso para o remetente
        emit('message_sent', {'success': True, 'message_id': message_id})
        
        # Criar notificações usando o sistema melhorado de notificações
        try:
            from services.enhanced_notifications import enhanced_notification_service, NotificationType, NotificationChannel
            from services.notification_socketio import send_realtime_notification
            
            conn_notif = sqlite3.connect(DB_PATH)
            cursor_notif = conn_notif.cursor()
            
            # Buscar informações do RNC
            cursor_notif.execute('SELECT rnc_number, title, user_id, assigned_user_id FROM rncs WHERE id = ?', (rnc_id,))
            rnc_info_notif = cursor_notif.fetchone()
            
            if rnc_info_notif:
                rnc_number, rnc_title, creator_id, assigned_user_id = rnc_info_notif
                
                # Buscar todos os usuários que participaram do chat
                cursor_notif.execute('''
                    SELECT DISTINCT user_id 
                    FROM chat_messages 
                    WHERE rnc_id = ? AND user_id != ?
                ''', (rnc_id, user_id))
                
                participants = [row[0] for row in cursor_notif.fetchall()]
                
                # Adicionar criador e responsável aos participantes
                if creator_id and creator_id != user_id and creator_id not in participants:
                    participants.append(creator_id)
                if assigned_user_id and assigned_user_id != user_id and assigned_user_id not in participants:
                    participants.append(assigned_user_id)
                
                # Criar notificação para cada participante
                for participant_id in participants:
                    try:
                        notification_id = enhanced_notification_service.create_notification(
                            notification_type=NotificationType.RNC_COMMENTED,
                            to_user_id=participant_id,
                            data={
                                'rnc_number': rnc_number,
                                'commenter_name': user_info[0] if user_info else 'Usuário',
                                'action_url': f'/rnc/{rnc_id}/chat',
                                'message_preview': message[:50] + ('...' if len(message) > 50 else '')
                            },
                            from_user_id=user_id,
                            rnc_id=rnc_id,
                            channels=[NotificationChannel.IN_APP, NotificationChannel.BROWSER]
                        )
                        
                        # Enviar notificação em tempo real se usuário estiver online
                        if notification_id and 'socketio' in globals() and socketio is not None:
                            notif_data = {
                                'id': notification_id,
                                'type': 'rnc_commented',
                                'title': f' Nova mensagem no {rnc_number}',
                                'message': f'{user_info[0] if user_info else "Usuário"}: {message[:50]}{"..." if len(message) > 50 else ""}',
                                'data': {
                                    'rnc_number': rnc_number,
                                    'commenter_name': user_info[0] if user_info else 'Usuário',
                                    'action_url': f'/rnc/{rnc_id}/chat'
                                }
                            }
                            try:
                                send_realtime_notification(participant_id, notif_data, socketio)
                            except Exception as e:
                                logger.error(f" Erro ao enviar notificação em tempo real: {e}")
                        
                        logger.info(f" Notificação criada para usuário {participant_id}")
                    except Exception as e:
                        logger.error(f" Erro ao criar notificação para usuário {participant_id}: {e}")
                
                logger.info(f" Notificações criadas para {len(participants)} participantes")
            
            conn_notif.close()
        except Exception as e:
            logger.error(f" Erro ao criar notificações: {e}")
            if 'conn_notif' in locals():
                conn_notif.close()
        
        logger.info(f"Nova mensagem no RNC {rnc_id}: {message}")
        
    except Exception as e:
        logger.error(f"Erro ao enviar mensagem: {e}")
        # Não emitir erro - deixar o cliente usar fallback HTTP

@socketio.on('typing')
def handle_typing(data):
    if 'user_id' not in session:
        return
    
    try:
        chat_type = data.get('chat_type', 'rnc')  # 'rnc' ou 'private'
        is_typing = data['is_typing']
        user_id = session['user_id']
        
        # Buscar informações do usuário
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT name FROM users WHERE id = ?', (user_id,))
        user_info = cursor.fetchone()
        conn.close()
        
        typing_data = {
            'user_id': user_id,
            'user_name': user_info[0] if user_info else 'Usuário',
            'is_typing': is_typing
        }
        
        if chat_type == 'rnc':
            rnc_id = data.get('rnc_id')
            if rnc_id:
                # Enviar indicador de digitação para chat de RNC
                emit('user_typing', typing_data, room=f'rnc_{rnc_id}', include_self=False)
        elif chat_type == 'private':
            contact_id = data.get('contact_id')
            if contact_id:
                # Enviar indicador de digitação para chat privado
                room_name = f'private_{min(user_id, contact_id)}_{max(user_id, contact_id)}'
                emit('user_typing', typing_data, room=room_name, include_self=False)
        
    except Exception as e:
        logger.error(f"Erro ao processar indicador de digitação: {e}")

@app.route('/<path:filename>')
def serve_static(filename):
    """Servir arquivos estáticos"""
    # Evitar path traversal
    if '..' in filename or filename.startswith(('/', '\\')):
        return abort(400)
    return send_from_directory('.', filename)

def cleanup_old_deleted_rncs():
    """(Desativado) Lixeira removida."""
    return

def schedule_cleanup():
    """(Desativado) Lixeira removida."""
    logger.info("Rotina de limpeza desativada (lixeira removida)")

@app.route('/api/charts/data')
def get_chart_data():
    """API para fornecer dados dos gráficos - CORRIGIDA"""
    if 'user_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    # Remover verificação de permissão para permitir acesso aos gráficos
    # if not has_permission(session['user_id'], 'view_charts'):
    #     return jsonify({'error': 'Acesso negado: usuário não tem permissão para visualizar gráficos'}), 403
    
    try:
        period = request.args.get('period', '30', type=int)
        cache_key = f"charts_{session['user_id']}_{period}"
        cached = get_cached_query(cache_key)
        if cached:
            return jsonify(cached)
        logger.info(f"Gerando dados de gráficos para período de {period} dias")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Data limite baseada no período
        from datetime import datetime, timedelta
        limit_date = datetime.now() - timedelta(days=period)
        # Evitar DeprecationWarning do sqlite3 (Py 3.12+) convertendo datetime para string explícita
        limit_date_str = limit_date.strftime('%Y-%m-%d %H:%M:%S')
        
        # 1. Gráfico de Status
        cursor.execute('''
            SELECT status, COUNT(*) as count
            FROM rncs 
            WHERE created_at >= ? AND is_deleted = 0
            GROUP BY status
            ORDER BY count DESC
        ''', (limit_date_str,))
        status_data = cursor.fetchall()
        
        # 2. Gráfico de Tendência Temporal (RNCs por dia)
        cursor.execute('''
            SELECT DATE(created_at) as date, COUNT(*) as count
            FROM rncs 
            WHERE created_at >= ? AND is_deleted = 0
            GROUP BY DATE(created_at)
            ORDER BY date
        ''', (limit_date_str,))
        trend_data = cursor.fetchall()
        
        # 3. Gráfico de Clientes
        cursor.execute('''
            SELECT client, COUNT(*) as count
            FROM rncs 
            WHERE created_at >= ? AND is_deleted = 0 AND client IS NOT NULL AND client != ''
            GROUP BY client
            ORDER BY count DESC
            LIMIT 10
        ''', (limit_date_str,))
        client_data = cursor.fetchall()
        
        # 4. Gráfico de Equipamentos
        cursor.execute('''
            SELECT equipment, COUNT(*) as count
            FROM rncs 
            WHERE created_at >= ? AND is_deleted = 0 AND equipment IS NOT NULL AND equipment != ''
            GROUP BY equipment
            ORDER BY count DESC
            LIMIT 10
        ''', (limit_date_str,))
        equipment_data = cursor.fetchall()
        
        # 5. Gráfico de Departamentos
        cursor.execute('''
            SELECT u.department, COUNT(*) as count
            FROM rncs r
            JOIN users u ON r.user_id = u.id
            WHERE r.created_at >= ? AND r.is_deleted = 0 AND u.department IS NOT NULL
            GROUP BY u.department
            ORDER BY count DESC 
        ''', (limit_date_str,))
        department_data = cursor.fetchall()
        
        # 6. Gráfico de Prioridades
        cursor.execute('''
            SELECT priority, COUNT(*) as count
            FROM rncs 
            WHERE created_at >= ? AND is_deleted = 0 AND priority IS NOT NULL
            GROUP BY priority
            ORDER BY count DESC
        ''', (limit_date_str,))
        priority_data = cursor.fetchall()
        
        # 7. Gráfico de Disposição
        disposition_data = []
        dispositions = [
            ('disposition_usar', 'Usar'),
            ('disposition_retrabalhar', 'Retrabalhar'),
            ('disposition_rejeitar', 'Rejeitar'),
            ('disposition_sucata', 'Sucata'),
            ('disposition_devolver_estoque', 'Devolver ao Estoque'),
            ('disposition_devolver_fornecedor', 'Devolver ao Fornecedor')
        ]
        
        for field, label in dispositions:
            cursor.execute(f'''
                SELECT COUNT(*) as count
                FROM rncs 
                WHERE created_at >= ? AND is_deleted = 0 AND {field} = 1
            ''', (limit_date_str,))
            count = cursor.fetchone()[0]
            if count > 0:
                disposition_data.append((label, count))
        
        # 8. Gráfico de Usuários
        cursor.execute('''
            SELECT u.name, COUNT(*) as count
            FROM rncs r
            JOIN users u ON r.user_id = u.id
            WHERE r.created_at >= ? AND r.is_deleted = 0
            GROUP BY r.user_id, u.name
            ORDER BY count DESC
            LIMIT 10
        ''', (limit_date_str,))
        user_data = cursor.fetchall()
        
        return_db_connection(conn)
        
        result = {
            'status': [{'label': row[0], 'count': row[1]} for row in status_data],
            'trend': [{'date': row[0], 'count': row[1]} for row in trend_data],
            'clients': [{'label': row[0], 'count': row[1]} for row in client_data],
            'equipment': [{'label': row[0], 'count': row[1]} for row in equipment_data],
            'departments': [{'label': row[0], 'count': row[1]} for row in department_data],
            'priorities': [{'label': row[0], 'count': row[1]} for row in priority_data],
            'dispositions': [{'label': row[0], 'count': row[1]} for row in disposition_data],
            'users': [{'label': row[0], 'count': row[1]} for row in user_data]
        }
        cache_query(cache_key, result, ttl=60)
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Erro ao gerar dados dos gráficos: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/charts/simple-data')
def get_simple_charts_data():
    """API simplificada para dados dos gráficos - CORRIGIDA"""
    try:
        logger.info("Iniciando API simple-data")
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Buscar dados básicos
        cur.execute('SELECT COUNT(*) FROM rncs WHERE is_deleted = 0')
        total_rncs = cur.fetchone()[0]
        
        cur.execute('SELECT status, COUNT(*) FROM rncs WHERE is_deleted = 0 GROUP BY status')
        status_data = cur.fetchall()
        
        # Buscar usuários que criaram RNCs
        cur.execute('''
            SELECT u.name, COUNT(*) as count
            FROM rncs r
            JOIN users u ON r.user_id = u.id
            WHERE r.is_deleted = 0
            GROUP BY r.user_id, u.name
            ORDER BY count DESC
            LIMIT 10
        ''')
        users_data = cur.fetchall()
        
        return_db_connection(conn)
        
        result = {
            'success': True,
            'data': {
                'total_rncs': total_rncs,
                'status': [{'label': row[0], 'count': row[1]} for row in status_data],
                'users': [{'label': row[0], 'count': row[1]} for row in users_data]
            }
        }
        
        logger.info(f"API simple-data retornando: {len(users_data)} usuários")
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Erro na API simple-data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500



# Route para importação de dados (mantido para compatibilidade)
@app.route('/api/importar-dados', methods=['GET'])
def importar_dados():
    """Route simples para compatibilidade"""
    return jsonify({'success': True, 'message': 'Dados importados com sucesso'})

@app.route('/api/test/performance')
def test_performance_api():
    """API de teste para verificar se todos os usuários estão sendo retornados"""
    if 'user_id' not in session:
        return jsonify({'error': 'Usuário não autenticado'}), 401
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 1. VERIFICAR TODOS OS USUÁRIOS NO BANCO
        cursor.execute("SELECT COUNT(*) FROM users")
        total_users_db = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM users WHERE is_active = 1 OR is_active IS NULL")
        active_users_db = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM users WHERE name IS NOT NULL AND name != ''")
        named_users_db = cursor.fetchone()[0]
        
        print(f" DIAGNÓSTICO COMPLETO:")
        print(f"   Total de usuários no banco: {total_users_db}")
        print(f"   Usuários ativos: {active_users_db}")
        print(f"   Usuários com nome: {named_users_db}")
        
        # 2. BUSCAR TODOS OS USUÁRIOS ATIVOS COM NOME
        cursor.execute("""
            SELECT id, name, department, is_active FROM users 
            WHERE name IS NOT NULL 
              AND name != '' 
              AND (is_active = 1 OR is_active IS NULL)
            ORDER BY id
        """)
        all_users = cursor.fetchall()
        
        print(f" Usuários encontrados na consulta: {len(all_users)}")
        for user_id, name, dept, active in all_users[:10]:
            print(f"   ID {user_id}: {name} - {dept or 'Sem dept'} - Ativo: {active}")
        
        # 3. BUSCAR RNCs BASEADAS NAS ASSINATURAS DE ENGENHARIA
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN r.signature_engineering_name IS NOT NULL AND r.signature_engineering_name != '' 
                    THEN r.signature_engineering_name
                    ELSE r.user_id
                END as owner_id,
                COUNT(r.id) as rnc_count
            FROM rncs r
            WHERE r.status IN ('Finalizado','finalized')
              AND r.is_deleted = 0
            GROUP BY owner_id
            ORDER BY rnc_count DESC
        """)
        rnc_rows = cursor.fetchall()
        rnc_data = {row[0]: row[1] for row in rnc_rows}
        
        print(f" RNCs encontradas por assinatura: {rnc_data}")
        
        # 4. PROCESSAR RESULTADO BASEADO NAS ASSINATURAS
        meta_mensal = 5
        result = []
        
        # Criar lista de assinaturas únicas
        unique_signatures = set()
        for owner_id, count in rnc_rows:
            if isinstance(owner_id, str) and owner_id.strip():
                unique_signatures.add(owner_id.strip())
            elif isinstance(owner_id, int):
                # Se for ID, buscar nome do usuário
                cursor.execute("SELECT name FROM users WHERE id = ?", (owner_id,))
                user_result = cursor.fetchone()
                if user_result:
                    unique_signatures.add(user_result[0])
        
        print(f" Assinaturas únicas encontradas: {unique_signatures}")
        
        # Processar cada assinatura
        for signature in sorted(unique_signatures):
            rncs = rnc_data.get(signature, 0)
            percentage = (rncs / meta_mensal * 100) if meta_mensal > 0 else 0
            status = 'Acima da Meta' if rncs >= meta_mensal else 'Abaixo da Meta'
            
            print(f" PROCESSANDO - Assinatura: {signature} - RNCs: {rncs} - Status: {status}")
            
            result.append({
                'id': signature,
                'name': signature,
                'rncs': rncs,
                'meta': meta_mensal,
                'percentage': round(percentage, 1),
                'status': status,
                'department': 'Engenharia',
                'is_active': True
            })
        
        # 5. VERIFICAR USUÁRIOS QUE NÃO APARECERAM
        users_with_rncs = set(rnc_data.keys())
        all_user_ids = {user[0] for user in all_users}
        users_without_rncs = all_user_ids - users_with_rncs
        
        print(f" Usuários sem RNCs: {users_without_rncs}")
        
        conn.close()
        
        return jsonify({
            'success': True,
            'diagnostic': {
                'total_users_db': total_users_db,
                'active_users_db': active_users_db,
                'named_users_db': named_users_db,
                'users_found_in_query': len(all_users),
                'users_with_rncs': len(users_with_rncs),
                'users_without_rncs': len(users_without_rncs)
            },
            'total_users': len(all_users),
            'total_rncs_found': len(rnc_data),
            'data': result,
            'debug': {
                'users_processed': len(result),
                'rnc_data': rnc_data,
                'users_without_rncs': list(users_without_rncs)
            }
        })
        
    except Exception as e:
        print(f" Erro no teste de performance: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/available-years')
def get_available_years():
    """API para retornar os anos disponíveis no banco de dados"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        # Gerar todos os anos de 2013 até 2025
        all_years = [str(year) for year in range(2013, 2026)]
        
        # Adicionar anos futuros até 2030
        current_year = datetime.now().year
        for year in range(2026, 2031):
            all_years.append(str(year))
        
        # Ordenar anos (mais recentes primeiro)
        all_years.sort(reverse=True)
        
        print(f" Anos gerados: {all_years}")
        
        return jsonify({
            'success': True,
            'years': all_years,
            'current_year': str(current_year),
            'total_years': len(all_years)
        })
        
    except Exception as e:
        print(f"Erro ao buscar anos disponíveis: {e}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

@app.route('/api/available-months')
def get_available_months():
    """API para retornar todos os meses disponíveis"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Usuário não autenticado'}), 401
    
    try:
        # Todos os meses do ano
        months = [
            {'value': '01', 'name': 'Janeiro'},
            {'value': '02', 'name': 'Fevereiro'},
            {'value': '03', 'name': 'Março'},
            {'value': '04', 'name': 'Abril'},
            {'value': '05', 'name': 'Maio'},
            {'value': '06', 'name': 'Junho'},
            {'value': '07', 'name': 'Julho'},
            {'value': '08', 'name': 'Agosto'},
            {'value': '09', 'name': 'Setembro'},
            {'value': '10', 'name': 'Outubro'},
            {'value': '11', 'name': 'Novembro'},
            {'value': '12', 'name': 'Dezembro'}
        ]
        
        return jsonify({
            'success': True,
            'months': months,
            'total_months': len(months)
        })
        
    except Exception as e:
        print(f"Erro ao buscar meses disponíveis: {e}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500

# ===================== ROTA PARA GERAR RELATÓRIO POR DATA =====================
@app.route('/generate_report_by_date', methods=['POST'])
def generate_report_by_date():
    """Gera relatório de RNCs por período de data especificado."""
    if 'user_id' not in session:
        return redirect('/')
    
    try:
        # Obter dados do formulário
        start_date = request.form.get('reportStartDate')
        end_date = request.form.get('reportEndDate')
        report_format = request.form.get('reportFormat', 'pdf')
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Datas inicial e final são obrigatórias'}), 400
        
        # Buscar RNCs no período especificado
        conn = sqlite3.connect('ippel_system.db')
        cursor = conn.cursor()
        
        # Query para buscar RNCs no período
        query = """
            SELECT 
                numero_rnc, titulo, cliente, equipamento, setor, responsavel, 
                data_finalizacao, prioridade, status
            FROM rncs 
            WHERE data_finalizacao BETWEEN ? AND ?
            ORDER BY data_finalizacao DESC
        """
        
        cursor.execute(query, (start_date, end_date))
        rncs = cursor.fetchall()
        conn.close()
        
        # Para agora, retornar uma mensagem simples
        # (Você pode implementar a geração do relatório aqui mais tarde)
        response_data = {
            'success': True,
            'message': f'Relatório gerado com sucesso! Período: {start_date} a {end_date}',
            'format': report_format,
            'total_rncs': len(rncs),
            'start_date': start_date,
            'end_date': end_date
        }
        
        # Por enquanto, retornar dados em JSON para teste
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Erro ao gerar relatório por data: {e}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500


# ============================================================================
# VALORES/HORA - API para gerenciar tabela de valores
# ============================================================================

@app.route('/api/valores-hora/list', methods=['GET'])
def api_valores_hora_list():
    """Lista todos os valores/hora cadastrados"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, codigo, setor, descricao, valor_hora, created_at, updated_at
            FROM valores_hora
            ORDER BY codigo
        ''')
        
        rows = cursor.fetchall()
        valores = []
        for row in rows:
            valores.append({
                'id': row[0],
                'codigo': row[1],
                'setor': row[2],
                'descricao': row[3],
                'valor_hora': row[4],
                'created_at': row[5],
                'updated_at': row[6]
            })
        
        return_db_connection(conn)
        return jsonify({'success': True, 'valores': valores})
        
    except Exception as e:
        print(f"Erro ao listar valores/hora: {e}")
        return jsonify({'success': False, 'message': 'Erro ao listar valores'}), 500


@app.route('/api/valores-hora/save', methods=['POST'])
def api_valores_hora_save():
    valor_hora = request.json.get('valor_hora', '0')
    valor_hora = valor_hora.replace(',', '.')
    valor_hora_float = float(valor_hora)
    try:
        valor_hora_float = float(valor_hora)
    except ValueError:
        return jsonify({'success': False, 'message': 'Valor inválido'}), 400
    """Salva ou atualiza um valor/hora"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    try:
        data = request.get_json()
        codigo = data.get('codigo', '').strip()
        setor = data.get('setor', '').strip()
        descricao = data.get('descricao', '').strip()
        valor_hora_str = data.get('valor_hora', '').strip()
        
        # Validações
        if not codigo or not descricao or not valor_hora_str:
            return jsonify({'success': False, 'message': 'Código, descrição e valor são obrigatórios'}), 400
        
        # Extrair apenas números e vírgula/ponto do valor
        valor_limpo = re.sub(r'[^\d,.]', '', valor_hora_str)
        valor_limpo = valor_limpo.replace(',', '.')
        
        try:
            valor_hora = float(valor_limpo)
            if valor_hora < 0:
                return jsonify({'success': False, 'message': 'Valor não pode ser negativo'}), 400
        except ValueError:
            return jsonify({'success': False, 'message': 'Valor inválido. Use apenas números'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar se já existe
        cursor.execute('SELECT id FROM valores_hora WHERE codigo = ?', (codigo,))
        existing = cursor.fetchone()
        
        if existing:
            # Atualizar
            cursor.execute('''
                UPDATE valores_hora 
                SET setor = ?, descricao = ?, valor_hora = ?, updated_at = CURRENT_TIMESTAMP
                WHERE codigo = ?
            ''', (setor, descricao, valor_hora, codigo))
        else:
            # Inserir
            cursor.execute('''
                INSERT INTO valores_hora (codigo, setor, descricao, valor_hora)
                VALUES (?, ?, ?, ?)
            ''', (codigo, setor, descricao, valor_hora))
        
        conn.commit()
        return_db_connection(conn)
        
        return jsonify({'success': True, 'message': 'Valor salvo com sucesso'})
        
    except Exception as e:
        print(f"Erro ao salvar valor/hora: {e}")
        return jsonify({'success': False, 'message': f'Erro ao salvar: {str(e)}'}), 500


@app.route('/api/valores-hora/save-bulk', methods=['POST'])
def api_valores_hora_save_bulk():
    """Salva múltiplos valores de uma vez"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    # CONTROLE DE ACESSO: Apenas Ronaldo (ID 11) ou admins podem salvar valores
    user_id = session['user_id']
    conn_check = get_db_connection()
    cursor_check = conn_check.cursor()
    cursor_check.execute('SELECT role FROM users WHERE id = ?', (user_id,))
    user_row = cursor_check.fetchone()
    return_db_connection(conn_check)
    
    is_admin = user_row and user_row[0] == 'admin'
    is_ronaldo = user_id == 11
    
    if not (is_admin or is_ronaldo):
        return jsonify({'success': False, 'message': 'Você não tem permissão para editar valores'}), 403
    
    try:
        data = request.get_json()
        valores = data.get('valores', [])
        
        if not valores:
            return jsonify({'success': False, 'message': 'Nenhum valor para salvar'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        saved_count = 0
        errors = []
        
        for item in valores:
            try:
                codigo = item.get('codigo', '').strip()
                setor = item.get('setor', '').strip()
                descricao = item.get('descricao', '').strip()
                valor_hora_str = item.get('valor_hora', '').strip()
                
                if not codigo or not descricao or not valor_hora_str:
                    continue
                
                # Extrair apenas números do valor
                valor_limpo = re.sub(r'[^\d,.]', '', valor_hora_str)
                valor_limpo = valor_limpo.replace(',', '.')
                valor_hora = float(valor_limpo)
                
                # Verificar se existe
                cursor.execute('SELECT id FROM valores_hora WHERE codigo = ?', (codigo,))
                existing = cursor.fetchone()
                
                if existing:
                    cursor.execute('''
                        UPDATE valores_hora 
                        SET setor = ?, descricao = ?, valor_hora = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE codigo = ?
                    ''', (setor, descricao, valor_hora, codigo))
                else:
                    cursor.execute('''
                        INSERT INTO valores_hora (codigo, setor, descricao, valor_hora)
                        VALUES (?, ?, ?, ?)
                    ''', (codigo, setor, descricao, valor_hora))
                
                saved_count += 1
                
            except Exception as e:
                errors.append(f"Erro no código {codigo}: {str(e)}")
        
        conn.commit()
        return_db_connection(conn)
        
        return jsonify({
            'success': True, 
            'message': f'{saved_count} valores salvos com sucesso',
            'saved_count': saved_count,
            'errors': errors if errors else None
        })
        
    except Exception as e:
        print(f"Erro ao salvar valores em lote: {e}")
        return jsonify({'success': False, 'message': f'Erro ao salvar: {str(e)}'}), 500


@app.route('/api/valores-hora/delete/<codigo>', methods=['DELETE'])
def api_valores_hora_delete(codigo):
    """Deleta um valor/hora"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    # CONTROLE DE ACESSO: Apenas Ronaldo (ID 11) ou admins podem deletar valores
    user_id = session['user_id']
    conn_check = get_db_connection()
    cursor_check = conn_check.cursor()
    cursor_check.execute('SELECT role FROM users WHERE id = ?', (user_id,))
    user_row = cursor_check.fetchone()
    return_db_connection(conn_check)
    
    is_admin = user_row and user_row[0] == 'admin'
    is_ronaldo = user_id == 11
    
    if not (is_admin or is_ronaldo):
        return jsonify({'success': False, 'message': 'Você não tem permissão para deletar valores'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM valores_hora WHERE codigo = ?', (codigo,))
        conn.commit()
        return_db_connection(conn)
        
        return jsonify({'success': True, 'message': 'Valor deletado com sucesso'})
        
    except Exception as e:
        print(f"Erro ao deletar valor/hora: {e}")
        return jsonify({'success': False, 'message': 'Erro ao deletar'}), 500


if __name__ == '__main__':
    # Inicializar banco de dados
    init_database()
    
    # Inicializar pool de conexões otimizado (reduzido para startup mais rápido)
    try:
        warm_pool(5)  # Reduzido de 150 para 5 conexões iniciais
    except Exception as _e:
        try:
            logger.warning(f"Falha ao pré-aquecer pool: {_e}")
        except Exception:
            pass
    
    # Iniciar backup automático (a cada 12 horas, adiado para não atrasar startup)
    try:
        # Atraso de 30 segundos antes do primeiro backup para não impactar a inicialização
        def delayed_backup_start():
            time.sleep(30)
        start_backup_scheduler(interval_seconds=43200)
        threading.Thread(target=delayed_backup_start, daemon=True).start()
    except Exception as e:
        try:
            logger.error(f"Falha ao iniciar backup automático: {e}")
        except Exception:
            print(f"Falha ao iniciar backup automático: {e}")
    
    # Iniciar monitor de performance em background
    performance_thread = threading.Thread(target=performance_monitor, daemon=True)
    performance_thread.start()
    
    # Iniciar configurações em background após inicialização completa
    def delayed_setup_compression():
        time.sleep(1.0)
        setup_compression()
    threading.Thread(target=delayed_setup_compression, daemon=True).start()
    
    # Obter IP e usar porta fixa para acesso em rede
    local_ip = get_local_ip()
    port = 5001
    
    print(" Iniciando Servidor do Formulário RNC")
    print("=" * 50)
    
    # Configuração SSL/HTTPS
    import os as _os
    ssl_cert_path = _os.path.join(_os.path.dirname(__file__), 'ssl_certs', 'cert.pem')
    ssl_key_path = _os.path.join(_os.path.dirname(__file__), 'ssl_certs', 'key.pem')
    
    use_https = _os.path.exists(ssl_cert_path) and _os.path.exists(ssl_key_path)
    
    if use_https:
        print(" HTTPS ATIVADO - Conexão Segura!")
        print(f" Login/Formulário: https://{local_ip}:{port}")
        print(f" Painel Admin: https://{local_ip}:5000")
        print("\n  AVISO: Certificado auto-assinado")
        print("   Aceite o aviso de segurança no navegador")
    else:
        print("  HTTP (não seguro) - Gere certificados SSL")
        print(f" Login/Formulário: http://{local_ip}:{port}")
        print(f" Painel Admin: http://{local_ip}:5000")
        print("\n Para ativar HTTPS:")
        print("   Execute: python gerar_certificado_ssl.py")
    print("=" * 50)
    print(" Usuário Admin criado automaticamente:")
    print("   Email: admin@ippel.com.br")
    print("   Senha: admin123")
    print("=" * 50)
    print(" Lixeira desativada:")
    print("   - Exclusão agora é definitiva (sem retenção)")
    print("=" * 50)
    print(" Otimizações de Performance:")
    print("   - Pool de conexões: 20 conexões")
    print("   - Cache de consultas: Ativo")
    print("   - Monitor de performance: Ativo")
    print("   - SocketIO threading: Ativo")
    print("   - SQLite WAL mode: Ativo")
    print("=" * 50)
    
    try:
        # Definir cookie de sessão para o domínio/IP acessado
        app.config['SESSION_COOKIE_DOMAIN'] = None  # IP
        app.config['SESSION_COOKIE_PATH'] = '/'
        
        # Configurar cookies seguros se HTTPS estiver ativo
        if use_https:
            app.config['SESSION_COOKIE_SECURE'] = True  # Apenas HTTPS
            app.config['SESSION_COOKIE_HTTPONLY'] = True
            app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
        
        # Executar servidor com SocketIO habilitado
        print(" Executando com SocketIO - funcionalidades de chat completas")
        if use_https:
            socketio.run(app, host='0.0.0.0', port=port, debug=False, 
                        allow_unsafe_werkzeug=True,
                        certfile=ssl_cert_path, keyfile=ssl_key_path)
        else:
            socketio.run(app, host='0.0.0.0', port=port, debug=False, 
                        allow_unsafe_werkzeug=True)
    except KeyboardInterrupt:
        print("\n Servidor do formulário encerrado")
    except Exception as e:
        print(f" Erro ao iniciar servidor: {e}")
